from pathlib import Path
import json
import tempfile
import unittest

from openpyxl import load_workbook

from cadtocae.workbook import (
    create_material_workbook,
    export_abaqus_json,
    read_component_rows_for_processing,
    read_raw_material_csv,
)


ROOT = Path(__file__).resolve().parents[1]


class WorkbookTest(unittest.TestCase):
    def test_create_material_workbook(self):
        raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "test_components.xlsx"
            create_material_workbook(
                raw_rows,
                support_type="单桩单立柱",
                angle="20",
                array_layout="2行7列竖向",
                output_path=output,
            )

            wb = load_workbook(output, data_only=False)
            resolved_rows, _headers = read_component_rows_for_processing(output)
        self.assertEqual(wb.sheetnames, ["原始材料表", "建模构件表"])

        raw_ws = wb["原始材料表"]
        self.assertEqual(raw_ws["G4"].value, "Q355 B")

        component_ws = wb["建模构件表"]
        headers = [cell.value for cell in component_ws[1]]
        part_name_col = headers.index("abaqus_part_name") + 1
        name_col = headers.index("构件名称") + 1
        self.assertEqual(part_name_col, name_col + 1)
        self.assertIn("'原始材料表'!C2", component_ws.cell(row=2, column=name_col).value)
        self.assertTrue(str(component_ws.cell(row=2, column=part_name_col).value).startswith("="))
        self.assertIn("INCLINED_BEAM", component_ws.cell(row=2, column=part_name_col).value)
        self.assertEqual(resolved_rows[0]["abaqus_part_name"], "P_SP_SC_ANG20_INCLINED_BEAM")
        self.assertNotIn("构件代码", headers)
        self.assertNotIn("是否重点分析", headers)
        self.assertNotIn("校核状态", headers)
        self.assertNotIn("Step02建模提示", headers)

        spec_col = headers.index("规格") + 1
        length_col = headers.index("长度_mm") + 1
        length_m_col = headers.index("长度_m") + 1
        material_col = headers.index("材料牌号") + 1
        model_policy_col = headers.index("建模方式") + 1
        self.assertIn("'原始材料表'!D2", component_ws.cell(row=2, column=spec_col).value)
        self.assertIn("/1000", component_ws.cell(row=2, column=length_m_col).value)
        self.assertIn("Q355B", component_ws.cell(row=2, column=material_col).value)
        self.assertEqual(component_ws.cell(row=1, column=spec_col).fill.fgColor.rgb, "FFC00000")
        self.assertIn("Step02", component_ws.cell(row=1, column=spec_col).comment.text)
        self.assertEqual(component_ws.cell(row=12, column=length_col).fill.fgColor.rgb, "FFFFC7CE")
        self.assertIn("长度", component_ws.cell(row=12, column=length_col).comment.text)
        self.assertEqual(component_ws.cell(row=15, column=model_policy_col).fill.fgColor.rgb, "FFFFC7CE")
        self.assertIn("人工模板", component_ws.cell(row=15, column=model_policy_col).comment.text)

    def test_export_approved_components_json(self):
        raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "test_components_approved.xlsx"
            create_material_workbook(
                raw_rows,
                support_type="单桩单立柱",
                angle="20",
                array_layout="2行7列竖向",
                output_path=workbook_path,
            )

            wb = load_workbook(workbook_path)
            ws = wb["建模构件表"]
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col).value = "校核状态"
            ws.cell(row=2, column=status_col).value = "已确认"
            wb.save(workbook_path)

            json_path = Path(tmp) / "test_abaqus_components.json"
            export_abaqus_json(workbook_path, json_path)
            payload = json.loads(json_path.read_text(encoding="utf-8"))
        self.assertEqual(len(payload["components"]), 1)
        self.assertEqual(payload["components"][0]["part_name"], "P_SP_SC_ANG20_INCLINED_BEAM")

    def test_export_complete_components_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "SP_SC_ANG20_components.xlsx"
            raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
            create_material_workbook(
                raw_rows,
                support_type="单桩单立柱",
                angle="20",
                array_layout="2行7列竖向",
                output_path=workbook_path,
            )
            json_path = Path(tmp) / "test_complete_components.json"
            export_abaqus_json(workbook_path, json_path, selection="complete")
            payload = json.loads(json_path.read_text(encoding="utf-8"))
        names = {component["part_name"] for component in payload["components"]}
        self.assertIn("P_SP_SC_ANG20_INCLINED_BEAM", names)
        self.assertIn("P_SP_SC_ANG20_HOOP", names)
        self.assertIn("P_SP_SC_ANG20_TIE_ROD", names)
        self.assertNotIn("P_SP_SC_ANG20_U_BOLT", names)
        self.assertNotIn("P_SP_SC_ANG20_PRESS_BLOCK", names)
        tie_rod = next(component for component in payload["components"] if component["part_name"] == "P_SP_SC_ANG20_TIE_ROD")
        self.assertEqual(tie_rod["model_policy"], "SOLID")
        self.assertEqual(tie_rod["component_code"], "TIE_ROD")
        inclined = next(component for component in payload["components"] if component["part_name"] == "P_SP_SC_ANG20_INCLINED_BEAM")
        self.assertEqual(inclined["length_m"], 3.948)
        self.assertEqual(inclined["thickness_m"], 0.002)
        self.assertEqual(inclined["section_params_m"]["h_m"], 0.075)
        self.assertEqual(inclined["material"]["density_kg_per_m3"], 7850.0)

    def test_export_uses_step01_section_columns_when_spec_needs_manual_parse(self):
        raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "SP_SC_ANG20_components.xlsx"
            create_material_workbook(
                raw_rows,
                support_type="单桩单立柱",
                angle="20",
                array_layout="2行7列竖向",
                output_path=workbook_path,
            )

            wb = load_workbook(workbook_path)
            ws = wb["建模构件表"]
            headers = [cell.value for cell in ws[1]]
            header_to_col = {header: index + 1 for index, header in enumerate(headers)}
            ws.cell(row=2, column=header_to_col["规格"]).value = "CUSTOM_C_CHANNEL"
            ws.cell(row=2, column=header_to_col["截面类型"]).value = "C型钢"
            ws.cell(row=2, column=header_to_col["截面参数"]).value = "高度_mm=75; 翼缘宽_mm=40; 卷边_mm=15; 厚度_mm=2"
            ws.cell(row=2, column=header_to_col["厚度_mm"]).value = 2
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col).value = "校核状态"
            ws.cell(row=2, column=status_col).value = "已确认"
            wb.save(workbook_path)

            json_path = Path(tmp) / "components.json"
            export_abaqus_json(workbook_path, json_path, selection="approved")
            payload = json.loads(json_path.read_text(encoding="utf-8"))

        component = payload["components"][0]
        self.assertEqual(component["section_kind"], "C_CHANNEL")
        self.assertEqual(component["section_params_m"]["h_m"], 0.075)
        self.assertEqual(component["section_params_m"]["b_m"], 0.04)
        self.assertEqual(component["section_params_m"]["lip_m"], 0.015)
        self.assertEqual(component["thickness_m"], 0.002)


if __name__ == "__main__":
    unittest.main()
