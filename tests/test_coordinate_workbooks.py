from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from cadtocae.coordinate_workbooks import create_coordinate_formula_workbooks


ROOT = Path(__file__).resolve().parents[1]
LAYOUT = ROOT / "examples" / "sp_sc_ang20_coordinate_layout.json"


class CoordinateWorkbookTest(unittest.TestCase):
    def test_coordinate_workbooks_use_project_prefix(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            full, simple = create_coordinate_formula_workbooks(
                LAYOUT,
                tmp_path / "DP_ANG26P5_coordinate_formula_full_fixed.xlsx",
                tmp_path / "DP_ANG26P5_coordinate_formula_simple_fixed.xlsx",
                "DP_ANG26P5",
            )
            self.assertTrue(full.exists())
            self.assertTrue(simple.exists())

            wb = load_workbook(full, data_only=False)
            ws = wb["构件轴线校核"]
            part_names = [ws.cell(row=row, column=2).value for row in range(4, 7)]
            instance_names = [ws.cell(row=row, column=3).value for row in range(4, 7)]
            self.assertIn("P_DP_ANG26P5_BRACE_FRONT", part_names)
            self.assertIn("I_DP_ANG26P5_BRACE_FRONT_01", instance_names)

    def test_coordinate_workbooks_include_purlin_axis_points_and_checks(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            full, simple = create_coordinate_formula_workbooks(
                LAYOUT,
                tmp_path / "SP_SC_ANG20_coordinate_formula_full_fixed.xlsx",
                tmp_path / "SP_SC_ANG20_coordinate_formula_simple_fixed.xlsx",
                "SP_SC_ANG20",
            )

            simple_ws = load_workbook(simple, data_only=False)["坐标计算总表"]
            simple_first_col = [simple_ws.cell(row=row, column=1).value for row in range(1, simple_ws.max_row + 1)]
            for name in ["HF_mm", "HS_mm", "HP_mm", "HQ_mm", "HR_mm", "pv_axis_angle_tolerance_deg", "H", "S", "P", "Q", "R"]:
                self.assertIn(name, simple_first_col)
            for check_name in ["SPQR_COLLINEAR", "SPQR_ANGLE"]:
                self.assertIn(check_name, simple_first_col)

            h_row = simple_first_col.index("H") + 1
            s_row = simple_first_col.index("S") + 1
            self.assertIn("SIN(RADIANS", simple_ws.cell(row=h_row, column=2).value)
            self.assertIn("-($C$", simple_ws.cell(row=s_row, column=2).value)

            full_wb = load_workbook(full, data_only=False)
            point_names = [full_wb["控制点坐标"].cell(row=row, column=1).value for row in range(1, full_wb["控制点坐标"].max_row + 1)]
            check_names = [full_wb["长度与角度校核"].cell(row=row, column=1).value for row in range(1, full_wb["长度与角度校核"].max_row + 1)]
            annotation_names = [full_wb["图纸标注记录"].cell(row=row, column=1).value for row in range(1, full_wb["图纸标注记录"].max_row + 1)]
            for name in ["H", "S", "P", "Q", "R"]:
                self.assertIn(name, point_names)
                self.assertIn(name, annotation_names)
            self.assertIn("SPQR_COLLINEAR", check_names)
            self.assertIn("SPQR_ANGLE", check_names)


if __name__ == "__main__":
    unittest.main()
