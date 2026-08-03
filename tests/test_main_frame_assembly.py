from __future__ import annotations

import json
import math
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from cadtocae.assembly_script import generate_assembly_scripts_from_workbook
from cadtocae.coordinate_workbooks import create_coordinate_formula_workbooks
from cadtocae.main_frame_assembly import (
    PASSED,
    PURLIN_GROUP_Y_OFFSET_M,
    PURLIN_SHORT_LENGTH_M,
    _section_reference_xy,
    build_payload,
    export_main_frame_assembly,
    transform_rotation_sequence,
    rotate_z,
    rotate_y,
)
from cadtocae.workbook import create_material_workbook, export_abaqus_json, read_raw_material_csv


ROOT = Path(__file__).resolve().parents[1]
LAYOUT = ROOT / "examples" / "sp_sc_ang20_coordinate_layout.json"


def build_sample_inputs(tmp: str | Path) -> tuple[Path, Path]:
    tmp_path = Path(tmp)
    raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
    components_xlsx = tmp_path / "SP_SC_ANG20_components.xlsx"
    create_material_workbook(raw_rows, "单桩单立柱", "20", "2行7列竖向", components_xlsx)
    components_json = tmp_path / "components.json"
    export_abaqus_json(components_xlsx, components_json, selection="complete")
    _full, simple = create_coordinate_formula_workbooks(
        LAYOUT,
        tmp_path / "SP_SC_ANG20_coordinate_formula_full_fixed.xlsx",
        tmp_path / "SP_SC_ANG20_coordinate_formula_simple_fixed.xlsx",
        "SP_SC_ANG20",
    )
    return simple, components_json


def fill_purlin_axis_inputs(excel: Path) -> None:
    values = {
        "HF_mm": 115,
        "HS_mm": 1849,
        "HP_mm": 449,
        "HQ_mm": 449,
        "HR_mm": 1849,
    }
    wb = load_workbook(excel)
    try:
        for ws in wb.worksheets:
            for row in range(1, ws.max_row + 1):
                name = ws.cell(row=row, column=1).value
                if name in values:
                    ws.cell(row=row, column=3).value = values[name]
        wb.save(excel)
    finally:
        wb.close()


class MainFrameAssemblyTest(unittest.TestCase):
    def test_section_reference_for_pipe_c_channel_and_angle(self):
        pipe_ref = _section_reference_xy({"section_kind": "PIPE", "section_params_m": {"od_m": 0.14, "t_m": 0.0035}})
        self.assertEqual(pipe_ref["x_m"], 0.0)
        self.assertEqual(pipe_ref["y_m"], 0.0)
        self.assertEqual(pipe_ref["rule"], "SECTION_ORIGIN")

        c_ref = _section_reference_xy(
            {"section_kind": "C_CHANNEL", "section_params_m": {"h_m": 0.075, "b_m": 0.04, "lip_m": 0.015, "t_m": 0.002}}
        )
        self.assertAlmostEqual(c_ref["x_m"], 0.0)
        self.assertAlmostEqual(c_ref["y_m"], 0.0375)
        self.assertEqual(c_ref["rule"], "C_CHANNEL_WEB_MIDPOINT")
        self.assertEqual(c_ref["open_side_target_global"], "-Y")

        angle_ref = _section_reference_xy(
            {"section_kind": "ANGLE", "section_params_m": {"leg_a_m": 0.09, "leg_b_m": 0.056, "t_m": 0.005}}
        )
        area = 0.09 * 0.005 + 0.005 * 0.056 - 0.005 * 0.005
        expected_x = (0.09 * 0.005 * 0.045 + 0.005 * 0.056 * 0.0025 - 0.005 * 0.005 * 0.0025) / area
        expected_y = (0.09 * 0.005 * 0.0025 + 0.005 * 0.056 * 0.028 - 0.005 * 0.005 * 0.0025) / area
        self.assertAlmostEqual(angle_ref["x_m"], expected_x)
        self.assertAlmostEqual(angle_ref["y_m"], expected_y)
        self.assertEqual(angle_ref["rule"], "ANGLE_SOLID_SECTION_CENTROID")

    def test_excel_inputs_recompute_g_local_control_points(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel, components = build_sample_inputs(tmp)
            payload = build_payload(excel, components)
        inputs = payload["inputs"]
        points = payload["points"]

        self.assertAlmostEqual(inputs["theta_deg"], 18.0)
        self.assertAlmostEqual(inputs["GC_m"], 0.439)
        self.assertAlmostEqual(inputs["GF_m"], 2.119)
        self.assertAlmostEqual(inputs["GE_m"], 3.739)

        self.assertAlmostEqual(points["F"]["x_m"], 0.0)
        self.assertAlmostEqual(points["F"]["z_m"], 3.1)
        self.assertAlmostEqual(points["C"]["x_m"], -1.597774947375858)
        self.assertAlmostEqual(points["C"]["z_m"], 2.5808514494500883)
        self.assertAlmostEqual(points["E"]["x_m"], 1.5407115563981484)
        self.assertAlmostEqual(points["E"]["z_m"], 3.600607530887415)
        self.assertAlmostEqual(points["G_global"]["x_m"], -2.0152887580294307)
        self.assertAlmostEqual(points["G_global"]["z_m"], 2.4451929889194863)

    def test_beam_anchor_uses_section_centroid_and_y_rotation(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel, components = build_sample_inputs(tmp)
            payload = build_payload(excel, components)
        beam = payload["beam_anchor"]

        self.assertAlmostEqual(beam["section_reference"]["x_m"], 0.0)
        self.assertAlmostEqual(beam["section_reference"]["y_m"], 0.0375)
        self.assertEqual(beam["section_reference"]["rule"], "C_CHANNEL_WEB_MIDPOINT")
        self.assertEqual(beam["axis_local_point"], [0.0, 0.0, 2.119])
        self.assertAlmostEqual(beam["local_point"][0], beam["section_reference"]["x_m"])
        self.assertAlmostEqual(beam["local_point"][1], beam["section_reference"]["y_m"])
        self.assertAlmostEqual(beam["local_point"][2], 2.119)
        self.assertAlmostEqual(beam["rotate_y_deg"], 72.0)
        self.assertAlmostEqual(beam["roll_about_axis_deg"], 90.0)

        rotated_f = rotate_y(beam["local_point"], beam["rotate_y_deg"])
        transformed_f = [rotated_f[index] + beam["translation"][index] for index in range(3)]
        self.assertAlmostEqual(transformed_f[0], payload["points"]["F"]["x_m"])
        self.assertAlmostEqual(transformed_f[1], payload["points"]["F"]["y_m"])
        self.assertAlmostEqual(transformed_f[2], payload["points"]["F"]["z_m"])
        open_side = rotate_y(rotate_z([1.0, 0.0, 0.0], beam["roll_about_axis_deg"]), beam["rotate_y_deg"])
        self.assertAlmostEqual(open_side[0], 0.0, places=7)
        self.assertAlmostEqual(open_side[1], 1.0, places=7)
        self.assertAlmostEqual(open_side[2], 0.0, places=7)

        local_z_after_rotation = rotate_y([0.0, 0.0, 1.0], beam["rotate_y_deg"])
        self.assertAlmostEqual(local_z_after_rotation[0], math.cos(math.radians(18.0)))
        self.assertAlmostEqual(local_z_after_rotation[2], math.sin(math.radians(18.0)))

    def test_excel_checks_are_exported_for_current_tolerance(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel, components = build_sample_inputs(tmp)
            payload = build_payload(excel, components)
        checks = payload["checks"]
        self.assertEqual(checks["GC_GF_GE_ORDER"]["passed"], PASSED)
        self.assertEqual(checks["CE_ANGLE"]["passed"], PASSED)
        self.assertEqual(checks["BC"]["passed"], "不通过")
        self.assertEqual(checks["DE"]["passed"], "不通过")
        self.assertAlmostEqual(checks["BC"]["error"], 0.0012409188706739016)
        self.assertAlmostEqual(checks["DE"]["error"], -0.013053394071534719)

    def test_purlin_axis_points_checks_and_members_are_exported(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel, components = build_sample_inputs(tmp)
            fill_purlin_axis_inputs(excel)
            payload = build_payload(excel, components)

        points = payload["points"]
        checks = payload["checks"]
        self.assertEqual(checks["SPQR_COLLINEAR"]["passed"], PASSED)
        self.assertEqual(checks["SPQR_ANGLE"]["passed"], PASSED)
        self.assertAlmostEqual(checks["SPQR_ANGLE"]["error"], 0.0)
        for name in ["H", "S", "P", "Q", "R"]:
            self.assertIn(name, points)

        members_by_name = {member["name"]: member for member in payload["members"]}
        purlin_names = ["PURLIN_%s" % name for name in ["S", "P", "Q", "R"]]
        support_names = ["PURLIN_SUPPORT_%s" % name for name in ["S", "P", "Q", "R"]]
        for name in purlin_names + support_names:
            self.assertIn(name, members_by_name)
        self.assertEqual(len({members_by_name[name]["instance_name"] for name in purlin_names + support_names}), 8)
        self.assertEqual(payload["purlin_axis"]["enabled"], True)
        self.assertEqual(payload["purlin_axis"]["purlin_derived_part_name"], "P_SP_SC_ANG20_PURLIN_50MM")

        theta = math.radians(payload["inputs"]["theta_deg"])
        expected_u = [math.cos(theta), 0.0, math.sin(theta)]
        expected_n = [-math.sin(theta), 0.0, math.cos(theta)]
        purlin_s = members_by_name["PURLIN_S"]
        support_s = members_by_name["PURLIN_SUPPORT_S"]
        self.assertEqual(purlin_s["part_name"], "P_SP_SC_ANG20_PURLIN_50MM")
        self.assertEqual(purlin_s["source_part_name"], "P_SP_SC_ANG20_PURLIN")
        self.assertAlmostEqual(purlin_s["part_length_m"], PURLIN_SHORT_LENGTH_M)
        self.assertEqual(purlin_s["local_anchor"], [0.025, 0.115, 0.025])

        rotated_anchor = transform_rotation_sequence(purlin_s["local_anchor"], purlin_s["rotation_sequence"])
        transformed_anchor = [rotated_anchor[index] + purlin_s["translation"][index] for index in range(3)]
        expected_purlin_anchor = list(points["S"]["coords"])
        expected_purlin_anchor[1] += PURLIN_GROUP_Y_OFFSET_M
        for index, expected in enumerate(expected_purlin_anchor):
            self.assertAlmostEqual(transformed_anchor[index], expected)
        self.assertAlmostEqual(payload["purlin_axis"]["group_y_offset_m"], PURLIN_GROUP_Y_OFFSET_M)

        flange_axis = transform_rotation_sequence([1.0, 0.0, 0.0], purlin_s["rotation_sequence"])
        web_axis = transform_rotation_sequence([0.0, 1.0, 0.0], purlin_s["rotation_sequence"])
        length_axis = transform_rotation_sequence([0.0, 0.0, 1.0], purlin_s["rotation_sequence"])
        for index in range(3):
            self.assertAlmostEqual(flange_axis[index], expected_u[index])
            self.assertAlmostEqual(web_axis[index], expected_n[index])
        self.assertAlmostEqual(abs(length_axis[1]), 1.0)
        self.assertAlmostEqual(length_axis[0], 0.0, places=7)
        self.assertAlmostEqual(length_axis[2], 0.0, places=7)

        support_long_leg = transform_rotation_sequence([1.0, 0.0, 0.0], support_s["rotation_sequence"])
        support_short_leg = transform_rotation_sequence([0.0, 1.0, 0.0], support_s["rotation_sequence"])
        for index in range(3):
            self.assertAlmostEqual(support_long_leg[index], expected_n[index])
            self.assertAlmostEqual(support_short_leg[index], -expected_u[index])
        self.assertEqual(support_s["part_name"], "P_SP_SC_ANG20_PURLIN_SUPPORT")
        self.assertEqual(support_s["local_anchor"], [0.0, 0.0, 0.025])

        expected_web_bottom = [
            points["S"]["coords"][index] - 0.025 * expected_u[index] - 0.115 * expected_n[index]
            for index in range(3)
        ]
        expected_web_bottom[1] += PURLIN_GROUP_Y_OFFSET_M
        support_rotated_anchor = transform_rotation_sequence(support_s["local_anchor"], support_s["rotation_sequence"])
        support_transformed_anchor = [support_rotated_anchor[index] + support_s["translation"][index] for index in range(3)]
        for index, expected in enumerate(expected_web_bottom):
            self.assertAlmostEqual(support_transformed_anchor[index], expected)
        self.assertEqual(payload["member_checks"]["PURLIN_SUPPORT_S_PLACEMENT"]["anchor"], "S_WEB_BOTTOM")

    def test_export_generates_single_embedded_assembly_script(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            excel, components = build_sample_inputs(tmp)
            out_json = out_dir / "SP_SC_ANG20_main_frame_assembly_inputs.json"
            json_path, scripts, payload = export_main_frame_assembly(
                excel,
                components,
                out_json,
                out_dir,
                project_code="SP_SC_ANG20",
            )

            self.assertTrue(json_path.exists())
            data = json.loads(json_path.read_text(encoding="utf-8"))
            self.assertEqual(data["beam_anchor"]["section_sets"]["F"], "SET_BEAM_SEC_F")
            self.assertEqual(data["beam_anchor"]["section_reference"]["rule"], "C_CHANNEL_WEB_MIDPOINT")
            self.assertIn("section_reference", data["members"][2])
            self.assertIn("roll_about_axis_deg", data["members"][2])
            self.assertEqual(data["members"][2]["open_side_global"], [0.0, 1.0, 0.0])
            self.assertEqual(data["members"][2]["roll_about_axis_deg"], 90.0)
            self.assertEqual(data["members"][3]["roll_about_axis_deg"], -90.0)
            self.assertEqual(data["members"][4]["roll_about_axis_deg"], -90.0)
            self.assertNotIn("reference_point_set", payload["beam_anchor"])
            self.assertNotIn("reference_point_set", data["beam_anchor"])

            script_names = {path.name for path in scripts}
            self.assertEqual(
                script_names,
                {
                    "SP_SC_ANG20_assembly_frame.py",
                },
            )
            assembly_script = (out_dir / "SP_SC_ANG20_assembly_frame.py").read_text(encoding="utf-8")
            self.assertIn('PHASE = "full_main_frame"', assembly_script)
            self.assertIn("ASSEMBLY_DATA = json.loads", assembly_script)
            self.assertIn("SET_BEAM_SEC_F", assembly_script)
            self.assertIn("roll_about_axis_deg", assembly_script)
            self.assertIn("_member_axis_direction", assembly_script)
            self.assertNotIn("ReferencePoint", assembly_script)
            self.assertNotIn("RP_GLOBAL", assembly_script)
            self.assertNotIn("RP_BEAM_F_LOCAL", assembly_script)
            self.assertIn("does not match project prefix", assembly_script)
            self.assertIn("Project model %s not found", assembly_script)
            self.assertNotIn("ASSEMBLY_JSON_PATH", assembly_script)
            self.assertNotIn("_resolve_script_dir", assembly_script)

    def test_build_payload_accepts_single_column_component_code(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            excel, components = build_sample_inputs(tmp)
            payload = json.loads(components.read_text(encoding="utf-8"))
            source_column = next(row for row in payload["components"] if row.get("component_code") == "COLUMN_UP")
            single_column = dict(source_column)
            single_column["component_code"] = "COLUMN"
            single_column["component_name"] = "立柱"
            single_column["part_name"] = "P_SP_SC_ANG20_COLUMN"
            payload["components"] = [
                row for row in payload["components"] if row.get("component_code") not in {"COLUMN_UP", "COLUMN_DOWN"}
            ]
            payload["components"].append(single_column)
            single_components = tmp_path / "single_column_components.json"
            single_components.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

            assembly_payload = build_payload(excel, single_components)

            member_names = {member["name"] for member in assembly_payload["members"]}
            self.assertIn("COLUMN", member_names)
            self.assertNotIn("COLUMN_UP", member_names)
            self.assertNotIn("COLUMN_DOWN", member_names)
            self.assertIn("COLUMN_PLACEMENT", assembly_payload["member_checks"])

    def test_build_payload_accepts_column_up_without_column_down(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            excel, components = build_sample_inputs(tmp)
            payload = json.loads(components.read_text(encoding="utf-8"))
            payload["components"] = [
                row for row in payload["components"] if row.get("component_code") != "COLUMN_DOWN"
            ]
            column_up_only = tmp_path / "column_up_only_components.json"
            column_up_only.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

            assembly_payload = build_payload(excel, column_up_only)

            member_names = {member["name"] for member in assembly_payload["members"]}
            self.assertIn("COLUMN_UP", member_names)
            self.assertNotIn("COLUMN_DOWN", member_names)
            self.assertIn("COLUMN_UP_PLACEMENT", assembly_payload["member_checks"])
            self.assertIn("Only COLUMN_UP is available", assembly_payload["member_checks"]["COLUMN_UP_PLACEMENT"]["note"])

    def test_step04_auto_locates_flat_step02_part_script_components(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            excel, components = build_sample_inputs(tmp)
            part_script = tmp_path / "SP_SC_ANG20_create_parts_in_cae.py"
            part_script.write_text(
                "COMPONENTS_JSON = r'''%s'''\n" % components.read_text(encoding="utf-8"),
                encoding="utf-8",
            )

            result = generate_assembly_scripts_from_workbook(
                excel,
                tmp_path / "assembly_out",
                overwrite=True,
            )

            self.assertIn(result.status, {"ok", "needs_review"})
            self.assertIsNone(result.assembly_json_path)
            self.assertEqual(Path(result.copied_components_json_path or "").name, "SP_SC_ANG20_create_parts_in_cae.py")
            self.assertEqual({"SP_SC_ANG20_assembly_frame.py"}, {Path(path).name for path in result.script_paths})
            expected_report = tmp_path / "assembly_out" / "过程文件" / "调试文件" / "SP_SC_ANG20_step04_assembly_script_report.json"
            self.assertEqual(Path(result.report_path), expected_report)
            self.assertTrue(expected_report.exists())
            self.assertFalse((tmp_path / "assembly_out" / "SP_SC_ANG20_step04_assembly_script_report.json").exists())
            self.assertFalse((tmp_path / "assembly_out" / "SP_SC_ANG20_assembly_inputs.json").exists())

    def test_step04_rejects_model_name_that_does_not_match_prefix(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            excel, components = build_sample_inputs(tmp)
            part_script = tmp_path / "SP_SC_ANG20_create_parts_in_cae.py"
            part_script.write_text(
                "COMPONENTS_JSON = r'''%s'''\n" % components.read_text(encoding="utf-8"),
                encoding="utf-8",
            )

            result = generate_assembly_scripts_from_workbook(
                excel,
                tmp_path,
                model_name="Model-1",
                overwrite=True,
            )

            self.assertEqual(result.status, "failed")
            self.assertIn("must match project prefix", result.messages[0])


if __name__ == "__main__":
    unittest.main()
