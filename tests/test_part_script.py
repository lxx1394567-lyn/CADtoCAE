from __future__ import annotations

import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from cadtocae.main_frame_assembly import read_components_payload
from cadtocae.part_script import batch_generate_part_scripts, infer_project_prefix_from_workbook
from cadtocae.workbook import create_material_workbook, read_raw_material_csv


ROOT = Path(__file__).resolve().parents[1]


class PartScriptGenerationTest(unittest.TestCase):
    def test_batch_generates_part_script_from_step01_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "SP_SC_ANG20_components.xlsx"
            raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
            create_material_workbook(raw_rows, "单桩单立柱", "20", "2行7列竖向", workbook)

            outputs = batch_generate_part_scripts([workbook], tmp_path / "out", selection="complete")

            self.assertEqual(len(outputs), 1)
            result = outputs[0]
            self.assertEqual(result.status, "ok")
            self.assertEqual(result.project_prefix, "SP_SC_ANG20")
            self.assertEqual(Path(result.project_dir), tmp_path / "out")
            self.assertIsNone(result.components_json_path)
            expected_report = tmp_path / "out" / "过程文件" / "调试文件" / "SP_SC_ANG20_step02_part_script_report.json"
            self.assertEqual(Path(result.report_path), expected_report)
            self.assertTrue(expected_report.exists())
            self.assertEqual(Path(result.part_script_path or ""), tmp_path / "out" / "SP_SC_ANG20_create_parts_in_cae.py")
            self.assertTrue(Path(result.part_script_path or "").exists())
            self.assertFalse((tmp_path / "out" / "SP_SC_ANG20_components.json").exists())
            self.assertFalse((tmp_path / "out" / "SP_SC_ANG20_step02_part_script_report.json").exists())
            self.assertGreater(result.exported_count, 0)

            payload = read_components_payload(result.part_script_path or "")
            part_names = {component["part_name"] for component in payload["components"]}
            self.assertIn("P_SP_SC_ANG20_INCLINED_BEAM", part_names)

            script = Path(result.part_script_path or "").read_text(encoding="utf-8")
            self.assertIn('MODEL_NAME = "SP_SC_ANG20"', script)
            self.assertIsNone(result.cae_save_path)
            self.assertNotIn("CADtoCAE_PARTS.cae", script)
            self.assertNotIn("openMdb", script)
            self.assertNotIn("SAVE_AS_PATH", script)
            self.assertNotIn("saveAs", script)
            self.assertIn("COMPONENTS_JSON", script)
            self.assertIn("P_SP_SC_ANG20_INCLINED_BEAM", script)
            self.assertNotIn("光伏抗风", script)
            self.assertNotIn("create_parts_in_cae_mkg", script)

    def test_infer_project_prefix_from_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "components.xlsx"
            raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
            create_material_workbook(raw_rows, "双桩", "26.5", "2行7列竖向", workbook)

            self.assertEqual(infer_project_prefix_from_workbook(workbook), "DP_ANG26P5")

    def test_filename_prefix_overrides_stale_angle_inside_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "SP_SC_ANG18_components.xlsx"
            raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
            create_material_workbook(raw_rows, "单桩单立柱", "20", "2行7列竖向", workbook)

            outputs = batch_generate_part_scripts([workbook], tmp_path / "out", selection="complete")

            self.assertEqual(outputs[0].project_prefix, "SP_SC_ANG18")
            self.assertTrue(str(outputs[0].part_script_path).endswith("SP_SC_ANG18_create_parts_in_cae.py"))
            script = Path(outputs[0].part_script_path or "").read_text(encoding="utf-8")
            self.assertIn('MODEL_NAME = "SP_SC_ANG18"', script)
            self.assertNotIn('MODEL_NAME = "SP_SC_ANG18_PARTS"', script)
            payload = read_components_payload(outputs[0].part_script_path or "")
            part_names = {component["part_name"] for component in payload["components"]}
            self.assertIn("P_SP_SC_ANG18_INCLINED_BEAM", part_names)
            self.assertNotIn("P_SP_SC_ANG20_INCLINED_BEAM", part_names)

    def test_single_step02_cli_uses_filename_prefix_for_generated_part_names(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "SP_SC_ANG18_components.xlsx"
            raw_rows = read_raw_material_csv(ROOT / "examples" / "single_pile_single_column_2x7_raw_materials.csv")
            create_material_workbook(raw_rows, "单桩单立柱", "20", "2行7列竖向", workbook)

            result = subprocess.run(
                [
                    sys.executable,
                    str(ROOT / "scripts" / "step02_generate_part_script.py"),
                    "--xlsx",
                    str(workbook),
                    "--outputs-root",
                    str(tmp_path / "outputs"),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            scripts = list((tmp_path / "outputs").glob("SP_SC_ANG18_runs/*/abaqus_scripts/SP_SC_ANG18_create_parts_in_cae.py"))
            self.assertEqual(len(scripts), 1)
            payload = read_components_payload(scripts[0])
            part_names = {component["part_name"] for component in payload["components"]}
            self.assertIn("P_SP_SC_ANG18_INCLINED_BEAM", part_names)
            self.assertNotIn("P_SP_SC_ANG20_INCLINED_BEAM", part_names)

    def test_step02_reads_raw_material_updates_without_excel_recalc(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "SP_SC_ANG20_components.xlsx"
            raw_rows = [
                {
                    "类别": "支架",
                    "序号": "1",
                    "名称": "斜梁",
                    "规格": "C75×40×15×2.0",
                    "长度_mm": "3948",
                    "数量": "2",
                    "备注": "Q355 B",
                    "来源页码": "1",
                    "识别置信度": "1.00",
                }
            ]
            create_material_workbook(raw_rows, "单桩单立柱", "20", "2行7列竖向", workbook)

            wb = load_workbook(workbook)
            raw_ws = wb["原始材料表"]
            raw_ws["C2"] = "上立柱"
            raw_ws["G2"] = "Q235B"
            wb.save(workbook)

            outputs = batch_generate_part_scripts([workbook], tmp_path / "out", selection="complete")

            self.assertEqual(outputs[0].status, "ok")
            payload = read_components_payload(outputs[0].part_script_path or "")
            self.assertEqual(len(payload["components"]), 1)
            component = payload["components"][0]
            self.assertEqual(component["part_name"], "P_SP_SC_ANG20_COLUMN_UP")
            self.assertEqual(component["component_code"], "COLUMN_UP")
            self.assertEqual(component["material"]["material_grade"], "Q235 B")

    def test_step02_prefers_manual_component_sheet_overrides(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "SP_SC_ANG20_components.xlsx"
            raw_rows = [
                {
                    "类别": "支架",
                    "序号": "1",
                    "名称": "斜梁",
                    "规格": "C75×40×15×2.0",
                    "长度_mm": "3948",
                    "数量": "2",
                    "备注": "Q355 B",
                    "来源页码": "1",
                    "识别置信度": "1.00",
                }
            ]
            create_material_workbook(raw_rows, "单桩单立柱", "20", "2行7列竖向", workbook)

            wb = load_workbook(workbook)
            ws = wb["建模构件表"]
            headers = [cell.value for cell in ws[1]]
            part_col = headers.index("abaqus_part_name") + 1
            ws.cell(row=2, column=part_col).value = "P_SP_SC_ANG20_MANUAL_BEAM"
            wb.save(workbook)

            outputs = batch_generate_part_scripts([workbook], tmp_path / "out", selection="complete")

            self.assertEqual(outputs[0].status, "ok")
            payload = read_components_payload(outputs[0].part_script_path or "")
            self.assertEqual(payload["components"][0]["part_name"], "P_SP_SC_ANG20_MANUAL_BEAM")


if __name__ == "__main__":
    unittest.main()
