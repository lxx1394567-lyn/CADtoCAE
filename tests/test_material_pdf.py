from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from cadtocae.material_pdf import (
    OcrToken,
    PdfMaterialResult,
    batch_extract_material_workbooks,
    detect_project_info,
    rows_from_positioned_words,
    rows_from_tables,
    rows_from_text,
)
from cadtocae.workbook import read_component_rows_for_processing


class MaterialPdfExtractionTest(unittest.TestCase):
    def test_rows_from_headered_pdf_table(self):
        table = [
            ["序号", "名称", "规格", "长度", "数量", "备注"],
            ["1", "斜梁", "C75×40×15×2.0", "3948", "2", "Q355 B"],
            ["2", "上立柱", "Φ127×2.5", "2050", "2", "Q355 B"],
        ]

        rows = rows_from_tables([(1, table)])

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["名称"], "斜梁")
        self.assertEqual(rows[0]["规格"], "C75×40×15×2.0")
        self.assertEqual(rows[0]["来源页码"], "1")

    def test_rows_from_ocr_like_text(self):
        text = "\n".join(
            [
                "材料表",
                "1 斜梁 C75×40×15×2.0 3948 2 Q355 B",
                "2 上立柱 Φ127×2.5 2050 2 Q355 B",
            ]
        )

        rows = rows_from_text([(1, text)])

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1]["名称"], "上立柱")
        self.assertEqual(rows[1]["数量"], "2")

    def test_rows_from_positioned_words_uses_table_grid(self):
        verticals = [0, 100, 200, 400, 650, 820, 950, 1150]
        horizontals = [0, 60, 120, 180]
        tokens = [
            OcrToken("1", 130, 80, 10, 20, 95),
            OcrToken("斜梁", 260, 80, 40, 20, 92),
            OcrToken("C75×40×15×2.0", 470, 80, 120, 20, 90),
            OcrToken("3948", 720, 80, 50, 20, 96),
            OcrToken("2", 880, 80, 10, 20, 97),
            OcrToken("Q355", 1010, 80, 45, 20, 94),
            OcrToken("B", 1060, 80, 10, 20, 94),
            OcrToken("2", 130, 140, 10, 20, 95),
            OcrToken("上立柱", 250, 140, 60, 20, 92),
            OcrToken("Φ127×2.5", 480, 140, 90, 20, 91),
            OcrToken("2050", 720, 140, 50, 20, 96),
            OcrToken("2", 880, 140, 10, 20, 97),
        ]

        rows = rows_from_positioned_words(tokens, verticals, horizontals)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["名称"], "斜梁")
        self.assertEqual(rows[0]["规格"], "C75×40×15×2.0")
        self.assertEqual(rows[0]["长度_mm"], "3948")
        self.assertEqual(rows[0]["备注"], "Q355 B")
        self.assertEqual(rows[1]["名称"], "上立柱")

    def test_detect_project_info_uses_filename_and_angle_text(self):
        support_type, angle, prefix, messages = detect_project_info(
            "项目A-双桩.pdf",
            ["光伏板倾角 26.5°"],
            fallback_support_type="单桩单立柱",
            fallback_angle="20",
        )

        self.assertEqual(support_type, "双桩")
        self.assertEqual(angle, "26.5")
        self.assertEqual(prefix, "DP_ANG26P5")
        self.assertTrue(any("识别支架类型" in message for message in messages))

    def test_batch_writes_prefixed_component_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            pdf = root / "项目A-单桩双立柱-20度.pdf"
            pdf.write_bytes(b"%PDF-1.4\n")
            result = PdfMaterialResult(
                pdf_path=str(pdf),
                project_prefix="SP_DC_ANG20",
                support_type="单桩双立柱",
                angle="20",
                rows=[
                    {
                        "类别": "支架",
                        "序号": "1",
                        "名称": "斜梁",
                        "规格": "C75×40×15×2.0",
                        "长度_mm": "3948",
                        "数量": "2",
                        "备注": "Q355 B",
                        "来源页码": "1",
                        "识别置信度": "0.90",
                    }
                ],
                status="ok",
                messages=["识别到材料表行数: 1"],
                used_pages=[1],
                extraction_method="pdf_table",
            )

            with patch("cadtocae.material_pdf.extract_material_table_from_document", return_value=result):
                outputs = batch_extract_material_workbooks([pdf], root / "out")

            self.assertEqual(len(outputs), 1)
            workbook_path = Path(outputs[0].workbook_path or "")
            self.assertEqual(workbook_path.name, "SP_DC_ANG20_components.xlsx")
            self.assertTrue(workbook_path.exists())

            wb = load_workbook(workbook_path, data_only=False)
            ws = wb["建模构件表"]
            headers = [cell.value for cell in ws[1]]
            part_col = headers.index("abaqus_part_name") + 1
            self.assertTrue(str(ws.cell(2, part_col).value).startswith("="))
            rows, _headers = read_component_rows_for_processing(workbook_path)
            self.assertEqual(rows[0]["abaqus_part_name"], "P_SP_DC_ANG20_INCLINED_BEAM")

    def test_batch_failure_writes_template_without_component_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            pdf = root / "项目B-双桩-26.5度.pdf"
            pdf.write_bytes(b"%PDF-1.4\n")
            result = PdfMaterialResult(
                pdf_path=str(pdf),
                project_prefix="DP_ANG26P5",
                support_type="双桩",
                angle="26.5",
                rows=[],
                status="needs_review",
                messages=["未识别到可用材料表"],
                used_pages=[],
                extraction_method="none",
            )

            with patch("cadtocae.material_pdf.extract_material_table_from_document", return_value=result):
                outputs = batch_extract_material_workbooks([pdf], root / "out")

            self.assertIsNone(outputs[0].workbook_path)
            self.assertTrue(Path(outputs[0].manual_template_path or "").exists())
            self.assertFalse(list((root / "out").glob("**/*_components.xlsx")))
            self.assertIsNone(outputs[0].report_path)
            self.assertFalse(list((root / "out").glob("**/step01_material_recognition_report.json")))


if __name__ == "__main__":
    unittest.main()
