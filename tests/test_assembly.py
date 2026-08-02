from __future__ import annotations

from pathlib import Path
import copy
import json
import math
import tempfile
import unittest

from openpyxl import load_workbook
from PIL import Image

from cadtocae.assembly import (
    MANUAL_CHECK,
    PASSED,
    beam_local_sections,
    brace_length_checks,
    control_length_checks,
    create_coordinate_workbook,
    distance_3d,
    export_assembly_inputs,
    length_and_angle_checks,
    load_layout,
    member_checks,
    ready_members,
    solve_control_points,
    world_to_pixel,
)


ROOT = Path(__file__).resolve().parents[1]
LAYOUT = ROOT / "examples" / "sp_sc_ang20_coordinate_layout.json"


def set_input_status(layout: dict, names: set[str], status: str) -> None:
    for row in layout["dimension_inputs"]:
        if row["name"] in names:
            row["status"] = status


class AssemblyCoordinateTest(unittest.TestCase):
    def test_distance_3d_uses_meter_coordinates(self):
        self.assertAlmostEqual(
            distance_3d({"x_m": 0, "y_m": 0, "z_m": 0}, {"x_m": 3, "y_m": 4, "z_m": 12}),
            13.0,
        )

    def test_member_checks_keep_unconfirmed_members_out_of_ready_inputs(self):
        layout = load_layout(LAYOUT)
        checks = member_checks(layout)
        self.assertGreaterEqual(len(checks), 3)
        self.assertTrue(all(check["passed"] == MANUAL_CHECK for check in checks))
        self.assertEqual(ready_members(layout), [])

    def test_control_points_are_solved_from_beam_local_stations(self):
        layout = load_layout(LAYOUT)
        points = {point["name"]: point for point in solve_control_points(layout)}
        theta = math.radians(18.0)
        gc = 0.439
        gf = 2.119
        ge = 3.739

        self.assertAlmostEqual(points["A"]["z_m"], 3.0)
        self.assertAlmostEqual(points["F"]["x_m"], 0.0)
        self.assertAlmostEqual(points["F"]["z_m"], 3.1)
        self.assertAlmostEqual(points["B"]["x_m"], -0.1)
        self.assertAlmostEqual(points["D"]["x_m"], 0.1)
        self.assertAlmostEqual(points["C"]["x_m"], (gc - gf) * math.cos(theta))
        self.assertAlmostEqual(points["C"]["z_m"], 3.1 + (gc - gf) * math.sin(theta))
        self.assertAlmostEqual(points["E"]["x_m"], (ge - gf) * math.cos(theta))
        self.assertAlmostEqual(points["E"]["z_m"], 3.1 + (ge - gf) * math.sin(theta))
        self.assertAlmostEqual(points["G_global"]["x_m"], -gf * math.cos(theta))
        self.assertAlmostEqual(points["G_global"]["z_m"], 3.1 - gf * math.sin(theta))

    def test_beam_local_sections_and_angle_are_reported(self):
        layout = load_layout(LAYOUT)
        sections = {row["section"]: row for row in beam_local_sections(layout)}
        self.assertEqual(list(sections.keys()), ["G", "C", "F", "E"])
        self.assertAlmostEqual(sections["C"]["station_m"], 0.439)
        self.assertAlmostEqual(sections["F"]["station_m"], 2.119)

        checks = {check["check_item"]: check for check in length_and_angle_checks(layout)}
        self.assertEqual(checks["GC_GF_GE_ORDER"]["passed"], PASSED)
        self.assertEqual(checks["CE_ANGLE"]["passed"], PASSED)
        self.assertAlmostEqual(checks["CE_ANGLE"]["calc_value"], 18.0)

        bad_layout = copy.deepcopy(layout)
        for row in bad_layout["dimension_inputs"]:
            if row["name"] == "GE_mm":
                row["value"] = 2000
        bad_checks = {check["check_item"]: check for check in length_and_angle_checks(bad_layout)}
        self.assertEqual(bad_checks["GC_GF_GE_ORDER"]["passed"], "不通过")

    def test_control_length_checks_use_one_millimeter_tolerance(self):
        layout = copy.deepcopy(load_layout(LAYOUT))
        checks = {check["segment"]: check for check in control_length_checks(layout)}
        self.assertEqual(checks["BC"]["passed"], "不通过")

        set_input_status(
            layout,
            {"Z_A_mm", "X_F_mm", "Z_F_mm", "R_hoop_mm", "GC_mm", "GF_mm", "GE_mm"},
            "已确认",
        )

        exact = {check["segment"]: check["calc_length_m"] for check in brace_length_checks(layout)}
        for row in layout["dimension_inputs"]:
            if row["name"] == "L_BC_draw_mm":
                row["value"] = exact["BC"] * 1000
            if row["name"] == "L_DE_draw_mm":
                row["value"] = exact["DE"] * 1000

        checks = {check["segment"]: check for check in control_length_checks(layout)}
        self.assertEqual(checks["BC"]["passed"], PASSED)
        self.assertEqual(checks["DE"]["passed"], PASSED)

    def test_coordinate_workbook_contains_expected_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "coordinate_check.xlsx"
            create_coordinate_workbook(LAYOUT, output)
            wb = load_workbook(output, data_only=True)
            self.assertEqual(
                wb.sheetnames,
                [
                    "关键尺寸输入",
                    "斜梁局部截面",
                    "控制点坐标",
                    "长度与角度校核",
                    "构件轴线校核",
                    "图纸标注记录",
                    "坐标系说明",
                ],
            )
            ws = wb["控制点坐标"]
            self.assertEqual(ws["A3"].value, "A")
            self.assertEqual(ws["E3"].value, 3.0)

    def test_export_assembly_inputs_requires_confirmed_checked_members(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "assembly_inputs.json"
            export_assembly_inputs(LAYOUT, output, include_draft=True)
            payload = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(payload["ready_members"], [])
            self.assertIn("control_points", payload)
            self.assertIn("beam_local_sections", payload)
            self.assertIn("geometry_checks", payload)
            self.assertIn("length_checks", payload)
            self.assertIn("draft_member_checks", payload)

    def test_world_to_pixel_uses_x_right_and_z_up(self):
        layout = load_layout(LAYOUT)
        calibration = layout["image_calibration"]
        origin_px = tuple(calibration["origin_px"])
        self.assertEqual(world_to_pixel({"x_m": 0, "z_m": 0}, calibration), origin_px)
        px, py = world_to_pixel({"x_m": 1, "z_m": 1}, calibration)
        self.assertGreater(px, origin_px[0])
        self.assertLess(py, origin_px[1])

    def test_sample_image_path_exists_when_available(self):
        sample = Path(r"C:\Users\LDT\AppData\Local\Temp\codex-clipboard-f7777794-a0d7-4343-9d7b-1086c9666147.png")
        if sample.exists():
            with Image.open(sample) as image:
                self.assertGreater(image.size[0], 500)
                self.assertGreater(image.size[1], 500)


if __name__ == "__main__":
    unittest.main()
