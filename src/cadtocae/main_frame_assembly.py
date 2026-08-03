from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from string import Template
from typing import Any

from openpyxl import load_workbook

from .runs import create_run_paths, update_manifest


CONFIRMED = "\u5df2\u786e\u8ba4"
MANUAL_CHECK = "\u9700\u4eba\u5de5\u786e\u8ba4"
PASSED = "\u901a\u8fc7"
FAILED = "\u4e0d\u901a\u8fc7"

DEFAULT_PROJECT_CODE = "PV_SUPPORT"
POINT_COMPARE_TOLERANCE_M = 1.0e-6
VALUE_COMPARE_TOLERANCE = 1.0e-6

MAIN_COMPONENT_CODES = {
    "INCLINED_BEAM": "INCLINED_BEAM",
    "BRACE_FRONT": "BRACE_FRONT",
    "BRACE_REAR": "BRACE_REAR",
}
PURLIN_AXIS_INPUT_NAMES = ("HF_mm", "HS_mm", "HP_mm", "HQ_mm", "HR_mm")
PURLIN_AXIS_POINT_NAMES = ("S", "P", "Q", "R")
PURLIN_SHORT_LENGTH_M = 0.05
PURLIN_GROUP_Y_OFFSET_M = 0.025
SPQR_COLLINEAR_TOLERANCE_M = 1.0e-6


@dataclass(frozen=True)
class ExcelInput:
    name: str
    meaning: str
    value: Any
    unit: str
    status: str
    note: str
    row: int


def _float(value: Any, name: str) -> float:
    if value is None or value == "":
        raise ValueError("Missing numeric input: %s" % name)
    return float(value)


def _length_m(inputs: dict[str, ExcelInput], name: str) -> float:
    row = inputs[name]
    value = _float(row.value, name)
    unit = (row.unit or "").strip().lower()
    if unit == "mm":
        return value / 1000.0
    if unit == "m":
        return value
    raise ValueError("Input %s must use mm or m, got %r" % (name, row.unit))


def _value(inputs: dict[str, ExcelInput], name: str) -> float:
    return _float(inputs[name].value, name)


def _point(x: float, y: float, z: float, status: str = CONFIRMED, note: str = "") -> dict[str, Any]:
    return {
        "coords": [x, y, z],
        "x_m": x,
        "y_m": y,
        "z_m": z,
        "status": status,
        "note": note,
    }


def _coords(point: dict[str, Any]) -> list[float]:
    return [float(point["x_m"]), float(point["y_m"]), float(point["z_m"])]


def _sub(a: list[float], b: list[float]) -> list[float]:
    return [a[0] - b[0], a[1] - b[1], a[2] - b[2]]


def _add(a: list[float], b: list[float]) -> list[float]:
    return [a[0] + b[0], a[1] + b[1], a[2] + b[2]]


def _scale(a: list[float], factor: float) -> list[float]:
    return [a[0] * factor, a[1] * factor, a[2] * factor]


def _dot(a: list[float], b: list[float]) -> float:
    return a[0] * b[0] + a[1] * b[1] + a[2] * b[2]


def _cross(a: list[float], b: list[float]) -> list[float]:
    return [
        a[1] * b[2] - a[2] * b[1],
        a[2] * b[0] - a[0] * b[2],
        a[0] * b[1] - a[1] * b[0],
    ]


def _norm(a: list[float]) -> float:
    return math.sqrt(_dot(a, a))


def _point_line_distance(point: list[float], start: list[float], end: list[float]) -> float:
    line = _sub(end, start)
    line_length = _norm(line)
    if line_length <= 1.0e-15:
        return distance(point, start)
    return _norm(_cross(_sub(point, start), line)) / line_length


def distance(a: list[float], b: list[float]) -> float:
    dx = a[0] - b[0]
    dy = a[1] - b[1]
    dz = a[2] - b[2]
    return math.sqrt(dx * dx + dy * dy + dz * dz)


def rotate_x(point: list[float], angle_deg: float) -> list[float]:
    angle = math.radians(angle_deg)
    c = math.cos(angle)
    s = math.sin(angle)
    x, y, z = point
    return [x, y * c - z * s, y * s + z * c]


def rotate_y(point: list[float], angle_deg: float) -> list[float]:
    angle = math.radians(angle_deg)
    c = math.cos(angle)
    s = math.sin(angle)
    x, y, z = point
    return [x * c + z * s, y, -x * s + z * c]


def rotate_z(point: list[float], angle_deg: float) -> list[float]:
    angle = math.radians(angle_deg)
    c = math.cos(angle)
    s = math.sin(angle)
    x, y, z = point
    return [x * c - y * s, x * s + y * c, z]


def transform_local(point: list[float], rotate_y_deg: float, roll_about_axis_deg: float = 0.0) -> list[float]:
    return rotate_y(rotate_z(point, roll_about_axis_deg), rotate_y_deg)


def transform_rotation_sequence(point: list[float], rotation_sequence: list[dict[str, Any]]) -> list[float]:
    result = [float(point[0]), float(point[1]), float(point[2])]
    for rotation in rotation_sequence:
        axis = str(rotation.get("axis") or "").upper()
        angle_deg = float(rotation.get("angle_deg") or 0.0)
        if axis == "X":
            result = rotate_x(result, angle_deg)
        elif axis == "Y":
            result = rotate_y(result, angle_deg)
        elif axis == "Z":
            result = rotate_z(result, angle_deg)
        else:
            raise ValueError("Unsupported rotation axis: %s" % axis)
    return result


def vector_angle_from_x(vector: list[float]) -> float:
    return math.degrees(math.atan2(vector[2], vector[0]))


def rotate_y_for_local_z_to_vector(start: list[float], end: list[float]) -> float:
    dx = end[0] - start[0]
    dz = end[2] - start[2]
    return math.degrees(math.atan2(dx, dz))


def _angle_error_deg(calc: float, reference: float) -> float:
    return (calc - reference + 180.0) % 360.0 - 180.0


def _pass_fail(error: float, tolerance: float) -> str:
    return PASSED if abs(error) <= tolerance else FAILED


def _find_header_row(ws, required: tuple[str, ...]) -> int:
    for row in range(1, ws.max_row + 1):
        values = [ws.cell(row, column).value for column in range(1, len(required) + 1)]
        if tuple(values) == required:
            return row
    raise ValueError("Cannot find header row: %s" % (required,))


def _find_table_header(wb, required_headers: tuple[str, ...], preferred_sheets: tuple[str, ...] = ()):
    sheet_names = list(preferred_sheets) + [name for name in wb.sheetnames if name not in preferred_sheets]
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            headers = [ws.cell(row, column).value for column in range(1, ws.max_column + 1)]
            header_map = {str(value): index + 1 for index, value in enumerate(headers) if value is not None}
            if all(header in header_map for header in required_headers):
                return ws, row, header_map
    raise ValueError("Cannot find table headers: %s" % (required_headers,))


def read_excel_inputs(excel_path: str | Path) -> tuple[dict[str, ExcelInput], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    path = Path(excel_path)
    wb_values = load_workbook(path, data_only=True)

    ws_values, input_header, input_cols = _find_table_header(
        wb_values,
        ("参数名", "参数含义", "数值", "单位", "校核状态", "备注"),
        preferred_sheets=("关键尺寸输入",),
    )
    inputs: dict[str, ExcelInput] = {}
    row = input_header + 1
    while row <= ws_values.max_row:
        name = ws_values.cell(row, input_cols["参数名"]).value
        if not name:
            break
        inputs[str(name)] = ExcelInput(
            name=str(name),
            meaning=str(ws_values.cell(row, input_cols["参数含义"]).value or ""),
            value=ws_values.cell(row, input_cols["数值"]).value,
            unit=str(ws_values.cell(row, input_cols["单位"]).value or ""),
            status=str(ws_values.cell(row, input_cols["校核状态"]).value or ""),
            note=str(ws_values.cell(row, input_cols["备注"]).value or ""),
            row=row,
        )
        row += 1

    point_ws, output_header, point_cols = _find_table_header(
        wb_values,
        ("点名", "X_m", "Y_m", "Z_m", "校核状态"),
        preferred_sheets=("控制点坐标",),
    )
    cached_points: dict[str, dict[str, Any]] = {}
    row = output_header + 1
    point_note_col = point_cols.get("说明") or point_cols.get("标注说明")
    while row <= point_ws.max_row:
        name = point_ws.cell(row, point_cols["点名"]).value
        if not name:
            break
        cached_points[str(name)] = {
            "coords": [
                point_ws.cell(row, point_cols["X_m"]).value,
                point_ws.cell(row, point_cols["Y_m"]).value,
                point_ws.cell(row, point_cols["Z_m"]).value,
            ],
            "status": str(point_ws.cell(row, point_cols["校核状态"]).value or ""),
            "note": str(point_ws.cell(row, point_note_col).value or "") if point_note_col else "",
            "row": row,
        }
        row += 1

    check_ws, check_header, check_cols = _find_table_header(
        wb_values,
        ("校核项", "计算值", "图纸/输入值", "误差", "允许误差", "是否通过", "校核状态"),
        preferred_sheets=("长度与角度校核",),
    )
    cached_checks: dict[str, dict[str, Any]] = {}
    row = check_header + 1
    while row <= check_ws.max_row:
        name = check_ws.cell(row, check_cols["校核项"]).value
        if not name:
            break
        cached_checks[str(name)] = {
            "calc_value": check_ws.cell(row, check_cols["计算值"]).value,
            "reference_value": check_ws.cell(row, check_cols["图纸/输入值"]).value,
            "error": check_ws.cell(row, check_cols["误差"]).value,
            "tolerance": check_ws.cell(row, check_cols["允许误差"]).value,
            "passed": str(check_ws.cell(row, check_cols["是否通过"]).value or ""),
            "status": str(check_ws.cell(row, check_cols["校核状态"]).value or ""),
            "row": row,
        }
        row += 1

    return inputs, cached_points, cached_checks


def input_status(inputs: dict[str, ExcelInput], names: list[str]) -> str:
    for name in names:
        if inputs.get(name) is None or inputs[name].status != CONFIRMED:
            return MANUAL_CHECK
    return CONFIRMED


def _has_numeric_inputs(inputs: dict[str, ExcelInput], names: tuple[str, ...]) -> bool:
    for name in names:
        row = inputs.get(name)
        if row is None or row.value in (None, ""):
            return False
        try:
            _float(row.value, name)
        except (TypeError, ValueError):
            return False
    return True


def solve_points_from_inputs(inputs: dict[str, ExcelInput]) -> dict[str, dict[str, Any]]:
    theta_deg = _value(inputs, "theta_deg")
    theta = math.radians(theta_deg)
    cos_theta = math.cos(theta)
    sin_theta = math.sin(theta)

    z_a = _length_m(inputs, "Z_A_mm")
    x_f = _length_m(inputs, "X_F_mm")
    z_f = _length_m(inputs, "Z_F_mm")
    z_bd = _length_m(inputs, "Z_BD_mm")
    r_hoop = _length_m(inputs, "R_hoop_mm")
    gc = _length_m(inputs, "GC_mm")
    gf = _length_m(inputs, "GF_mm")
    ge = _length_m(inputs, "GE_mm")

    points = {
        "O": _point(0.0, 0.0, 0.0, CONFIRMED, "Origin"),
        "A": _point(0.0, 0.0, z_a, input_status(inputs, ["Z_A_mm"]), "Upper column top"),
        "F": _point(x_f, 0.0, z_f, input_status(inputs, ["X_F_mm", "Z_F_mm"]), "Beam-column reference section"),
        "B": _point(-r_hoop, 0.0, z_bd, input_status(inputs, ["Z_BD_mm", "R_hoop_mm"]), "Front brace hoop point"),
        "D": _point(r_hoop, 0.0, z_bd, input_status(inputs, ["Z_BD_mm", "R_hoop_mm"]), "Rear brace hoop point"),
    }

    f = _coords(points["F"])
    u = [cos_theta, 0.0, sin_theta]
    c = _add(f, _scale(u, gc - gf))
    e = _add(f, _scale(u, ge - gf))
    g_global = _add(f, _scale(u, -gf))

    points["C"] = _point(
        c[0],
        c[1],
        c[2],
        input_status(inputs, ["theta_deg", "X_F_mm", "Z_F_mm", "GC_mm", "GF_mm"]),
        "Beam-front brace section",
    )
    points["E"] = _point(
        e[0],
        e[1],
        e[2],
        input_status(inputs, ["theta_deg", "X_F_mm", "Z_F_mm", "GE_mm", "GF_mm"]),
        "Beam-rear brace section",
    )
    points["G_global"] = _point(
        g_global[0],
        g_global[1],
        g_global[2],
        input_status(inputs, ["theta_deg", "X_F_mm", "Z_F_mm", "GF_mm"]),
        "Derived global position of beam local origin G",
    )
    if _has_numeric_inputs(inputs, PURLIN_AXIS_INPUT_NAMES):
        hf = _length_m(inputs, "HF_mm")
        hs = _length_m(inputs, "HS_mm")
        hp = _length_m(inputs, "HP_mm")
        hq = _length_m(inputs, "HQ_mm")
        hr = _length_m(inputs, "HR_mm")
        n = [-sin_theta, 0.0, cos_theta]
        h = _add(f, _scale(n, hf))
        purlin_status_names = ["theta_deg", "X_F_mm", "Z_F_mm", "HF_mm"]
        points["H"] = _point(h[0], h[1], h[2], input_status(inputs, purlin_status_names), "PV/purlin axis reference point")
        points["S"] = _point(
            h[0] - hs * cos_theta,
            h[1],
            h[2] - hs * sin_theta,
            input_status(inputs, [*purlin_status_names, "HS_mm"]),
            "Purlin axis point S, negative beam-axis side from H",
        )
        points["P"] = _point(
            h[0] - hp * cos_theta,
            h[1],
            h[2] - hp * sin_theta,
            input_status(inputs, [*purlin_status_names, "HP_mm"]),
            "Purlin axis point P, negative beam-axis side from H",
        )
        points["Q"] = _point(
            h[0] + hq * cos_theta,
            h[1],
            h[2] + hq * sin_theta,
            input_status(inputs, [*purlin_status_names, "HQ_mm"]),
            "Purlin axis point Q, positive beam-axis side from H",
        )
        points["R"] = _point(
            h[0] + hr * cos_theta,
            h[1],
            h[2] + hr * sin_theta,
            input_status(inputs, [*purlin_status_names, "HR_mm"]),
            "Purlin axis point R, positive beam-axis side from H",
        )
    return points


def build_checks(inputs: dict[str, ExcelInput], points: dict[str, dict[str, Any]], beam_length_m: float | None) -> dict[str, dict[str, Any]]:
    theta_deg = _value(inputs, "theta_deg")
    gc = _length_m(inputs, "GC_mm")
    gf = _length_m(inputs, "GF_mm")
    ge = _length_m(inputs, "GE_mm")
    l_bc = _length_m(inputs, "L_BC_draw_mm")
    l_de = _length_m(inputs, "L_DE_draw_mm")
    control_tol = _length_m(inputs, "control_tolerance_m")
    angle_tol = _value(inputs, "angle_tolerance_deg")

    c = _coords(points["C"])
    e = _coords(points["E"])
    b = _coords(points["B"])
    d = _coords(points["D"])
    ce_angle = vector_angle_from_x(_sub(e, c))
    ce_error = _angle_error_deg(ce_angle, theta_deg)
    bc_calc = distance(b, c)
    de_calc = distance(d, e)
    beam_length_ok = True
    if beam_length_m is not None:
        beam_length_ok = ge < beam_length_m

    checks = {
        "GC_GF_GE_ORDER": {
            "calc_value": "GC=%.6f, GF=%.6f, GE=%.6f" % (gc, gf, ge),
            "reference_value": "0 < GC < GF < GE < beam_length",
            "error": None,
            "tolerance": None,
            "passed": PASSED if 0.0 < gc < gf < ge and beam_length_ok else FAILED,
            "status": input_status(inputs, ["GC_mm", "GF_mm", "GE_mm"]),
        },
        "CF_LOCAL": {
            "calc_value": gf - gc,
            "reference_value": "GF-GC",
            "error": None,
            "tolerance": None,
            "passed": PASSED if gf - gc > 0.0 else FAILED,
            "status": input_status(inputs, ["GC_mm", "GF_mm"]),
        },
        "FE_LOCAL": {
            "calc_value": ge - gf,
            "reference_value": "GE-GF",
            "error": None,
            "tolerance": None,
            "passed": PASSED if ge - gf > 0.0 else FAILED,
            "status": input_status(inputs, ["GF_mm", "GE_mm"]),
        },
        "CE_ANGLE": {
            "calc_value": ce_angle,
            "reference_value": theta_deg,
            "error": ce_error,
            "tolerance": angle_tol,
            "passed": _pass_fail(ce_error, angle_tol),
            "status": input_status(inputs, ["theta_deg", "X_F_mm", "Z_F_mm", "GC_mm", "GF_mm", "GE_mm"]),
        },
        "BC": {
            "calc_value": bc_calc,
            "reference_value": l_bc,
            "error": bc_calc - l_bc,
            "tolerance": control_tol,
            "passed": _pass_fail(bc_calc - l_bc, control_tol),
            "status": input_status(
                inputs,
                ["theta_deg", "X_F_mm", "Z_F_mm", "Z_BD_mm", "R_hoop_mm", "GC_mm", "GF_mm", "L_BC_draw_mm"],
            ),
        },
        "DE": {
            "calc_value": de_calc,
            "reference_value": l_de,
            "error": de_calc - l_de,
            "tolerance": control_tol,
            "passed": _pass_fail(de_calc - l_de, control_tol),
            "status": input_status(
                inputs,
                ["theta_deg", "X_F_mm", "Z_F_mm", "Z_BD_mm", "R_hoop_mm", "GE_mm", "GF_mm", "L_DE_draw_mm"],
            ),
        },
    }
    if all(name in points for name in PURLIN_AXIS_POINT_NAMES) and inputs.get("pv_axis_angle_tolerance_deg") is not None:
        s = _coords(points["S"])
        p = _coords(points["P"])
        q = _coords(points["Q"])
        r = _coords(points["R"])
        max_offset = max(_point_line_distance(p, s, r), _point_line_distance(q, s, r))
        spqr_angle = vector_angle_from_x(_sub(r, s))
        spqr_angle_error = _angle_error_deg(spqr_angle, theta_deg)
        spqr_angle_tol = _value(inputs, "pv_axis_angle_tolerance_deg")
        purlin_status = input_status(inputs, ["theta_deg", "X_F_mm", "Z_F_mm", *PURLIN_AXIS_INPUT_NAMES, "pv_axis_angle_tolerance_deg"])
        checks["SPQR_COLLINEAR"] = {
            "calc_value": max_offset,
            "reference_value": 0.0,
            "error": max_offset,
            "tolerance": SPQR_COLLINEAR_TOLERANCE_M,
            "passed": PASSED if max_offset <= SPQR_COLLINEAR_TOLERANCE_M else FAILED,
            "status": purlin_status,
        }
        checks["SPQR_ANGLE"] = {
            "calc_value": spqr_angle,
            "reference_value": theta_deg,
            "error": spqr_angle_error,
            "tolerance": spqr_angle_tol,
            "passed": _pass_fail(spqr_angle_error, spqr_angle_tol),
            "status": purlin_status,
        }
    return checks


def read_components_payload(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    text = source.read_text(encoding="utf-8")
    if source.suffix.lower() == ".py" or "COMPONENTS_JSON" in text:
        match = re.search(r"COMPONENTS_JSON\s*=\s*r?'''(?P<payload>.*?)'''", text, re.DOTALL)
        if not match:
            match = re.search(r'COMPONENTS_JSON\s*=\s*r?"""(?P<payload>.*?)"""', text, re.DOTALL)
        if not match:
            raise ValueError("Cannot find embedded COMPONENTS_JSON in %s" % source)
        return json.loads(match.group("payload"))
    return json.loads(text)


def load_components(path: str | Path) -> dict[str, dict[str, Any]]:
    payload = read_components_payload(path)
    by_code: dict[str, dict[str, Any]] = {}
    for row in payload.get("components", []):
        code = row.get("component_code")
        if code:
            by_code[str(code)] = row
    return by_code


def _component(by_code: dict[str, dict[str, Any]], code: str) -> dict[str, Any]:
    if code not in by_code:
        raise ValueError("Missing component_code in components JSON: %s" % code)
    return by_code[code]


def _optional_component(by_code: dict[str, dict[str, Any]], code: str) -> dict[str, Any] | None:
    return by_code.get(code)


def compare_cached_points(points: dict[str, dict[str, Any]], cached_points: dict[str, dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    for name, point in points.items():
        if name == "O":
            continue
        cached = cached_points.get(name)
        if not cached:
            warnings.append("Excel cached point %s is missing." % name)
            continue
        cached_coords = cached.get("coords") or []
        if len(cached_coords) != 3 or any(value is None for value in cached_coords):
            warnings.append("Excel cached point %s has empty coordinate cache." % name)
            continue
        err = distance(_coords(point), [float(cached_coords[0]), float(cached_coords[1]), float(cached_coords[2])])
        if err > POINT_COMPARE_TOLERANCE_M:
            warnings.append("Excel cached point %s differs from Python recompute by %.9g m." % (name, err))
    return warnings


def compare_cached_checks(checks: dict[str, dict[str, Any]], cached_checks: dict[str, dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    for name, check in checks.items():
        cached = cached_checks.get(name)
        if not cached:
            warnings.append("Excel cached check %s is missing." % name)
            continue
        for key in ("calc_value", "reference_value", "error", "tolerance"):
            calc = check.get(key)
            cached_value = cached.get(key)
            if calc is None or cached_value in (None, ""):
                continue
            if isinstance(calc, (int, float)) and isinstance(cached_value, (int, float)):
                if abs(float(calc) - float(cached_value)) > VALUE_COMPARE_TOLERANCE:
                    warnings.append(
                        "Excel cached check %s.%s differs from Python recompute by %.9g."
                        % (name, key, float(calc) - float(cached_value))
                    )
        if cached.get("passed") and check.get("passed") != cached.get("passed"):
            warnings.append("Excel cached check %s pass state is %r, Python recompute is %r." % (name, cached.get("passed"), check.get("passed")))
    return warnings


def member_length_check(points: dict[str, dict[str, Any]], start: str, end: str, part_length_m: float | None, tolerance_m: float) -> dict[str, Any]:
    axis_length = distance(_coords(points[start]), _coords(points[end]))
    error = None if part_length_m is None else axis_length - float(part_length_m)
    passed = MANUAL_CHECK if error is None else _pass_fail(error, tolerance_m)
    return {
        "start": start,
        "end": end,
        "axis_length_m": axis_length,
        "part_length_m": part_length_m,
        "error_m": error,
        "tolerance_m": tolerance_m,
        "passed": passed,
    }


def _section_reference_xy(component: dict[str, Any]) -> dict[str, Any]:
    kind = component.get("section_kind")
    params = component.get("section_params_m") or {}
    if kind == "C_CHANNEL":
        h = float(params.get("h_m") or 0.0)
        if h > 0.0:
            return {
                "x_m": 0.0,
                "y_m": h / 2.0,
                "rule": "C_CHANNEL_WEB_MIDPOINT",
                "open_side_local": "+X",
                "open_side_target_global": "-Y",
            }
    if kind == "ANGLE":
        a = float(params.get("leg_a_m") or 0.0)
        b = float(params.get("leg_b_m") or 0.0)
        t = float(params.get("t_m") or 0.0)
        area = a * t + t * b - t * t
        if area > 0.0:
            x = (a * t * (a / 2.0) + t * b * (t / 2.0) - t * t * (t / 2.0)) / area
            y = (a * t * (t / 2.0) + t * b * (b / 2.0) - t * t * (t / 2.0)) / area
            return {
                "x_m": x,
                "y_m": y,
                "rule": "ANGLE_SOLID_SECTION_CENTROID",
                "open_side_local": "+X,+Y",
                "open_side_target_global": "",
            }
    return {
        "x_m": 0.0,
        "y_m": 0.0,
        "rule": "SECTION_ORIGIN",
        "open_side_local": "",
        "open_side_target_global": "",
    }


def _default_roll_about_axis_deg(name: str, component: dict[str, Any]) -> float:
    roll = 0.0
    if component.get("section_kind") == "C_CHANNEL":
        roll = -90.0
    if name == "INCLINED_BEAM":
        roll += 180.0
    return roll


def _local_reference_point(component: dict[str, Any], station_m: float) -> list[float]:
    ref = _section_reference_xy(component)
    return [float(ref["x_m"]), float(ref["y_m"]), float(station_m)]


def _clean_vector(vector: list[float], tolerance: float = 1.0e-12) -> list[float]:
    return [0.0 if abs(float(value)) <= tolerance else float(value) for value in vector]


def _translation_for_anchor(
    local_anchor: list[float],
    global_anchor: list[float],
    rotate_y_deg: float,
    roll_about_axis_deg: float = 0.0,
) -> list[float]:
    rotated = rotate_y(local_anchor, rotate_y_deg)
    return [global_anchor[0] - rotated[0], global_anchor[1] - rotated[1], global_anchor[2] - rotated[2]]


def _translation_for_rotation_sequence(local_anchor: list[float], global_anchor: list[float], rotation_sequence: list[dict[str, Any]]) -> list[float]:
    rotated = transform_rotation_sequence(local_anchor, rotation_sequence)
    return [global_anchor[0] - rotated[0], global_anchor[1] - rotated[1], global_anchor[2] - rotated[2]]


def _instance_name(component: dict[str, Any], suffix: str | None = None) -> str:
    part_name = str(component["part_name"])
    base = part_name[2:] if part_name.startswith("P_") else part_name
    return "I_%s_%s" % (base, suffix) if suffix else "I_%s" % base


def _short_purlin_part_name(component: dict[str, Any]) -> str:
    return "%s_50MM" % component["part_name"]


def _purlin_rotation_sequence(theta_deg: float) -> list[dict[str, Any]]:
    return [
        {"axis": "X", "angle_deg": 90.0},
        {"axis": "Y", "angle_deg": -theta_deg},
    ]


def _purlin_support_rotation_sequence(theta_deg: float) -> list[dict[str, Any]]:
    return [
        {"axis": "Z", "angle_deg": 90.0},
        {"axis": "X", "angle_deg": 90.0},
        {"axis": "Y", "angle_deg": -theta_deg},
    ]


def _purlin_upper_flange_anchor(component: dict[str, Any]) -> list[float]:
    params = component.get("section_params_m") or {}
    return [float(params["b_m"]) / 2.0, float(params["h_m"]), PURLIN_SHORT_LENGTH_M / 2.0]


def _purlin_support_inside_corner_anchor(component: dict[str, Any]) -> list[float]:
    length = float(component.get("length_m") or PURLIN_SHORT_LENGTH_M)
    return [0.0, 0.0, length / 2.0]


def _copy_component_for_payload(component: dict[str, Any], part_name: str, length_m: float) -> dict[str, Any]:
    copied = json.loads(json.dumps(component, ensure_ascii=False))
    copied["part_name"] = part_name
    copied["length_m"] = length_m
    return copied


def _member_with_rotation_sequence(
    name: str,
    phase: str,
    component: dict[str, Any],
    part_name: str,
    instance_name: str,
    local_anchor: list[float],
    global_anchor_name: str,
    global_anchor: list[float],
    rotation_sequence: list[dict[str, Any]],
    part_length_m: float | None = None,
    source_part_name: str | None = None,
    part_component: dict[str, Any] | None = None,
    section_reference: dict[str, Any] | None = None,
    axis_checks: list[dict[str, Any]] | None = None,
    placement_note: str = "",
) -> dict[str, Any]:
    member = {
        "name": name,
        "phase": phase,
        "part_name": part_name,
        "source_part_name": source_part_name or component["part_name"],
        "component_code": component.get("component_code"),
        "instance_name": instance_name,
        "local_anchor": local_anchor,
        "global_anchor_name": global_anchor_name,
        "global_anchor": global_anchor,
        "rotation_sequence": rotation_sequence,
        "translation": _translation_for_rotation_sequence(local_anchor, global_anchor, rotation_sequence),
        "part_length_m": part_length_m if part_length_m is not None else component.get("length_m"),
        "section_kind": component.get("section_kind"),
        "section_params_m": component.get("section_params_m") or {},
        "section_reference": section_reference or _section_reference_xy(component),
        "model_policy": component.get("model_policy"),
        "placement_note": placement_note,
    }
    if part_component is not None:
        member["part_component"] = part_component
    if axis_checks is not None:
        member["axis_checks"] = axis_checks
    return member


def _build_purlin_members(
    purlin_component: dict[str, Any],
    purlin_support_component: dict[str, Any],
    points: dict[str, dict[str, Any]],
    theta_deg: float,
    u: list[float],
    n: list[float],
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if purlin_component.get("section_kind") != "C_CHANNEL":
        return [], {"enabled": False, "reason": "PURLIN section_kind is not C_CHANNEL."}, {}, ["PURLIN section_kind is not C_CHANNEL; purlin assembly skipped."]
    if purlin_support_component.get("section_kind") != "ANGLE":
        return [], {"enabled": False, "reason": "PURLIN_SUPPORT section_kind is not ANGLE."}, {}, ["PURLIN_SUPPORT section_kind is not ANGLE; purlin assembly skipped."]

    purlin_params = purlin_component.get("section_params_m") or {}
    purlin_b = float(purlin_params["b_m"])
    purlin_h = float(purlin_params["h_m"])
    derived_part_name = _short_purlin_part_name(purlin_component)
    purlin_part_component = _copy_component_for_payload(purlin_component, derived_part_name, PURLIN_SHORT_LENGTH_M)
    purlin_rotation = _purlin_rotation_sequence(theta_deg)
    support_rotation = _purlin_support_rotation_sequence(theta_deg)
    purlin_local_anchor = _purlin_upper_flange_anchor(purlin_component)
    support_local_anchor = _purlin_support_inside_corner_anchor(purlin_support_component)
    support_length = float(purlin_support_component.get("length_m") or PURLIN_SHORT_LENGTH_M)
    if abs(support_length - PURLIN_SHORT_LENGTH_M) > 1.0e-9:
        warnings.append("PURLIN_SUPPORT length is %.6f m; expected 0.050000 m for the XZ simplified slice." % support_length)

    members: list[dict[str, Any]] = []
    checks: dict[str, dict[str, Any]] = {}
    for point_name in PURLIN_AXIS_POINT_NAMES:
        control_anchor = _coords(points[point_name])
        global_anchor = _add(control_anchor, [0.0, PURLIN_GROUP_Y_OFFSET_M, 0.0])
        web_top = _add(global_anchor, _scale(u, -purlin_b / 2.0))
        web_bottom = _add(web_top, _scale(n, -purlin_h))
        purlin_member = _member_with_rotation_sequence(
            "PURLIN_%s" % point_name,
            "step04_purlins",
            purlin_component,
            derived_part_name,
            _instance_name(purlin_component, point_name),
            purlin_local_anchor,
            point_name,
            global_anchor,
            purlin_rotation,
            part_length_m=PURLIN_SHORT_LENGTH_M,
            source_part_name=purlin_component["part_name"],
            part_component=purlin_part_component,
            section_reference={
                "x_m": purlin_local_anchor[0],
                "y_m": purlin_local_anchor[1],
                "z_m": purlin_local_anchor[2],
                "rule": "C_CHANNEL_UPPER_FLANGE_CENTER",
                "open_side_local": "+X",
                "open_side_target_global": "PV_AXIS_POSITIVE",
            },
            axis_checks=[
                {"name": "flange_axis", "local_vector": [1.0, 0.0, 0.0], "expected_global": u, "tolerance": 1.0e-6},
                {"name": "web_axis", "local_vector": [0.0, 1.0, 0.0], "expected_global": n, "tolerance": 1.0e-6},
                {"name": "length_axis_parallel_y", "local_vector": [0.0, 0.0, 1.0], "expected_global": [0.0, -1.0, 0.0], "tolerance": 1.0e-6},
            ],
            placement_note="50mm purlin slice; upper flange center maps to %s; length axis is parallel to global Y." % point_name,
        )
        support_member = _member_with_rotation_sequence(
            "PURLIN_SUPPORT_%s" % point_name,
            "step04_purlins",
            purlin_support_component,
            purlin_support_component["part_name"],
            _instance_name(purlin_support_component, point_name),
            support_local_anchor,
            "%s_WEB_BOTTOM" % point_name,
            web_bottom,
            support_rotation,
            part_length_m=support_length,
            source_part_name=purlin_support_component["part_name"],
            section_reference={
                "x_m": support_local_anchor[0],
                "y_m": support_local_anchor[1],
                "z_m": support_local_anchor[2],
                "rule": "ANGLE_INSIDE_CORNER_AT_PURLIN_WEB_BOTTOM",
                "open_side_local": "+Y",
                "open_side_target_global": "PV_AXIS_NEGATIVE",
            },
            axis_checks=[
                {"name": "long_leg_axis", "local_vector": [1.0, 0.0, 0.0], "expected_global": n, "tolerance": 1.0e-6},
                {"name": "short_leg_axis", "local_vector": [0.0, 1.0, 0.0], "expected_global": _scale(u, -1.0), "tolerance": 1.0e-6},
                {"name": "length_axis_parallel_y", "local_vector": [0.0, 0.0, 1.0], "expected_global": [0.0, -1.0, 0.0], "tolerance": 1.0e-6},
            ],
            placement_note="Angle support inside corner contacts purlin web-bottom corner; long leg follows the web and short leg is flush with the lower flange line.",
        )
        members.extend([purlin_member, support_member])
        checks["PURLIN_%s_PLACEMENT" % point_name] = {
            "anchor": point_name,
            "control_point": control_anchor,
            "y_offset_m": PURLIN_GROUP_Y_OFFSET_M,
            "upper_flange_center": global_anchor,
            "web_top": web_top,
            "web_bottom": web_bottom,
            "lower_flange_center": _add(global_anchor, _scale(n, -purlin_h)),
            "flange_axis_global": u,
            "web_axis_global": n,
            "length_axis_global": [0.0, -1.0, 0.0],
            "passed": PASSED,
        }
        checks["PURLIN_SUPPORT_%s_PLACEMENT" % point_name] = {
            "anchor": "%s_WEB_BOTTOM" % point_name,
            "y_offset_m": PURLIN_GROUP_Y_OFFSET_M,
            "inside_corner": web_bottom,
            "long_leg_axis_global": n,
            "short_leg_axis_global": _scale(u, -1.0),
            "length_axis_global": [0.0, -1.0, 0.0],
            "passed": PASSED,
        }

    metadata = {
        "enabled": True,
        "point_names": list(PURLIN_AXIS_POINT_NAMES),
        "short_length_m": PURLIN_SHORT_LENGTH_M,
        "purlin_source_part_name": purlin_component["part_name"],
        "purlin_derived_part_name": derived_part_name,
        "purlin_support_part_name": purlin_support_component["part_name"],
        "axis_unit": u,
        "normal_unit": n,
        "purlin_height_m": purlin_h,
        "purlin_width_m": purlin_b,
        "group_y_offset_m": PURLIN_GROUP_Y_OFFSET_M,
        "placement": "C-channel upper flange center at S/P/Q/R; member length axes parallel global Y.",
    }
    return members, metadata, checks, warnings


def _member(
    name: str,
    phase: str,
    component: dict[str, Any],
    local_anchor: list[float],
    global_anchor_name: str,
    points: dict[str, dict[str, Any]],
    rotate_y_deg: float,
    target_point_name: str | None = None,
    roll_about_axis_deg: float = 0.0,
) -> dict[str, Any]:
    global_anchor = _coords(points[global_anchor_name])
    section_reference = _section_reference_xy(component)
    open_side_local = [1.0, 0.0, 0.0] if component.get("section_kind") == "C_CHANNEL" else [0.0, 0.0, 0.0]
    open_side_global = _clean_vector(transform_local(open_side_local, rotate_y_deg, roll_about_axis_deg)) if open_side_local != [0.0, 0.0, 0.0] else [0.0, 0.0, 0.0]
    member = {
        "name": name,
        "phase": phase,
        "part_name": component["part_name"],
        "component_code": component.get("component_code"),
        "instance_name": _instance_name(component),
        "local_anchor": local_anchor,
        "global_anchor_name": global_anchor_name,
        "global_anchor": global_anchor,
        "rotate_y_deg": rotate_y_deg,
        "roll_about_axis_deg": roll_about_axis_deg,
        "translation": _translation_for_anchor(local_anchor, global_anchor, rotate_y_deg, roll_about_axis_deg),
        "part_length_m": component.get("length_m"),
        "section_kind": component.get("section_kind"),
        "section_params_m": component.get("section_params_m") or {},
        "section_reference": section_reference,
        "open_side_global": open_side_global,
        "model_policy": component.get("model_policy"),
    }
    if target_point_name:
        member["target_point_name"] = target_point_name
        member["target_point"] = _coords(points[target_point_name])
    return member


def build_payload(
    excel_path: str | Path,
    components_path: str | Path,
    project_code: str = DEFAULT_PROJECT_CODE,
    model_name: str | None = None,
) -> dict[str, Any]:
    model_name = model_name or project_code
    inputs, cached_points, cached_checks = read_excel_inputs(excel_path)
    components = load_components(components_path)
    main_components = {code: _component(components, code) for code in MAIN_COMPONENT_CODES}
    column_down_component = _optional_component(components, "COLUMN_DOWN")
    column_up_component = _optional_component(components, "COLUMN_UP")
    single_column_component = _optional_component(components, "COLUMN")
    purlin_component = _optional_component(components, "PURLIN")
    purlin_support_component = _optional_component(components, "PURLIN_SUPPORT")
    if not (column_down_component and column_up_component) and not single_column_component and not column_up_component:
        raise ValueError("Missing column components. Expected COLUMN_DOWN+COLUMN_UP, COLUMN, or COLUMN_UP in components JSON.")
    beam_component = main_components["INCLINED_BEAM"]
    beam_length = beam_component.get("length_m")

    points = solve_points_from_inputs(inputs)
    checks = build_checks(inputs, points, beam_length)
    warnings = compare_cached_points(points, cached_points) + compare_cached_checks(checks, cached_checks)
    errors: list[str] = []

    theta_deg = _value(inputs, "theta_deg")
    theta_rad = math.radians(theta_deg)
    gc = _length_m(inputs, "GC_mm")
    gf = _length_m(inputs, "GF_mm")
    ge = _length_m(inputs, "GE_mm")
    control_tolerance = _length_m(inputs, "control_tolerance_m")
    angle_tolerance = _value(inputs, "angle_tolerance_deg")
    rotate_beam_y = 90.0 - theta_deg
    u = [math.cos(theta_rad), 0.0, math.sin(theta_rad)]
    n = [-math.sin(theta_rad), 0.0, math.cos(theta_rad)]
    beam_roll = _default_roll_about_axis_deg("INCLINED_BEAM", beam_component)
    front_brace_roll = _default_roll_about_axis_deg("BRACE_FRONT", main_components["BRACE_FRONT"])
    rear_brace_roll = _default_roll_about_axis_deg("BRACE_REAR", main_components["BRACE_REAR"])

    for name, row in inputs.items():
        if row.status != CONFIRMED:
            warnings.append("Input %s status is %s." % (name, row.status or "empty"))
    if control_tolerance > 0.001 + 1.0e-12:
        warnings.append("control_tolerance_m is %.6f m; this is wider than +/-1 mm." % control_tolerance)
    if checks["GC_GF_GE_ORDER"]["passed"] != PASSED:
        errors.append("Beam stations must satisfy 0 < GC < GF < GE < beam length.")

    column_members: list[dict[str, Any]] = []
    member_checks: dict[str, dict[str, Any]] = {}
    if column_down_component and column_up_component:
        column_members.extend(
            [
                _member(
                    "COLUMN_DOWN",
                    "step01_columns",
                    column_down_component,
                    _local_reference_point(column_down_component, 0.0),
                    "O",
                    points,
                    0.0,
                ),
                _member(
                    "COLUMN_UP",
                    "step01_columns",
                    column_up_component,
                    _local_reference_point(column_up_component, float(column_up_component.get("length_m") or 0.0)),
                    "A",
                    points,
                    0.0,
                ),
            ]
        )
        column_up_bottom = _add(_coords(points["A"]), [0.0, 0.0, -float(column_up_component.get("length_m") or 0.0)])
        column_down_top = _add(_coords(points["O"]), [0.0, 0.0, float(column_down_component.get("length_m") or 0.0)])
        member_checks.update(
            {
                "COLUMN_DOWN_PLACEMENT": {
                    "anchor": "O",
                    "part_length_m": column_down_component.get("length_m"),
                    "derived_top": column_down_top,
                    "passed": PASSED,
                    "note": "Only the bottom point O is controlled in step01.",
                },
                "COLUMN_UP_PLACEMENT": {
                    "anchor": "A",
                    "part_length_m": column_up_component.get("length_m"),
                    "derived_bottom": column_up_bottom,
                    "passed": PASSED,
                    "note": "Only the top point A is controlled in step01.",
                },
            }
        )
    elif single_column_component is not None:
        assert single_column_component is not None
        column_members.append(
            _member(
                "COLUMN",
                "step01_columns",
                single_column_component,
                _local_reference_point(single_column_component, float(single_column_component.get("length_m") or 0.0)),
                "A",
                points,
                0.0,
            )
        )
        column_bottom = _add(_coords(points["A"]), [0.0, 0.0, -float(single_column_component.get("length_m") or 0.0)])
        member_checks["COLUMN_PLACEMENT"] = {
            "anchor": "A",
            "part_length_m": single_column_component.get("length_m"),
            "derived_bottom": column_bottom,
            "passed": PASSED,
            "note": "Single COLUMN component is controlled by its top point A.",
        }
    else:
        assert column_up_component is not None
        column_members.append(
            _member(
                "COLUMN_UP",
                "step01_columns",
                column_up_component,
                _local_reference_point(column_up_component, float(column_up_component.get("length_m") or 0.0)),
                "A",
                points,
                0.0,
            )
        )
        column_up_bottom = _add(_coords(points["A"]), [0.0, 0.0, -float(column_up_component.get("length_m") or 0.0)])
        member_checks["COLUMN_UP_PLACEMENT"] = {
            "anchor": "A",
            "part_length_m": column_up_component.get("length_m"),
            "derived_bottom": column_up_bottom,
            "passed": PASSED,
            "note": "Only COLUMN_UP is available; it is controlled as a single upper column by top point A.",
        }

    members = [
        *column_members,
        _member(
            "INCLINED_BEAM",
            "step02_beam",
            beam_component,
            _local_reference_point(beam_component, gf),
            "F",
            points,
            rotate_beam_y,
            target_point_name="E",
            roll_about_axis_deg=beam_roll,
        ),
        _member(
            "BRACE_FRONT",
            "step03_main_frame",
            main_components["BRACE_FRONT"],
            _local_reference_point(main_components["BRACE_FRONT"], 0.0),
            "B",
            points,
            rotate_y_for_local_z_to_vector(_coords(points["B"]), _coords(points["C"])),
            target_point_name="C",
            roll_about_axis_deg=front_brace_roll,
        ),
        _member(
            "BRACE_REAR",
            "step03_main_frame",
            main_components["BRACE_REAR"],
            _local_reference_point(main_components["BRACE_REAR"], 0.0),
            "D",
            points,
            rotate_y_for_local_z_to_vector(_coords(points["D"]), _coords(points["E"])),
            target_point_name="E",
            roll_about_axis_deg=rear_brace_roll,
        ),
    ]

    member_checks.update(
        {
            "INCLINED_BEAM_CE": {
                "start": "C",
                "end": "E",
                "axis_length_m": distance(_coords(points["C"]), _coords(points["E"])),
                "part_length_m": beam_length,
                "error_m": None,
                "tolerance_m": control_tolerance,
                "passed": PASSED,
            },
            "BRACE_FRONT_BC": {
                **member_length_check(points, "B", "C", main_components["BRACE_FRONT"].get("length_m"), control_tolerance),
                "passed": checks["BC"]["passed"],
                "drawing_length_check": checks["BC"],
                "note": "Abaqus placement uses B-C axis; material length can include connection offsets.",
            },
            "BRACE_REAR_DE": {
                **member_length_check(points, "D", "E", main_components["BRACE_REAR"].get("length_m"), control_tolerance),
                "passed": checks["DE"]["passed"],
                "drawing_length_check": checks["DE"],
                "note": "Abaqus placement uses D-E axis; material length can include connection offsets.",
            },
        }
    )
    purlin_axis_metadata: dict[str, Any] = {"enabled": False, "reason": "PURLIN/PURLIN_SUPPORT components or S/P/Q/R points are not available."}
    purlin_points_available = all(name in points for name in PURLIN_AXIS_POINT_NAMES)
    if purlin_component and purlin_support_component and purlin_points_available:
        purlin_members, purlin_axis_metadata, purlin_checks, purlin_warnings = _build_purlin_members(
            purlin_component,
            purlin_support_component,
            points,
            theta_deg,
            u,
            n,
        )
        members.extend(purlin_members)
        member_checks.update(purlin_checks)
        warnings.extend(purlin_warnings)
    elif purlin_component or purlin_support_component:
        missing = []
        if not purlin_component:
            missing.append("PURLIN")
        if not purlin_support_component:
            missing.append("PURLIN_SUPPORT")
        if not purlin_points_available:
            missing.append("S/P/Q/R")
        warnings.append("Purlin assembly skipped because %s is unavailable." % ", ".join(missing))

    beam_reference = _section_reference_xy(beam_component)
    beam_local_point = _local_reference_point(beam_component, gf)
    beam_local_origin = _local_reference_point(beam_component, 0.0)
    required_part_names = sorted({member.get("source_part_name") or member["part_name"] for member in members})
    payload = {
        "meta": {
            "project_code": project_code,
            "model_name": model_name,
            "source_excel": str(Path(excel_path).as_posix()),
            "source_components": str(Path(components_path).as_posix()),
            "coordinate_system": "X right, Y out of elevation plane, Z up; units m-kg-N-Pa",
        },
        "units": {
            "length": "m",
            "mass": "kg",
            "force": "N",
            "stress": "Pa",
            "density": "kg/m^3",
        },
        "inputs": {
            "theta_deg": theta_deg,
            "theta_rad": theta_rad,
            "GC_m": gc,
            "GF_m": gf,
            "GE_m": ge,
            "control_tolerance_m": control_tolerance,
            "angle_tolerance_deg": angle_tolerance,
            "beam_length_m": beam_length,
            "pv_axis_angle_tolerance_deg": _value(inputs, "pv_axis_angle_tolerance_deg") if inputs.get("pv_axis_angle_tolerance_deg") is not None else None,
            "purlin_short_length_m": PURLIN_SHORT_LENGTH_M,
        },
        "input_rows": {
            name: {
                "meaning": row.meaning,
                "value": row.value,
                "unit": row.unit,
                "status": row.status,
                "note": row.note,
                "excel_row": row.row,
            }
            for name, row in inputs.items()
        },
        "points": points,
        "beam_anchor": {
            "part_name": beam_component["part_name"],
            "local_point_name": "F",
            "local_point": beam_local_point,
            "axis_local_point": [0.0, 0.0, gf],
            "reference_local_origin": beam_local_origin,
            "global_point_name": "F",
            "global_point": _coords(points["F"]),
            "stations": {"C": gc, "F": gf, "E": ge},
            "direction_u": u,
            "rotate_y_deg": rotate_beam_y,
            "roll_about_axis_deg": beam_roll,
            "translation": _translation_for_anchor(beam_local_point, _coords(points["F"]), rotate_beam_y, beam_roll),
            "section_reference": beam_reference,
            "section_sets": {"C": "SET_BEAM_SEC_C", "F": "SET_BEAM_SEC_F", "E": "SET_BEAM_SEC_E"},
        },
        "purlin_axis": purlin_axis_metadata,
        "members": members,
        "required_part_names": required_part_names,
        "checks": checks,
        "member_checks": member_checks,
        "warnings": warnings,
        "errors": errors,
    }
    return payload


def write_json(payload: dict[str, Any], output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def generate_abaqus_scripts(
    payload: dict[str, Any],
    output_dir: str | Path,
    json_path: str | Path | None,
    reports_dir: str | Path | None = None,
) -> list[Path]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    report_dir = Path(reports_dir) if reports_dir else out_dir
    report_dir.mkdir(parents=True, exist_ok=True)
    project = payload["meta"]["project_code"]
    scripts = [
        ("full_main_frame", "%s_assembly_frame.py" % project),
    ]
    paths: list[Path] = []
    for phase, filename in scripts:
        script = build_abaqus_script(
            phase=phase,
            payload=payload,
            report_path=report_dir / ("%s_%s_report.json" % (project, phase)),
            cae_path=report_dir / ("%s_%s.cae" % (project, phase)),
        )
        path = out_dir / filename
        path.write_text(script, encoding="utf-8")
        paths.append(path)
    return paths


def build_abaqus_script(
    phase: str,
    payload: dict[str, Any],
    report_path: Path,
    cae_path: Path,
) -> str:
    return Template(ABAQUS_SCRIPT_TEMPLATE).substitute(
        phase=phase,
        assembly_payload_json=json.dumps(payload, ensure_ascii=False, indent=2),
        report_filename=Path(report_path).name,
        cae_filename=Path(cae_path).name,
        script_name="%s.py" % phase,
    )


ABAQUS_SCRIPT_TEMPLATE = r'''# -*- coding: utf-8 -*-
"""Abaqus Assembly stage script generated from coordinate Excel.

Run inside Abaqus/CAE after the five main Parts already exist in the model.
For example:
    abaqus cae noGUI=$script_name
"""
from __future__ import print_function

import codecs
import json
import math
import os

from abaqus import mdb
from abaqusConstants import *
import mesh
import regionToolset


PHASE = "$phase"
ASSEMBLY_DATA = json.loads(r"""$assembly_payload_json""")


def _resolve_report_path(filename):
    candidates = []
    try:
        script_folder = os.path.dirname(os.path.abspath(__file__))
        candidates.append(os.path.join(script_folder, "..", "reports"))
        candidates.append(script_folder)
    except Exception:
        pass
    candidates.append(os.path.join(os.getcwd(), "reports"))
    candidates.append(os.getcwd())
    for candidate in candidates:
        if candidate:
            try:
                if os.path.exists(candidate):
                    return os.path.join(candidate, filename)
            except Exception:
                pass
    return os.path.join(os.getcwd(), filename)


REPORT_PATH = _resolve_report_path("$report_filename")
SAVE_AS_PATH = r""
SUGGESTED_SAVE_AS_PATH = _resolve_report_path("$cae_filename")
DEFAULT_BEAM_SECTION_SET_NAMES = ("SET_BEAM_SEC_C", "SET_BEAM_SEC_F", "SET_BEAM_SEC_E")


try:
    unicode
except NameError:
    unicode = str


def _ascii(value):
    if value is None:
        return ""
    if isinstance(value, unicode):
        return value.encode("ascii")
    return str(value)


def _float_or_none(value):
    if value is None or str(value).strip() == "":
        return None
    return float(value)


def _ensure_parent(path):
    folder = os.path.dirname(os.path.abspath(path))
    if folder and not os.path.exists(folder):
        os.makedirs(folder)


def _model(data):
    project_code = _ascii(data.get("meta", {}).get("project_code") or "")
    model_name = _ascii(data.get("meta", {}).get("model_name") or project_code)
    if project_code and model_name != project_code:
        raise RuntimeError("Assembly model name %s does not match project prefix %s." % (model_name, project_code))
    if model_name not in mdb.models:
        raise RuntimeError("Project model %s not found. Run %s_create_parts_in_cae.py first." % (model_name, project_code or model_name))
    return mdb.models[model_name]


def _point(data, name):
    return tuple(float(v) for v in data["points"][name]["coords"])


def _distance(a, b):
    dx = a[0] - b[0]
    dy = a[1] - b[1]
    dz = a[2] - b[2]
    return math.sqrt(dx * dx + dy * dy + dz * dz)


def _rotate_x(point, angle_deg):
    angle = math.radians(float(angle_deg))
    c = math.cos(angle)
    s = math.sin(angle)
    x, y, z = point
    return (x, y * c - z * s, y * s + z * c)


def _rotate_y(point, angle_deg):
    angle = math.radians(float(angle_deg))
    c = math.cos(angle)
    s = math.sin(angle)
    x, y, z = point
    return (x * c + z * s, y, -x * s + z * c)


def _rotate_z(point, angle_deg):
    angle = math.radians(float(angle_deg))
    c = math.cos(angle)
    s = math.sin(angle)
    x, y, z = point
    return (x * c - y * s, x * s + y * c, z)


def _apply_rotation_sequence(point, rotation_sequence):
    result = tuple(float(v) for v in point)
    for rotation in rotation_sequence or ():
        axis = str(rotation.get("axis") or "").upper()
        angle = float(rotation.get("angle_deg") or 0.0)
        if axis == "X":
            result = _rotate_x(result, angle)
        elif axis == "Y":
            result = _rotate_y(result, angle)
        elif axis == "Z":
            result = _rotate_z(result, angle)
        else:
            raise RuntimeError("Unsupported rotation axis for %s: %s" % (rotation, axis))
    return result


def _rotation_axis_direction(axis):
    axis = str(axis or "").upper()
    if axis == "X":
        return (1.0, 0.0, 0.0)
    if axis == "Y":
        return (0.0, 1.0, 0.0)
    if axis == "Z":
        return (0.0, 0.0, 1.0)
    raise RuntimeError("Unsupported rotation axis: %s" % axis)


def _add3(a, b):
    return (float(a[0]) + float(b[0]), float(a[1]) + float(b[1]), float(a[2]) + float(b[2]))


def _sub3(a, b):
    return (float(a[0]) - float(b[0]), float(a[1]) - float(b[1]), float(a[2]) - float(b[2]))


def _dot(a, b):
    return float(a[0]) * float(b[0]) + float(a[1]) * float(b[1]) + float(a[2]) * float(b[2])


def _cross(a, b):
    return (
        float(a[1]) * float(b[2]) - float(a[2]) * float(b[1]),
        float(a[2]) * float(b[0]) - float(a[0]) * float(b[2]),
        float(a[0]) * float(b[1]) - float(a[1]) * float(b[0]),
    )


def _unit(vector):
    length = math.sqrt(_dot(vector, vector))
    if length <= 1.0e-15:
        return (0.0, 0.0, 1.0)
    return (float(vector[0]) / length, float(vector[1]) / length, float(vector[2]) / length)


def _rotate_about_axis(point, axis_point, axis_direction, angle_deg):
    axis = _unit(axis_direction)
    rel = _sub3(point, axis_point)
    angle = math.radians(float(angle_deg))
    c = math.cos(angle)
    s = math.sin(angle)
    cross = _cross(axis, rel)
    along = _dot(axis, rel)
    rotated = (
        rel[0] * c + cross[0] * s + axis[0] * along * (1.0 - c),
        rel[1] * c + cross[1] * s + axis[1] * along * (1.0 - c),
        rel[2] * c + cross[2] * s + axis[2] * along * (1.0 - c),
    )
    return _add3(axis_point, rotated)


def _member_axis_direction(member):
    if member.get("rotation_sequence"):
        return _unit(_apply_rotation_sequence((0.0, 0.0, 1.0), member.get("rotation_sequence")))
    return _unit(_rotate_y((0.0, 0.0, 1.0), float(member.get("rotate_y_deg") or 0.0)))


def _transform_member(local_point, member):
    if member.get("rotation_sequence"):
        rotated = _apply_rotation_sequence(tuple(float(v) for v in local_point), member.get("rotation_sequence"))
        return _add3(rotated, tuple(float(v) for v in member.get("translation", (0.0, 0.0, 0.0))))
    rotated = _rotate_y(tuple(float(v) for v in local_point), float(member.get("rotate_y_deg") or 0.0))
    translated = _add3(rotated, tuple(float(v) for v in member.get("translation", (0.0, 0.0, 0.0))))
    roll_about_axis_deg = float(member.get("roll_about_axis_deg") or 0.0)
    if abs(roll_about_axis_deg) <= 1.0e-12:
        return translated
    axis_point = tuple(float(v) for v in member.get("global_anchor", (0.0, 0.0, 0.0)))
    return _rotate_about_axis(translated, axis_point, _member_axis_direction(member), roll_about_axis_deg)


def _ensure_material(model, material):
    mat_name = _ascii((material or {}).get("abaqus_name") or "MAT_MANUAL_CHECK")
    if mat_name in model.materials:
        return mat_name
    mat = model.Material(name=mat_name)
    elastic_modulus = (material or {}).get("elastic_modulus_pa")
    poisson_ratio = (material or {}).get("poisson_ratio")
    if elastic_modulus is not None and poisson_ratio is not None:
        mat.Elastic(table=((float(elastic_modulus), float(poisson_ratio)),))
    density = (material or {}).get("density_kg_per_m3")
    if density is not None:
        mat.Density(table=((float(density),),))
    return mat_name


def _profile_points(component):
    params = component.get("section_params_m") or {}
    kind = component.get("section_kind")
    if kind == "C_CHANNEL":
        h = float(params["h_m"])
        b = float(params["b_m"])
        lip = float(params["lip_m"])
        return [(b, lip), (b, 0.0), (0.0, 0.0), (0.0, h), (b, h), (b, h - lip)]
    raise RuntimeError("Can only derive 50mm purlin from C_CHANNEL, got %s." % kind)


def _create_shell_part(model, component):
    part_name = _ascii(component["part_name"])
    if part_name in model.parts:
        return model.parts[part_name]

    length = _float_or_none(component.get("length_m")) or 0.05
    thickness = _float_or_none(component.get("thickness_m"))
    if thickness is None:
        raise RuntimeError("Shell Part %s requires thickness_m." % part_name)

    sketch = model.ConstrainedSketch(name=_ascii("SK_" + str(component["part_name"])), sheetSize=max(length, 1.0) * 2.0)
    points = _profile_points(component)
    for start, end in zip(points[:-1], points[1:]):
        sketch.Line(point1=start, point2=end)

    part = model.Part(name=part_name, dimensionality=THREE_D, type=DEFORMABLE_BODY)
    part.BaseShellExtrude(sketch=sketch, depth=length)

    material_name = _ensure_material(model, component.get("material", {}))
    section_name = _ascii("SEC_" + str(component["part_name"]))
    if section_name not in model.sections:
        model.HomogeneousShellSection(
            name=section_name,
            preIntegrate=OFF,
            material=material_name,
            thicknessType=UNIFORM,
            thickness=thickness,
        )
    region = regionToolset.Region(faces=part.faces[:])
    part.SectionAssignment(region=region, sectionName=section_name)
    elem_type = mesh.ElemType(elemCode=S4R, elemLibrary=STANDARD)
    part.seedPart(size=0.02, deviationFactor=0.1, minSizeFactor=0.1)
    part.setElementType(regions=(part.faces[:],), elemTypes=(elem_type,))
    part.generateMesh()
    return part


def _effective_part(model, member):
    part_component = member.get("part_component")
    if part_component:
        source_name = member.get("source_part_name")
        if source_name:
            _part(model, source_name)
        component = dict(part_component)
        component["part_name"] = member["part_name"]
        component["length_m"] = float(member.get("part_length_m") or component.get("length_m") or 0.05)
        return _create_shell_part(model, component)
    return _part(model, member["part_name"])


def _part(model, name):
    key = _ascii(name)
    if key not in model.parts:
        raise RuntimeError("Missing Part %s. Run the generated Part creation script first." % name)
    return model.parts[key]


def _delete_instance(assembly, name):
    key = _ascii(name)
    if key in assembly.instances:
        try:
            del assembly.features[key]
        except Exception:
            try:
                del assembly.instances[key]
            except Exception:
                pass


def _delete_set(container, name):
    key = _ascii(name)
    try:
        if key in container.sets:
            del container.sets[key]
    except Exception:
        pass


def _edges_at_station(part, station):
    # Prefer Abaqus EdgeArray selection. Creating a Set from a plain Python
    # list of Edge objects is less reliable in Abaqus/CAE 2020.
    for tol in (1.0e-7, 1.0e-6, 1.0e-5, 1.0e-4, 1.0e-3):
        try:
            edges = part.edges.getByBoundingBox(
                xMin=-1000.0,
                yMin=-1000.0,
                zMin=station - tol,
                xMax=1000.0,
                yMax=1000.0,
                zMax=station + tol,
            )
            if len(edges):
                return edges, tol, "bounding_box"
        except Exception:
            pass

    found = []
    for edge in part.edges:
        try:
            point = edge.pointOn[0]
            if abs(float(point[2]) - station) <= 1.0e-5:
                found.append(edge)
        except Exception:
            pass
    return tuple(found), 1.0e-5, "point_on"


def _instance(model, member):
    assembly = model.rootAssembly
    inst_name = _ascii(member["instance_name"])
    _delete_instance(assembly, inst_name)
    part = _effective_part(model, member)
    assembly.Instance(name=inst_name, part=part, dependent=ON)
    rotation_sequence = member.get("rotation_sequence") or []
    rotate_y_deg = float(member.get("rotate_y_deg") or 0.0)
    if rotation_sequence:
        for rotation in rotation_sequence:
            angle = float(rotation.get("angle_deg") or 0.0)
            if abs(angle) <= 1.0e-12:
                continue
            assembly.rotate(
                instanceList=(inst_name,),
                axisPoint=(0.0, 0.0, 0.0),
                axisDirection=_rotation_axis_direction(rotation.get("axis")),
                angle=angle,
            )
    elif abs(rotate_y_deg) > 1.0e-12:
        assembly.rotate(
            instanceList=(inst_name,),
            axisPoint=(0.0, 0.0, 0.0),
            axisDirection=(0.0, 1.0, 0.0),
            angle=rotate_y_deg,
        )
    translation = tuple(float(v) for v in member.get("translation", (0.0, 0.0, 0.0)))
    if max(abs(translation[0]), abs(translation[1]), abs(translation[2])) > 1.0e-12:
        assembly.translate(instanceList=(inst_name,), vector=translation)
    roll_about_axis_deg = float(member.get("roll_about_axis_deg") or 0.0)
    if not rotation_sequence and abs(roll_about_axis_deg) > 1.0e-12:
        axis_point = tuple(float(v) for v in member.get("global_anchor", (0.0, 0.0, 0.0)))
        assembly.rotate(
            instanceList=(inst_name,),
            axisPoint=axis_point,
            axisDirection=_member_axis_direction(member),
            angle=roll_about_axis_deg,
        )
    return {
        "instance_name": member["instance_name"],
        "part_name": member["part_name"],
        "source_part_name": member.get("source_part_name"),
        "rotate_y_deg": rotate_y_deg,
        "rotation_sequence": rotation_sequence,
        "roll_about_axis_deg": roll_about_axis_deg,
        "translation": list(translation),
        "section_reference": member.get("section_reference"),
        "open_side_global": member.get("open_side_global"),
        "placement_note": member.get("placement_note"),
    }


def _partition_beam(model, data):
    beam = data["beam_anchor"]
    part = _part(model, beam["part_name"])
    report = {"part_name": beam["part_name"], "stations": {}, "sets": [], "warnings": []}
    stations = beam.get("stations", {})
    for label in ("C", "F", "E"):
        if label not in stations:
            continue
        station = float(stations[label])
        datum = part.DatumPlaneByPrincipalPlane(principalPlane=XYPLANE, offset=station)
        try:
            part.PartitionFaceByDatumPlane(datumPlane=part.datums[datum.id], faces=part.faces[:])
        except Exception as exc:
            report["warnings"].append("Partition at %s=%.6f failed or already exists: %s" % (label, station, exc))

        set_name = _ascii(beam.get("section_sets", {}).get(label) or ("SET_BEAM_SEC_" + label))
        if set_name in part.sets:
            report["sets"].append(set_name)
            report["warnings"].append("Set %s already exists; reused it." % set_name)
            report["stations"][label] = station
            continue

        edges, edge_tol, edge_method = _edges_at_station(part, station)
        if edges:
            try:
                part.Set(edges=edges, name=set_name)
                report["sets"].append(set_name)
                report["warnings"].append("Created %s using %s tolerance %.1e." % (set_name, edge_method, edge_tol))
            except Exception as exc:
                report["warnings"].append("Could not create edge set %s at station %.6f: %s" % (set_name, station, exc))
        else:
            report["warnings"].append("No partition edge found for %s at station %.6f." % (set_name, station))
        report["stations"][label] = station

    return report


def _members_for_phase(data):
    if PHASE == "step01_columns":
        names = set(["COLUMN_DOWN", "COLUMN_UP", "COLUMN"])
    elif PHASE == "step02_beam":
        names = set(["INCLINED_BEAM"])
    elif PHASE == "step03_main_frame":
        names = set(["COLUMN_DOWN", "COLUMN_UP", "COLUMN", "INCLINED_BEAM", "BRACE_FRONT", "BRACE_REAR"])
    elif PHASE == "step04_purlins":
        names = set(member.get("name") for member in data.get("members", []) if member.get("phase") == "step04_purlins")
    else:
        return list(data.get("members", []))
    return [member for member in data.get("members", []) if member.get("name") in names]


def _transform_member_vector(local_vector, member):
    if member.get("rotation_sequence"):
        return _unit(_apply_rotation_sequence(tuple(float(v) for v in local_vector), member.get("rotation_sequence")))
    rotated = _rotate_y(_rotate_z(tuple(float(v) for v in local_vector), float(member.get("roll_about_axis_deg") or 0.0)), float(member.get("rotate_y_deg") or 0.0))
    return _unit(rotated)


def _validate_member(member, data):
    errors = []
    anchor = _transform_member(member.get("local_anchor", (0.0, 0.0, 0.0)), member)
    target_anchor = tuple(float(v) for v in member.get("global_anchor", (0.0, 0.0, 0.0)))
    anchor_error = _distance(anchor, target_anchor)
    if anchor_error > 1.0e-6:
        errors.append("%s anchor error %.9g m" % (member["name"], anchor_error))
    axis_validation = {}
    for check in member.get("axis_checks", []):
        actual = _transform_member_vector(check.get("local_vector", (0.0, 0.0, 1.0)), member)
        expected = _unit(tuple(float(v) for v in check.get("expected_global", (0.0, 0.0, 1.0))))
        error = _distance(actual, expected)
        tolerance = float(check.get("tolerance") or 1.0e-6)
        if error > tolerance:
            errors.append("%s %s axis error %.9g" % (member["name"], check.get("name"), error))
        axis_validation[str(check.get("name"))] = {
            "actual": list(actual),
            "expected": list(expected),
            "error": error,
            "tolerance": tolerance,
        }

    if member.get("target_point_name"):
        part_length = float(member.get("part_length_m") or 0.0)
        local_anchor = member.get("local_anchor", (0.0, 0.0, 0.0))
        local_end = (float(local_anchor[0]), float(local_anchor[1]), part_length)
        transformed_end = _transform_member(local_end, member)
        target = tuple(float(v) for v in member.get("target_point", (0.0, 0.0, 0.0)))
        # Braces may include connection offsets, so report this value rather than failing hard.
        end_error = _distance(transformed_end, target)
        return {"anchor_error_m": anchor_error, "end_error_m": end_error, "axis_validation": axis_validation, "errors": errors}
    return {"anchor_error_m": anchor_error, "axis_validation": axis_validation, "errors": errors}


def _validate_beam(data):
    beam_member = None
    for member in data.get("members", []):
        if member.get("name") == "INCLINED_BEAM":
            beam_member = member
            break
    if not beam_member:
        return {}
    stations = data["beam_anchor"]["stations"]
    beam_local_origin = data["beam_anchor"].get("reference_local_origin") or [0.0, 0.0, 0.0]
    ref_x = float(beam_local_origin[0])
    ref_y = float(beam_local_origin[1])
    validation = {}
    for label in ("C", "F", "E"):
        actual = _transform_member((ref_x, ref_y, float(stations[label])), beam_member)
        expected = _point(data, label)
        validation[label] = {"actual": list(actual), "expected": list(expected), "error_m": _distance(actual, expected)}
    actual_g = _transform_member((ref_x, ref_y, 0.0), beam_member)
    expected_g = _point(data, "G_global")
    validation["G_global"] = {"actual": list(actual_g), "expected": list(expected_g), "error_m": _distance(actual_g, expected_g)}
    return validation


def main():
    data = ASSEMBLY_DATA
    model = _model(data)
    assembly = model.rootAssembly
    report = {"phase": PHASE, "warnings": list(data.get("warnings", [])), "instances": [], "partition": None, "validation": {}}
    phase_members = _members_for_phase(data)

    missing = []
    for name in sorted(set(member.get("source_part_name") or member["part_name"] for member in phase_members)):
        if _ascii(name) not in model.parts:
            missing.append(name)
    if missing:
        project = data.get("meta", {}).get("project_code") or "project"
        raise RuntimeError("Missing required Parts for %s: %s. Run %s_create_parts_in_cae.py first." % (PHASE, ", ".join(missing), project))

    if PHASE in ("step02_beam", "step03_main_frame", "full_main_frame"):
        report["partition"] = _partition_beam(model, data)

    for member in phase_members:
        report["instances"].append(_instance(model, member))
        report["validation"][member["name"]] = _validate_member(member, data)

    if PHASE in ("step02_beam", "step03_main_frame", "full_main_frame"):
        report["beam_station_validation"] = _validate_beam(data)

    assembly.regenerate()

    _ensure_parent(REPORT_PATH)
    with codecs.open(REPORT_PATH, "w", "utf-8") as handle:
        json.dump(report, handle, ensure_ascii=False, indent=2)

    if SAVE_AS_PATH:
        mdb.saveAs(pathName=SAVE_AS_PATH)

    print("Assembly phase %s completed." % PHASE)
    print("Report: %s" % REPORT_PATH)
    print("Suggested save path, if needed: %s" % SUGGESTED_SAVE_AS_PATH)


main()
'''


def export_main_frame_assembly(
    excel_path: str | Path,
    components_path: str | Path,
    output_json: str | Path | None,
    output_dir: str | Path,
    reports_dir: str | Path | None = None,
    project_code: str = DEFAULT_PROJECT_CODE,
    model_name: str | None = None,
) -> tuple[Path | None, list[Path], dict[str, Any]]:
    payload = build_payload(excel_path, components_path, project_code=project_code, model_name=model_name)
    json_file = write_json(payload, output_json) if output_json is not None else None
    scripts = generate_abaqus_scripts(payload, output_dir, json_file, reports_dir=reports_dir)
    return json_file, scripts, payload


def _prefix_from_latest(outputs_root: str | Path) -> str | None:
    latest = Path(outputs_root) / "latest_run_manifest.json"
    if not latest.exists():
        return None
    payload = json.loads(latest.read_text(encoding="utf-8"))
    return payload.get("project_prefix") or payload.get("project_code")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export main-frame Assembly JSON and Abaqus stage scripts.")
    parser.add_argument("--excel", default=None)
    parser.add_argument("--components", default=None)
    parser.add_argument("--out-json", default=None)
    parser.add_argument("--out-dir", default=None, help="Abaqus script output directory. Defaults to run_dir/abaqus_scripts.")
    parser.add_argument("--reports-dir", default=None, help="Report output directory. Defaults to run_dir/reports.")
    parser.add_argument("--run-dir", default=None, help="Existing run directory. If omitted, a timestamped run directory is created.")
    parser.add_argument("--outputs-root", default="outputs", help="Root directory for timestamped runs.")
    parser.add_argument("--project-code", "--project-prefix", dest="project_code", default=None)
    parser.add_argument("--model-name", default=None, help="Abaqus model name. Defaults to <project_prefix>.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    project = args.project_code or _prefix_from_latest(args.outputs_root) or DEFAULT_PROJECT_CODE
    use_run_dir = args.run_dir is not None or args.out_json is None or args.out_dir is None
    run_paths = None
    if use_run_dir:
        run_paths = create_run_paths(project, outputs_root=args.outputs_root, run_dir=args.run_dir)
        out_json = Path(args.out_json) if args.out_json else run_paths.json / "assembly_inputs.json"
        out_dir = Path(args.out_dir) if args.out_dir else run_paths.abaqus_scripts
        reports_dir = Path(args.reports_dir) if args.reports_dir else run_paths.reports
        excel = Path(args.excel) if args.excel else run_paths.workbooks / ("%s_coordinate_formula_simple_fixed.xlsx" % project)
        components = Path(args.components) if args.components else run_paths.json / "components.json"
    else:
        out_json = Path(args.out_json)
        out_dir = Path(args.out_dir)
        reports_dir = Path(args.reports_dir) if args.reports_dir else None
        if not args.excel or not args.components:
            raise ValueError("--excel and --components are required when not using a run directory.")
        excel = Path(args.excel)
        components = Path(args.components)

    model_name = args.model_name or project
    json_file, scripts, payload = export_main_frame_assembly(
        excel,
        components,
        out_json,
        out_dir,
        reports_dir=reports_dir,
        project_code=project,
        model_name=model_name,
    )
    if run_paths:
        report_outputs = {
            "assembly_frame_report": str((reports_dir / ("%s_full_main_frame_report.json" % project)).resolve())
        }
        update_manifest(
            run_paths,
            project,
            "assembly_scripts",
            inputs={
                "coordinate_excel": str(excel.resolve()),
                "components_json": str(components.resolve()),
            },
            outputs={
                "json": {"assembly_inputs": str(json_file.resolve())},
                "abaqus_scripts": {script.stem: str(script.resolve()) for script in scripts},
                "reports": report_outputs,
            },
            warnings=list(payload.get("warnings", [])),
            errors=list(payload.get("errors", [])),
            metadata={"model_name": model_name},
        )
    print("Wrote Assembly JSON: %s" % json_file.resolve())
    for script in scripts:
        print("Wrote Abaqus script: %s" % script.resolve())
    if run_paths:
        print("Run manifest: %s" % run_paths.manifest.resolve())
    if payload.get("warnings"):
        print("Warnings: %d" % len(payload["warnings"]))
    if payload.get("errors"):
        print("Errors: %d" % len(payload["errors"]))
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
