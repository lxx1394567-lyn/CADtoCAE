from __future__ import annotations

import json
import re
import shutil
import sys
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .standards import component_code_from_part_name, has_complete_model_dimensions, part_name_from_prefix, project_prefix
from .workbook import export_abaqus_json, read_component_rows_for_processing


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
if SCRIPTS_DIR.exists() and str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from make_cae_runner import generate_cae_runner  # noqa: E402


DEBUG_REPORT_SUBDIR = Path("过程文件") / "调试文件"


COMPONENT_SHEET = "建模构件表"
REQUIRED_HEADERS = ["支架类型", "角度", "构件名称", "规格", "长度_mm", "数量", "材料牌号", "建模方式", "abaqus_part_name"]
PROJECT_PREFIX_RE = re.compile(r"(SP_SC|SP_DC|DP)_ANG\d+(?:P\d+)?", re.IGNORECASE)


@dataclass
class PartScriptOutput:
    workbook_path: str
    status: str
    project_prefix: str
    project_dir: str
    copied_workbook_path: str | None
    components_json_path: str | None
    part_script_path: str | None
    cae_save_path: str | None
    report_path: str
    exported_count: int
    complete_count: int
    row_count: int
    messages: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _safe_name(value: str, max_len: int = 90) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value).strip(" ._")
    cleaned = re.sub(r"_+", "_", cleaned)
    return (cleaned or "workbook")[:max_len]


def project_prefix_from_path(path: str | Path) -> str | None:
    match = PROJECT_PREFIX_RE.search(Path(path).stem)
    if not match:
        match = PROJECT_PREFIX_RE.search(str(Path(path).parent.name))
    return match.group(0).upper() if match else None


def angle_from_project_prefix(project_prefix_value: str) -> str:
    match = re.search(r"_ANG(?P<angle>\d+(?:P\d+)?)", project_prefix_value, re.IGNORECASE)
    if not match:
        return ""
    return match.group("angle").replace("P", ".")


def read_component_rows(xlsx: str | Path, standards_path: str | Path | None = None) -> tuple[list[dict[str, Any]], list[str]]:
    workbook = load_workbook(xlsx, read_only=True)
    try:
        if COMPONENT_SHEET not in workbook.sheetnames:
            raise ValueError("Excel 缺少 sheet: %s" % COMPONENT_SHEET)
    finally:
        workbook.close()

    rows, headers = read_component_rows_for_processing(xlsx, standards_path)
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError("%s 缺少必要列: %s" % (COMPONENT_SHEET, ", ".join(missing)))
    return rows, headers


def infer_project_prefix_from_workbook(xlsx: str | Path, standards_path: str | Path | None = None) -> str:
    prefix_from_name = project_prefix_from_path(xlsx)
    if prefix_from_name:
        return prefix_from_name

    rows, _headers = read_component_rows(xlsx, standards_path)
    for row in rows:
        support_type = row.get("支架类型")
        angle = row.get("角度")
        if support_type and angle:
            return project_prefix(str(support_type), angle, standards_path)
    raise ValueError("无法从建模构件表推断项目名称前缀，请检查“支架类型”和“角度”列。")


def normalize_copied_workbook_prefix(xlsx: str | Path, project_prefix_value: str) -> None:
    workbook = load_workbook(xlsx)
    try:
        if COMPONENT_SHEET not in workbook.sheetnames:
            return
        worksheet = workbook[COMPONENT_SHEET]
        headers = [cell.value for cell in worksheet[1]]
        header_to_col = {header: index + 1 for index, header in enumerate(headers)}
        angle_col = header_to_col.get("角度")
        part_col = header_to_col.get("abaqus_part_name")
        code_col = header_to_col.get("构件代码")
        name_col = header_to_col.get("构件名称")
        angle = angle_from_project_prefix(project_prefix_value)
        for row_index in range(2, worksheet.max_row + 1):
            if angle_col and angle:
                worksheet.cell(row=row_index, column=angle_col).value = angle
            if part_col:
                part_cell = worksheet.cell(row=row_index, column=part_col)
                part_value = part_cell.value
                if isinstance(part_value, str) and part_value.startswith("="):
                    continue
                component_code = worksheet.cell(row=row_index, column=code_col).value if code_col else None
                component_code = component_code or component_code_from_part_name(part_value)
                if component_code:
                    part_cell.value = "P_%s_%s" % (project_prefix_value, str(component_code).strip().upper())
                elif not part_value and name_col:
                    component_name = worksheet.cell(row=row_index, column=name_col).value
                    if component_name:
                        part_cell.value = part_name_from_prefix(project_prefix_value, str(component_name).strip())
        workbook.save(xlsx)
    finally:
        workbook.close()


def workbook_quality_summary(xlsx: str | Path, standards_path: str | Path | None = None) -> dict[str, Any]:
    rows, _headers = read_component_rows(xlsx, standards_path)
    complete = []
    incomplete = []
    approved = []
    for row in rows:
        is_complete, issues = has_complete_model_dimensions(row)
        if is_complete:
            complete.append(row)
        else:
            incomplete.append({"component_name": row.get("构件名称"), "issues": issues})
        if str(row.get("校核状态", "")).strip().lower() in {"已确认", "approved"}:
            approved.append(row)
    return {
        "row_count": len(rows),
        "complete_count": len(complete),
        "approved_count": len(approved),
        "incomplete": incomplete,
    }


def _unique_project_dir(output_root: Path, project_prefix_value: str, workbook: Path, overwrite: bool = False) -> Path:
    base = output_root / ("%s_%s" % (project_prefix_value, _safe_name(workbook.stem)))
    if overwrite or not base.exists():
        return base
    for index in range(2, 1000):
        candidate = output_root / ("%s_%02d" % (base.name, index))
        if not candidate.exists():
            return candidate
    raise RuntimeError("Cannot allocate output directory under %s" % output_root)


def write_json_report(path: str | Path, payload: dict[str, Any]) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def step02_debug_report_path(output_root: str | Path, project_prefix_value: str) -> Path:
    return Path(output_root) / DEBUG_REPORT_SUBDIR / ("%s_step02_part_script_report.json" % project_prefix_value)


def generate_part_script_from_workbook(
    workbook_path: str | Path,
    output_root: str | Path,
    selection: str = "complete",
    standards_path: str | Path | None = None,
    project_prefix_value: str | None = None,
    cae_save_path: str | Path | None = None,
    overwrite: bool = False,
) -> PartScriptOutput:
    workbook = Path(workbook_path)
    messages: list[str] = []
    try:
        prefix = project_prefix_value or infer_project_prefix_from_workbook(workbook, standards_path)
        summary = workbook_quality_summary(workbook, standards_path)
        output_dir = Path(output_root)
        output_dir.mkdir(parents=True, exist_ok=True)
        report_path = step02_debug_report_path(output_dir, prefix)
        part_script: Path | None = None
        model_name = prefix
        with tempfile.TemporaryDirectory() as tmp:
            copied_workbook = Path(tmp) / workbook.name
            components_json = Path(tmp) / ("%s_components.json" % prefix)
            shutil.copy2(workbook, copied_workbook)
            normalize_copied_workbook_prefix(copied_workbook, prefix)
            exported_json = export_abaqus_json(copied_workbook, components_json, standards_path, selection=selection)
            payload = json.loads(Path(exported_json).read_text(encoding="utf-8"))
            exported_count = len(payload.get("components", []))
            if exported_count:
                part_script = output_dir / ("%s_create_parts_in_cae.py" % prefix)
                generate_cae_runner(exported_json, part_script, model_name=model_name, save_as_path=None)

        if exported_count:
            status = "ok"
            messages.append("已导出 %s 个 Part 建模构件并嵌入到 Abaqus 脚本，Abaqus model 名称: %s。" % (exported_count, model_name))
        else:
            status = "needs_review"
            messages.append("未导出任何 Part 构件；请检查选择模式、规格、长度、数量、材料牌号和建模方式。")

        report_payload = {
            "status": status,
            "project_prefix": prefix,
            "selection": selection,
            "source_workbook": str(workbook.resolve()),
            "copied_workbook": None,
            "components_json": None,
            "data_mode": "embedded_in_part_script",
            "part_script": str(part_script.resolve()) if part_script else None,
            "model_name": model_name,
            "cae_save_path": None,
            "summary": summary,
            "messages": messages,
        }
        report = write_json_report(report_path, report_payload)
        return PartScriptOutput(
            workbook_path=str(workbook.resolve()),
            status=status,
            project_prefix=prefix,
            project_dir=str(output_dir.resolve()),
            copied_workbook_path=None,
            components_json_path=None,
            part_script_path=str(part_script.resolve()) if part_script else None,
            cae_save_path=None,
            report_path=str(report.resolve()),
            exported_count=exported_count,
            complete_count=int(summary["complete_count"]),
            row_count=int(summary["row_count"]),
            messages=messages,
        )
    except Exception as exc:
        prefix = project_prefix_value or "UNKNOWN_PROJECT"
        output_dir = Path(output_root)
        output_dir.mkdir(parents=True, exist_ok=True)
        messages.append(str(exc))
        report = write_json_report(
            step02_debug_report_path(output_dir, prefix),
            {
                "status": "failed",
                "source_workbook": str(workbook.resolve()),
                "messages": messages,
            },
        )
        return PartScriptOutput(
            workbook_path=str(workbook.resolve()),
            status="failed",
            project_prefix=prefix,
            project_dir=str(output_dir.resolve()),
            copied_workbook_path=None,
            components_json_path=None,
            part_script_path=None,
            cae_save_path=None,
            report_path=str(report.resolve()),
            exported_count=0,
            complete_count=0,
            row_count=0,
            messages=messages,
        )


def batch_generate_part_scripts(
    workbook_paths: Iterable[str | Path],
    output_root: str | Path,
    selection: str = "complete",
    standards_path: str | Path | None = None,
    cae_save_path: str | Path | None = None,
    overwrite: bool = False,
) -> list[PartScriptOutput]:
    root = Path(output_root)
    root.mkdir(parents=True, exist_ok=True)
    outputs: list[PartScriptOutput] = []
    for workbook in workbook_paths:
        outputs.append(
            generate_part_script_from_workbook(
                workbook,
                root,
                selection=selection,
                standards_path=standards_path,
                cae_save_path=cae_save_path,
                overwrite=overwrite,
            )
        )
    return outputs
