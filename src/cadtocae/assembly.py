from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


CONFIRMED = "已确认"
MANUAL_CHECK = "需人工确认"
NOT_USED = "暂不使用"
PASSED = "通过"
FAILED = "不通过"
CONTROL_LENGTH_TOLERANCE_M = 0.001
ANGLE_TOLERANCE_DEG = 0.05

DIMENSION_INPUT_HEADERS = [
    "参数名",
    "参数含义",
    "数值",
    "单位",
    "分段尺寸_mm",
    "分段合计_mm",
    "分段误差_mm",
    "来源尺寸链",
    "校核状态",
    "备注",
]
LOCAL_SECTION_HEADERS = [
    "局部截面",
    "关联全局点",
    "局部里程_s_m",
    "相对F里程_m",
    "来源参数",
    "计算公式",
    "校核状态",
    "备注",
]
POINT_HEADERS = [
    "点名",
    "点类型",
    "X_m",
    "Y_m",
    "Z_m",
    "来源参数",
    "计算公式",
    "校核状态",
    "标注说明",
]
GEOMETRY_CHECK_HEADERS = [
    "校核项",
    "校核类型",
    "起点",
    "终点",
    "计算值",
    "图纸/输入值",
    "误差",
    "允许误差",
    "是否通过",
    "校核状态",
    "计算公式",
    "备注",
]
MEMBER_HEADERS = [
    "构件名称",
    "abaqus_part_name",
    "实例名",
    "起点",
    "终点",
    "轴线长度_m",
    "校核长度_m",
    "误差_m",
    "是否通过",
    "槽口方向",
    "旋转规则",
    "校核状态",
    "备注",
]
ANNOTATION_HEADERS = [
    "标注对象",
    "对象类型",
    "图上标签",
    "X_m",
    "Z_m",
    "像素X",
    "像素Y",
    "校核状态",
]


def load_layout(path: str | Path) -> dict[str, Any]:
    with Path(path).open("r", encoding="utf-8") as handle:
        return json.load(handle)


def dimension_inputs(layout: dict[str, Any]) -> list[dict[str, Any]]:
    return list(layout.get("dimension_inputs", []))


def dimension_input_map(layout: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {row["name"]: row for row in dimension_inputs(layout)}


def _as_float(value: Any) -> float:
    if value in (None, ""):
        raise ValueError("empty numeric value")
    return float(value)


def _input_value(layout: dict[str, Any], name: str) -> float:
    row = dimension_input_map(layout)[name]
    return _as_float(row.get("value"))


def _input_value_m(layout: dict[str, Any], name: str) -> float:
    row = dimension_input_map(layout)[name]
    value = _as_float(row.get("value"))
    unit = str(row.get("unit", "")).lower()
    if unit == "mm":
        return value / 1000.0
    if unit == "m":
        return value
    raise ValueError("%s is not a length input" % name)


def _segments_text(row: dict[str, Any]) -> str:
    segments = row.get("segments_mm") or []
    return "+".join("%g" % float(value) for value in segments)


def _segments_sum(row: dict[str, Any]) -> float | None:
    segments = row.get("segments_mm") or []
    if not segments:
        return None
    return sum(float(value) for value in segments)


def _segments_error(row: dict[str, Any]) -> float | None:
    total = _segments_sum(row)
    if total is None or str(row.get("unit", "")).lower() != "mm":
        return None
    return total - _as_float(row.get("value"))


def _effective_input_status(row: dict[str, Any]) -> str:
    error = _segments_error(row)
    if error is not None and abs(error) > 1.0e-9:
        return MANUAL_CHECK
    return row.get("status") or MANUAL_CHECK


def _input_status(layout: dict[str, Any], names: list[str]) -> str:
    inputs = dimension_input_map(layout)
    for name in names:
        if name not in inputs or _effective_input_status(inputs[name]) != CONFIRMED:
            return MANUAL_CHECK
    return CONFIRMED


def _beam_unit(layout: dict[str, Any]) -> tuple[float, float, float]:
    theta_deg = _input_value(layout, "theta_deg")
    theta = math.radians(theta_deg)
    return theta_deg, math.cos(theta), math.sin(theta)


def solve_control_points(layout: dict[str, Any]) -> list[dict[str, Any]]:
    if not dimension_inputs(layout):
        return list(layout.get("points", []))

    theta_deg, cos_theta, sin_theta = _beam_unit(layout)
    z_a = _input_value_m(layout, "Z_A_mm")
    x_f = _input_value_m(layout, "X_F_mm")
    z_f = _input_value_m(layout, "Z_F_mm")
    z_bd = _input_value_m(layout, "Z_BD_mm")
    r_hoop = _input_value_m(layout, "R_hoop_mm")
    gc = _input_value_m(layout, "GC_mm")
    gf = _input_value_m(layout, "GF_mm")
    ge = _input_value_m(layout, "GE_mm")

    base_points = {
        point["name"]: dict(point)
        for point in layout.get("points", [])
        if point.get("name") not in {"A", "B", "C", "D", "E", "F", "G_global"}
    }
    solved = [
        {
            "name": "A",
            "type": "N",
            "x_m": 0.0,
            "y_m": 0.0,
            "z_m": z_a,
            "source_dimension": "Z_A_mm",
            "formula": "A=(0,0,Z_A)",
            "status": _input_status(layout, ["Z_A_mm"]),
            "annotation": "上立柱上顶点",
        },
        {
            "name": "B",
            "type": "N",
            "x_m": -r_hoop,
            "y_m": 0.0,
            "z_m": z_bd,
            "source_dimension": "Z_BD_mm, R_hoop_mm",
            "formula": "B=(-R_hoop,0,Z_BD)",
            "status": _input_status(layout, ["Z_BD_mm", "R_hoop_mm"]),
            "annotation": "前斜撑与抱箍交点",
        },
        {
            "name": "C",
            "type": "N",
            "x_m": x_f + (gc - gf) * cos_theta,
            "y_m": 0.0,
            "z_m": z_f + (gc - gf) * sin_theta,
            "source_dimension": "X_F_mm, Z_F_mm, GC_mm, GF_mm, theta_deg",
            "formula": "C=F+(GC-GF)*(cos(theta),0,sin(theta))",
            "status": _input_status(layout, ["X_F_mm", "Z_F_mm", "GC_mm", "GF_mm", "theta_deg"]),
            "annotation": "斜梁与前斜撑交点",
        },
        {
            "name": "D",
            "type": "N",
            "x_m": r_hoop,
            "y_m": 0.0,
            "z_m": z_bd,
            "source_dimension": "Z_BD_mm, R_hoop_mm",
            "formula": "D=(+R_hoop,0,Z_BD)",
            "status": _input_status(layout, ["Z_BD_mm", "R_hoop_mm"]),
            "annotation": "后斜撑与抱箍交点",
        },
        {
            "name": "E",
            "type": "N",
            "x_m": x_f + (ge - gf) * cos_theta,
            "y_m": 0.0,
            "z_m": z_f + (ge - gf) * sin_theta,
            "source_dimension": "X_F_mm, Z_F_mm, GE_mm, GF_mm, theta_deg",
            "formula": "E=F+(GE-GF)*(cos(theta),0,sin(theta))",
            "status": _input_status(layout, ["X_F_mm", "Z_F_mm", "GE_mm", "GF_mm", "theta_deg"]),
            "annotation": "斜梁与后斜撑交点",
        },
        {
            "name": "F",
            "type": "N",
            "x_m": x_f,
            "y_m": 0.0,
            "z_m": z_f,
            "source_dimension": "X_F_mm, Z_F_mm",
            "formula": "F=(X_F,0,Z_F)",
            "status": _input_status(layout, ["X_F_mm", "Z_F_mm"]),
            "annotation": "斜梁与上立柱/三角连接件参考交点",
        },
        {
            "name": "G_global",
            "type": "REF",
            "x_m": x_f - gf * cos_theta,
            "y_m": 0.0,
            "z_m": z_f - gf * sin_theta,
            "source_dimension": "X_F_mm, Z_F_mm, GF_mm, theta_deg",
            "formula": "G_global=F-GF*(cos(theta),0,sin(theta))",
            "status": _input_status(layout, ["X_F_mm", "Z_F_mm", "GF_mm", "theta_deg"]),
            "annotation": "斜梁局部起点派生全局位置，仅用于放置和标注",
        },
    ]

    for point in solved:
        base_points[point["name"]] = point
    order = ["O", "A", "B", "C", "D", "E", "F", "G_global"]
    return [base_points[name] for name in order if name in base_points] + [
        point for name, point in base_points.items() if name not in order
    ]


def beam_local_sections(layout: dict[str, Any]) -> list[dict[str, Any]]:
    if not dimension_inputs(layout):
        return []

    gc = _input_value_m(layout, "GC_mm")
    gf = _input_value_m(layout, "GF_mm")
    ge = _input_value_m(layout, "GE_mm")
    return [
        {
            "section": "G",
            "global_point": "G_global",
            "station_m": 0.0,
            "offset_from_f_m": -gf,
            "source_dimension": "0",
            "formula": "G=0",
            "status": _input_status(layout, ["X_F_mm", "Z_F_mm", "GF_mm", "theta_deg"]),
            "note": "斜梁 Part 局部起点；不是人工输入的全局控制点",
        },
        {
            "section": "C",
            "global_point": "C",
            "station_m": gc,
            "offset_from_f_m": gc - gf,
            "source_dimension": "GC_mm",
            "formula": "s_C=GC",
            "status": _input_status(layout, ["GC_mm"]),
            "note": "斜梁与前斜撑交点截面",
        },
        {
            "section": "F",
            "global_point": "F",
            "station_m": gf,
            "offset_from_f_m": 0.0,
            "source_dimension": "GF_mm",
            "formula": "s_F=GF",
            "status": _input_status(layout, ["GF_mm"]),
            "note": "斜梁与上立柱/三角连接件参考截面",
        },
        {
            "section": "E",
            "global_point": "E",
            "station_m": ge,
            "offset_from_f_m": ge - gf,
            "source_dimension": "GE_mm",
            "formula": "s_E=GE",
            "status": _input_status(layout, ["GE_mm"]),
            "note": "斜梁与后斜撑交点截面",
        },
    ]


def point_map(layout: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {point["name"]: point for point in solve_control_points(layout)}


def distance_3d(start: dict[str, Any], end: dict[str, Any]) -> float:
    dx = float(end.get("x_m", 0.0)) - float(start.get("x_m", 0.0))
    dy = float(end.get("y_m", 0.0)) - float(start.get("y_m", 0.0))
    dz = float(end.get("z_m", 0.0)) - float(start.get("z_m", 0.0))
    return math.sqrt(dx * dx + dy * dy + dz * dz)


def angle_xz_deg(start: dict[str, Any], end: dict[str, Any]) -> float:
    dx = float(end.get("x_m", 0.0)) - float(start.get("x_m", 0.0))
    dz = float(end.get("z_m", 0.0)) - float(start.get("z_m", 0.0))
    return math.degrees(math.atan2(dz, dx))


def _angle_error_deg(calc: float, target: float) -> float:
    return (calc - target + 180.0) % 360.0 - 180.0


def brace_length_checks(layout: dict[str, Any], tolerance_m: float = CONTROL_LENGTH_TOLERANCE_M) -> list[dict[str, Any]]:
    if not dimension_inputs(layout):
        return []

    points = point_map(layout)
    checks: list[dict[str, Any]] = []
    for segment, start_name, end_name, draw_input, required_inputs in [
        (
            "BC",
            "B",
            "C",
            "L_BC_draw_mm",
            ["theta_deg", "X_F_mm", "Z_F_mm", "Z_BD_mm", "R_hoop_mm", "GC_mm", "GF_mm", "L_BC_draw_mm"],
        ),
        (
            "DE",
            "D",
            "E",
            "L_DE_draw_mm",
            ["theta_deg", "X_F_mm", "Z_F_mm", "Z_BD_mm", "R_hoop_mm", "GE_mm", "GF_mm", "L_DE_draw_mm"],
        ),
    ]:
        start = points[start_name]
        end = points[end_name]
        calc_length = distance_3d(start, end)
        draw_length = _input_value_m(layout, draw_input)
        error = calc_length - draw_length
        passed = PASSED if abs(error) <= tolerance_m else FAILED
        status = CONFIRMED if passed == PASSED and _input_status(layout, required_inputs) == CONFIRMED else MANUAL_CHECK
        checks.append(
            {
                "segment": segment,
                "start": start_name,
                "end": end_name,
                "calc_length_m": calc_length,
                "drawing_length_m": draw_length,
                "error_m": error,
                "tolerance_m": tolerance_m,
                "passed": passed,
                "status": status,
                "formula": "sqrt((%s.x-%s.x)^2+(%s.z-%s.z)^2)" % (end_name, start_name, end_name, start_name),
                "note": "误差不超过 ±1mm 时通过",
            }
        )
    return checks


def control_length_checks(layout: dict[str, Any], tolerance_m: float = CONTROL_LENGTH_TOLERANCE_M) -> list[dict[str, Any]]:
    return brace_length_checks(layout, tolerance_m)


def length_and_angle_checks(
    layout: dict[str, Any],
    tolerance_m: float = CONTROL_LENGTH_TOLERANCE_M,
    angle_tolerance_deg: float = ANGLE_TOLERANCE_DEG,
) -> list[dict[str, Any]]:
    if not dimension_inputs(layout):
        return []

    gc = _input_value_m(layout, "GC_mm")
    gf = _input_value_m(layout, "GF_mm")
    ge = _input_value_m(layout, "GE_mm")
    theta_deg = _input_value(layout, "theta_deg")
    points = point_map(layout)
    c = points["C"]
    e = points["E"]
    order_passed = gc < gf < ge
    order_status = CONFIRMED if order_passed and _input_status(layout, ["GC_mm", "GF_mm", "GE_mm"]) == CONFIRMED else MANUAL_CHECK
    angle_calc = angle_xz_deg(c, e)
    angle_error = _angle_error_deg(angle_calc, theta_deg)
    angle_passed = abs(angle_error) <= angle_tolerance_deg
    angle_status = (
        CONFIRMED
        if angle_passed
        and _input_status(layout, ["theta_deg", "X_F_mm", "Z_F_mm", "GC_mm", "GF_mm", "GE_mm"]) == CONFIRMED
        else MANUAL_CHECK
    )

    checks: list[dict[str, Any]] = [
        {
            "check_item": "GC_GF_GE_ORDER",
            "check_type": "斜梁局部截面顺序",
            "start": "G/C/F",
            "end": "E",
            "calc_value": "GC=%.6f, GF=%.6f, GE=%.6f" % (gc, gf, ge),
            "reference_value": "GC < GF < GE",
            "error": "",
            "tolerance": "",
            "passed": PASSED if order_passed else FAILED,
            "status": order_status,
            "formula": "GC < GF < GE",
            "note": "不满足时 C/F/E 截面顺序需人工确认",
        },
        {
            "check_item": "CF_LOCAL",
            "check_type": "斜梁局部截面距离",
            "start": "C",
            "end": "F",
            "calc_value": gf - gc,
            "reference_value": "GF-GC",
            "error": "",
            "tolerance": "",
            "passed": PASSED if gf > gc else FAILED,
            "status": order_status,
            "formula": "CF_calc=GF-GC",
            "note": "用于复核图纸分段尺寸",
        },
        {
            "check_item": "FE_LOCAL",
            "check_type": "斜梁局部截面距离",
            "start": "F",
            "end": "E",
            "calc_value": ge - gf,
            "reference_value": "GE-GF",
            "error": "",
            "tolerance": "",
            "passed": PASSED if ge > gf else FAILED,
            "status": order_status,
            "formula": "FE_calc=GE-GF",
            "note": "用于复核图纸分段尺寸",
        },
        {
            "check_item": "CE_ANGLE",
            "check_type": "斜梁角度",
            "start": "C",
            "end": "E",
            "calc_value": angle_calc,
            "reference_value": theta_deg,
            "error": angle_error,
            "tolerance": angle_tolerance_deg,
            "passed": PASSED if angle_passed else FAILED,
            "status": angle_status,
            "formula": "atan2(E.z-C.z,E.x-C.x)",
            "note": "C/F/E 应在斜梁轴线上，CE 与 +X 夹角等于 theta_deg",
        },
    ]

    for brace in brace_length_checks(layout, tolerance_m):
        checks.append(
            {
                "check_item": brace["segment"],
                "check_type": "斜撑长度",
                "start": brace["start"],
                "end": brace["end"],
                "calc_value": brace["calc_length_m"],
                "reference_value": brace["drawing_length_m"],
                "error": brace["error_m"],
                "tolerance": brace["tolerance_m"],
                "passed": brace["passed"],
                "status": brace["status"],
                "formula": brace["formula"],
                "note": brace["note"],
            }
        )
    return checks


def beam_geometry_ready(layout: dict[str, Any]) -> bool:
    checks = {check["check_item"]: check for check in length_and_angle_checks(layout)}
    required = ["GC_GF_GE_ORDER", "CE_ANGLE"]
    return all(checks.get(name, {}).get("status") == CONFIRMED for name in required)


def member_checks(layout: dict[str, Any], tolerance_m: float = 0.001) -> list[dict[str, Any]]:
    points = point_map(layout)
    segment_status = {check["segment"]: check["status"] for check in brace_length_checks(layout, CONTROL_LENGTH_TOLERANCE_M)}
    rows: list[dict[str, Any]] = []

    for member in layout.get("members", []):
        start_name = member.get("start", "")
        end_name = member.get("end", "")
        start = points.get(start_name)
        end = points.get(end_name)
        issues: list[str] = []
        axis_length = None
        error = None
        passed = MANUAL_CHECK

        if not start:
            issues.append("起点不存在: %s" % start_name)
        if not end:
            issues.append("终点不存在: %s" % end_name)

        if start and end:
            axis_length = distance_3d(start, end)
            check_length = member.get("axis_check_length_m", member.get("material_length_m"))
            if check_length not in (None, ""):
                error = axis_length - float(check_length)
                passed = PASSED if abs(error) <= tolerance_m else FAILED
            if start.get("status") != CONFIRMED or end.get("status") != CONFIRMED:
                issues.append("起点或终点未确认")
                passed = MANUAL_CHECK

        linked_segment = member.get("control_segment")
        if linked_segment in {"BC", "DE"} and segment_status.get(linked_segment) != CONFIRMED:
            issues.append("%s 长度校核未通过或未确认" % linked_segment)
            passed = MANUAL_CHECK
        if linked_segment == "BEAM_LOCAL" and not beam_geometry_ready(layout):
            issues.append("斜梁局部截面顺序或角度校核未通过/未确认")
            passed = MANUAL_CHECK

        if member.get("status") != CONFIRMED:
            issues.append("构件轴线未确认")
            passed = MANUAL_CHECK

        if not issues and passed == MANUAL_CHECK:
            passed = PASSED

        row = dict(member)
        row.update(
            {
                "axis_length_m": axis_length,
                "length_error_m": error,
                "passed": passed,
                "issues": "；".join(issues),
            }
        )
        rows.append(row)

    return rows


def ready_members(layout: dict[str, Any], tolerance_m: float = 0.001) -> list[dict[str, Any]]:
    checks = member_checks(layout, tolerance_m)
    points = point_map(layout)
    ready: list[dict[str, Any]] = []

    for check in checks:
        if check["passed"] != PASSED:
            continue
        start = points[check["start"]]
        end = points[check["end"]]
        ready.append(
            {
                "component_name": check["component_name"],
                "abaqus_part_name": check["abaqus_part_name"],
                "instance_name": check["instance_name"],
                "start": {
                    "name": check["start"],
                    "x_m": start["x_m"],
                    "y_m": start.get("y_m", 0.0),
                    "z_m": start["z_m"],
                },
                "end": {
                    "name": check["end"],
                    "x_m": end["x_m"],
                    "y_m": end.get("y_m", 0.0),
                    "z_m": end["z_m"],
                },
                "axis_length_m": check["axis_length_m"],
                "slot_direction": check.get("slot_direction", ""),
                "rotation_rule": check.get("rotation_rule", ""),
                "control_segment": check.get("control_segment", ""),
            }
        )

    return ready


def export_assembly_inputs(
    layout_path: str | Path,
    output_path: str | Path,
    include_draft: bool = False,
    tolerance_m: float = 0.001,
    control_tolerance_m: float = CONTROL_LENGTH_TOLERANCE_M,
    angle_tolerance_deg: float = ANGLE_TOLERANCE_DEG,
) -> Path:
    layout = load_layout(layout_path)
    payload: dict[str, Any] = {
        "meta": layout.get("meta", {}),
        "tolerance_m": tolerance_m,
        "control_length_tolerance_m": control_tolerance_m,
        "angle_tolerance_deg": angle_tolerance_deg,
        "control_points": solve_control_points(layout),
        "beam_local_sections": beam_local_sections(layout),
        "geometry_checks": length_and_angle_checks(layout, control_tolerance_m, angle_tolerance_deg),
        "length_checks": brace_length_checks(layout, control_tolerance_m),
        "ready_members": ready_members(layout, tolerance_m),
    }
    if include_draft:
        payload["draft_member_checks"] = member_checks(layout, tolerance_m)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
    return output


def _append_table(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    last_row = max(ws.max_row, 1)
    last_col = get_column_letter(len(headers))
    ref = "A1:%s%d" % (last_col, last_row)
    if last_row >= 2:
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ref


def _style_sheet(ws, widths: dict[str, float]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    body_font = Font(name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


def _add_status_validation(ws, column_letter: str) -> None:
    dv = DataValidation(type="list", formula1='"%s,%s,%s"' % (CONFIRMED, MANUAL_CHECK, NOT_USED), allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("%s2:%s500" % (column_letter, column_letter))


def create_coordinate_workbook(
    layout_path: str | Path,
    output_path: str | Path,
    tolerance_m: float = CONTROL_LENGTH_TOLERANCE_M,
    angle_tolerance_deg: float = ANGLE_TOLERANCE_DEG,
) -> Path:
    layout = load_layout(layout_path)
    points = solve_control_points(layout)
    local_sections = beam_local_sections(layout)
    geometry_checks = length_and_angle_checks(layout, tolerance_m, angle_tolerance_deg)
    checks = member_checks(layout, tolerance_m)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("关键尺寸输入")
    input_rows = []
    for row in dimension_inputs(layout):
        input_rows.append(
            [
                row.get("name"),
                row.get("label", ""),
                row.get("value"),
                row.get("unit", ""),
                _segments_text(row),
                _segments_sum(row),
                _segments_error(row),
                row.get("source_dimension", ""),
                _effective_input_status(row),
                row.get("note", ""),
            ]
        )
    _append_table(ws, DIMENSION_INPUT_HEADERS, input_rows, "DimensionInputsTable")
    _style_sheet(ws, {"A": 18, "B": 28, "C": 12, "D": 10, "E": 22, "F": 16, "G": 16, "H": 36, "I": 14, "J": 48})
    _add_status_validation(ws, "I")

    ws = wb.create_sheet("斜梁局部截面")
    local_rows = [
        [
            row["section"],
            row["global_point"],
            round(row["station_m"], 6),
            round(row["offset_from_f_m"], 6),
            row["source_dimension"],
            row["formula"],
            row["status"],
            row["note"],
        ]
        for row in local_sections
    ]
    _append_table(ws, LOCAL_SECTION_HEADERS, local_rows, "BeamLocalSectionsTable")
    _style_sheet(ws, {"A": 12, "B": 14, "C": 16, "D": 16, "E": 18, "F": 28, "G": 14, "H": 42})
    _add_status_validation(ws, "G")

    ws = wb.create_sheet("控制点坐标")
    point_rows = [
        [
            p.get("name"),
            p.get("type"),
            round(p.get("x_m"), 6) if p.get("x_m") is not None else "",
            round(p.get("y_m", 0.0), 6) if p.get("y_m") is not None else "",
            round(p.get("z_m"), 6) if p.get("z_m") is not None else "",
            p.get("source_dimension", ""),
            p.get("formula", ""),
            p.get("status", ""),
            p.get("annotation", ""),
        ]
        for p in points
    ]
    _append_table(ws, POINT_HEADERS, point_rows, "ControlPointsTable")
    _style_sheet(ws, {"A": 12, "B": 10, "C": 12, "D": 12, "E": 12, "F": 42, "G": 56, "H": 14, "I": 40})
    _add_status_validation(ws, "H")

    ws = wb.create_sheet("长度与角度校核")
    geometry_rows = [
        [
            check["check_item"],
            check["check_type"],
            check["start"],
            check["end"],
            round(check["calc_value"], 6) if isinstance(check["calc_value"], (int, float)) else check["calc_value"],
            round(check["reference_value"], 6)
            if isinstance(check["reference_value"], (int, float))
            else check["reference_value"],
            round(check["error"], 6) if isinstance(check["error"], (int, float)) else check["error"],
            check["tolerance"],
            check["passed"],
            check["status"],
            check["formula"],
            check["note"],
        ]
        for check in geometry_checks
    ]
    _append_table(ws, GEOMETRY_CHECK_HEADERS, geometry_rows, "LengthAngleChecksTable")
    _style_sheet(ws, {"A": 18, "B": 20, "C": 10, "D": 10, "E": 24, "F": 18, "G": 12, "H": 12, "I": 12, "J": 14, "K": 42, "L": 48})

    ws = wb.create_sheet("构件轴线校核")
    member_rows = [
        [
            m.get("component_name"),
            m.get("abaqus_part_name"),
            m.get("instance_name"),
            m.get("start"),
            m.get("end"),
            round(m["axis_length_m"], 6) if m["axis_length_m"] is not None else "",
            m.get("axis_check_length_m", m.get("material_length_m", "")),
            round(m["length_error_m"], 6) if m["length_error_m"] is not None else "",
            m["passed"],
            m.get("slot_direction", ""),
            m.get("rotation_rule", ""),
            m.get("status", ""),
            (m.get("issues") or m.get("note") or ""),
        ]
        for m in checks
    ]
    _append_table(ws, MEMBER_HEADERS, member_rows, "MemberAxisCheckTable")
    _style_sheet(ws, {"A": 14, "B": 34, "C": 34, "D": 10, "E": 10, "F": 14, "G": 14, "H": 12, "I": 14, "J": 10, "K": 46, "L": 14, "M": 52})

    ws = wb.create_sheet("图纸标注记录")
    annotation_rows = []
    calibration = layout.get("image_calibration", {})
    for point in points:
        px, py = world_to_pixel(point, calibration)
        annotation_rows.append(
            [
                point.get("name"),
                point.get("type"),
                point_label(point),
                round(point.get("x_m"), 6) if point.get("x_m") is not None else "",
                round(point.get("z_m"), 6) if point.get("z_m") is not None else "",
                px,
                py,
                point.get("status"),
            ]
        )
    _append_table(ws, ANNOTATION_HEADERS, annotation_rows, "DrawingAnnotationTable")
    _style_sheet(ws, {"A": 14, "B": 10, "C": 34, "D": 12, "E": 12, "F": 12, "G": 12, "H": 14})

    ws = wb.create_sheet("坐标系说明")
    ws.append(["项目", layout.get("meta", {}).get("project", "")])
    ws.append(["单位体系", "m-kg-N-Pa"])
    ws.append(["坐标系", layout.get("meta", {}).get("coordinate_system", "")])
    ws.append(["原点规则", layout.get("meta", {}).get("origin_rule", "")])
    ws.append(["斜梁局部原点", "G 为斜梁 Part 左端局部起点，G_global 仅为派生全局参考点"])
    ws.append(["BC/DE允许误差_m", tolerance_m])
    ws.append(["斜梁角度允许误差_deg", angle_tolerance_deg])
    _style_sheet(ws, {"A": 28, "B": 100})

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def world_to_pixel(point: dict[str, Any], calibration: dict[str, Any]) -> tuple[int, int]:
    origin_x, origin_y = calibration.get("origin_px", [0, 0])
    scale = float(calibration.get("scale_px_per_m", 1.0))
    x = float(point.get("x_m", 0.0))
    z = float(point.get("z_m", 0.0))
    return int(round(origin_x + x * scale)), int(round(origin_y - z * scale))


def point_label(point: dict[str, Any]) -> str:
    return "%s (%.3f,%.3f)m" % (
        point.get("name"),
        float(point.get("x_m", 0.0)),
        float(point.get("z_m", 0.0)),
    )


def _font(size: int) -> ImageFont.ImageFont:
    candidates = [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\arial.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def _text_bbox(draw: ImageDraw.ImageDraw, xy: tuple[int, int], text: str, font: ImageFont.ImageFont) -> tuple[int, int, int, int]:
    if hasattr(draw, "textbbox"):
        return draw.textbbox(xy, text, font=font)
    width, height = draw.textsize(text, font=font)
    x, y = xy
    return x, y, x + width, y + height


def _draw_label(
    draw: ImageDraw.ImageDraw,
    image_size: tuple[int, int],
    anchor: tuple[int, int],
    text: str,
    font: ImageFont.ImageFont,
    outline: tuple[int, int, int],
) -> None:
    image_w, image_h = image_size
    px, py = anchor
    label_x = px + 10
    label_y = py - 24
    bbox = _text_bbox(draw, (label_x, label_y), text, font)
    width = bbox[2] - bbox[0]
    height = bbox[3] - bbox[1]

    if label_x + width + 6 > image_w:
        label_x = max(4, px - width - 12)
    if label_y < 4:
        label_y = py + 10
    if label_y + height + 6 > image_h:
        label_y = max(4, image_h - height - 8)

    bbox = _text_bbox(draw, (label_x, label_y), text, font)
    draw.rectangle((bbox[0] - 3, bbox[1] - 2, bbox[2] + 3, bbox[3] + 2), fill=(255, 255, 230), outline=outline)
    draw.text((label_x, label_y), text, fill=(0, 0, 0), font=font)


def _draw_segment_label(
    draw: ImageDraw.ImageDraw,
    image_size: tuple[int, int],
    p1: tuple[int, int],
    p2: tuple[int, int],
    text: str,
    font: ImageFont.ImageFont,
    color: tuple[int, int, int],
) -> None:
    x = int(round((p1[0] + p2[0]) / 2.0))
    y = int(round((p1[1] + p2[1]) / 2.0))
    _draw_label(draw, image_size, (x, y), text, font, color)


def _draw_polyline_for_points(
    draw: ImageDraw.ImageDraw,
    calibration: dict[str, Any],
    points: dict[str, dict[str, Any]],
    names: list[str],
    color: tuple[int, int, int],
    width: int,
) -> None:
    pixels = [world_to_pixel(points[name], calibration) for name in names if name in points]
    if len(pixels) >= 2:
        draw.line(pixels, fill=color, width=width)


def annotate_coordinates(
    layout_path: str | Path,
    image_path: str | Path,
    output_png: str | Path,
    output_pdf: str | Path | None = None,
) -> Path:
    layout = load_layout(layout_path)
    calibration = layout.get("image_calibration", {})
    image = Image.open(image_path).convert("RGB")
    draw = ImageDraw.Draw(image)
    font = _font(18)
    small_font = _font(14)

    origin = {"x_m": 0.0, "z_m": 0.0}
    ox, oy = world_to_pixel(origin, calibration)
    axis_len = int(float(calibration.get("scale_px_per_m", 1.0)) * 0.7)
    draw.line([(ox, oy), (ox + axis_len, oy)], fill=(220, 40, 40), width=4)
    draw.polygon(
        [(ox + axis_len, oy), (ox + axis_len - 12, oy - 6), (ox + axis_len - 12, oy + 6)],
        fill=(220, 40, 40),
    )
    draw.text((ox + axis_len + 6, oy - 12), "+X", fill=(220, 40, 40), font=font)
    draw.line([(ox, oy), (ox, oy - axis_len)], fill=(30, 120, 255), width=4)
    draw.polygon(
        [(ox, oy - axis_len), (ox - 6, oy - axis_len + 12), (ox + 6, oy - axis_len + 12)],
        fill=(30, 120, 255),
    )
    draw.text((ox + 8, oy - axis_len - 24), "+Z", fill=(30, 120, 255), font=font)

    colors = {
        CONFIRMED: (20, 150, 80),
        MANUAL_CHECK: (235, 130, 20),
        NOT_USED: (140, 140, 140),
        PASSED: (20, 150, 80),
        FAILED: (210, 40, 40),
    }
    points = point_map(layout)
    beam_color = (90, 65, 185)
    _draw_polyline_for_points(draw, calibration, points, ["G_global", "C", "F", "E"], beam_color, 3)
    _draw_polyline_for_points(draw, calibration, points, ["C", "E"], (70, 110, 210), 2)

    for check in brace_length_checks(layout, CONTROL_LENGTH_TOLERANCE_M):
        start = points[check["start"]]
        end = points[check["end"]]
        p1 = world_to_pixel(start, calibration)
        p2 = world_to_pixel(end, calibration)
        color = colors.get(check["passed"], (235, 130, 20))
        draw.line([p1, p2], fill=color, width=4)
        label = "%s %.3fm/Δ%.3fm" % (check["segment"], check["calc_length_m"], check["error_m"])
        _draw_segment_label(draw, image.size, p1, p2, label, small_font, color)

    angle_check = next((check for check in length_and_angle_checks(layout) if check["check_item"] == "CE_ANGLE"), None)
    if angle_check:
        c_px = world_to_pixel(points["C"], calibration)
        e_px = world_to_pixel(points["E"], calibration)
        text = "CE角度 %.2f°/Δ%.3f°" % (angle_check["calc_value"], angle_check["error"])
        _draw_segment_label(draw, image.size, c_px, e_px, text, small_font, beam_color)

    for member in layout.get("members", []):
        start = points.get(member.get("start"))
        end = points.get(member.get("end"))
        if not start or not end or member.get("start") == member.get("end"):
            continue
        p1 = world_to_pixel(start, calibration)
        p2 = world_to_pixel(end, calibration)
        draw.line([p1, p2], fill=(120, 80, 210), width=1)

    for point in solve_control_points(layout):
        px, py = world_to_pixel(point, calibration)
        color = beam_color if point.get("name") == "G_global" else colors.get(point.get("status"), (235, 130, 20))
        radius = 5 if point.get("name") == "G_global" else 7
        draw.ellipse((px - radius, py - radius, px + radius, py + radius), fill=color, outline=(0, 0, 0), width=2)
        label_anchor = (px, py + 28) if point.get("name") == "A" else (px, py)
        _draw_label(draw, image.size, label_anchor, point_label(point), small_font, color)

    legend_x, legend_y = 24, 24
    legend = [
        "坐标控制点校核图",
        "绿色：已确认/通过",
        "橙色：需人工确认",
        "红色：BC/DE长度误差超限",
        "紫色：斜梁局部G-C-F-E里程线",
        "单位：m, kg, N, Pa",
    ]
    for idx, text in enumerate(legend):
        draw.text((legend_x, legend_y + idx * 22), text, fill=(0, 0, 0), font=small_font)

    helper_y = legend_y + len(legend) * 22 + 12
    for idx, check in enumerate(length_and_angle_checks(layout, CONTROL_LENGTH_TOLERANCE_M)):
        if check["check_item"] not in {"BC", "DE", "CE_ANGLE"}:
            continue
        calc = check["calc_value"]
        reference = check["reference_value"]
        error = check["error"]
        text = "%s: calc=%s, ref=%s, err=%s" % (
            check["check_item"],
            "%.3f" % calc if isinstance(calc, (int, float)) else calc,
            "%.3f" % reference if isinstance(reference, (int, float)) else reference,
            "%.3f" % error if isinstance(error, (int, float)) else error,
        )
        color = colors.get(check["passed"], (25, 90, 170))
        draw.text((legend_x, helper_y + idx * 20), text, fill=color, font=small_font)

    output = Path(output_png)
    output.parent.mkdir(parents=True, exist_ok=True)
    image.save(output)
    if output_pdf:
        pdf_path = Path(output_pdf)
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        image.save(pdf_path, "PDF", resolution=200.0)
    return output
