from __future__ import annotations

import json
import tempfile
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from .standards import component_role


CONFIRMED = "已确认"
MANUAL_CHECK = "需人工确认"
NOT_USED = "暂不使用"
PASSED = "通过"
FAILED = "不通过"


def load_coordinate_layout(path: str | Path) -> dict[str, Any]:
    with Path(path).open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _copy_layout_with_prefix(layout: dict[str, Any], project_prefix: str) -> dict[str, Any]:
    copied = deepcopy(layout)
    copied.setdefault("meta", {})["project"] = project_prefix
    for member in copied.get("members", []):
        name = str(member.get("component_name", "")).strip()
        if not name:
            continue
        code = component_role(name)["code"]
        member["abaqus_part_name"] = "P_%s_%s" % (project_prefix, code)
        suffix = "_01" if code in {"BRACE_FRONT", "BRACE_REAR", "INCLINED_BEAM"} else ""
        member["instance_name"] = "I_%s_%s%s" % (project_prefix, code, suffix)
    return copied


def _dimension_rows(layout: dict[str, Any], control_tolerance_m: float, angle_tolerance_deg: float) -> list[dict[str, Any]]:
    rows = [dict(row) for row in layout.get("dimension_inputs", [])]
    calibration = layout.get("image_calibration") or {}
    rows.extend(
        [
            {
                "name": "origin_px_x",
                "label": "图纸标注原点像素 X",
                "value": (calibration.get("origin_px") or [0, 0])[0],
                "unit": "px",
                "source_dimension": "图纸标注显示用参数",
                "status": CONFIRMED,
                "note": "仅影响图纸标注记录中的像素位置，不影响 Abaqus 坐标。",
            },
            {
                "name": "origin_px_y",
                "label": "图纸标注原点像素 Y",
                "value": (calibration.get("origin_px") or [0, 0])[1],
                "unit": "px",
                "source_dimension": "图纸标注显示用参数",
                "status": CONFIRMED,
                "note": "仅影响图纸标注记录中的像素位置，不影响 Abaqus 坐标。",
            },
            {
                "name": "scale_px_per_m",
                "label": "图纸标注比例",
                "value": calibration.get("scale_px_per_m", 1),
                "unit": "px/m",
                "source_dimension": "图纸标注显示用参数",
                "status": CONFIRMED,
                "note": "仅影响图纸标注记录中的像素位置，不影响 Abaqus 坐标。",
            },
            {
                "name": "control_tolerance_m",
                "label": "BC/DE 长度允许误差",
                "value": control_tolerance_m,
                "unit": "m",
                "source_dimension": "校核规则",
                "status": CONFIRMED,
                "note": "默认 ±1mm。",
            },
            {
                "name": "angle_tolerance_deg",
                "label": "CE 角度允许误差",
                "value": angle_tolerance_deg,
                "unit": "deg",
                "source_dimension": "校核规则",
                "status": CONFIRMED,
                "note": "默认 0.05°。",
            },
        ]
    )
    return rows


def write_coordinate_layout_template(output_path: str | Path, project_prefix: str) -> Path:
    template = {
        "meta": {
            "project": project_prefix,
            "title": "待填写坐标尺寸输入",
            "units": {"length": "m", "mass": "kg", "force": "N", "stress": "Pa"},
            "coordinate_system": "Z 竖直向上，X 向图纸右侧，Y 向图纸外侧。",
            "origin_rule": "O 取下立柱中心线与地面线交点。",
        },
        "image_calibration": {
            "origin_px": [0, 0],
            "scale_px_per_m": 1,
            "note": "仅用于图上标注显示。",
        },
        "dimension_inputs": [
            {"name": "theta_deg", "label": "斜梁倾角", "value": "", "unit": "deg", "status": MANUAL_CHECK},
            {"name": "Z_A_mm", "label": "A 点高度", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "X_F_mm", "label": "F 点 X 坐标", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "Z_F_mm", "label": "F 点高度", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "Z_BD_mm", "label": "B/D 点共同高度", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "R_hoop_mm", "label": "抱箍连接点水平偏移", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "GC_mm", "label": "G 到 C 的斜梁局部里程", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "GF_mm", "label": "G 到 F 的斜梁局部里程", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "GE_mm", "label": "G 到 E 的斜梁局部里程", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "L_BC_draw_mm", "label": "图纸标注 BC 长度", "value": "", "unit": "mm", "status": MANUAL_CHECK},
            {"name": "L_DE_draw_mm", "label": "图纸标注 DE 长度", "value": "", "unit": "mm", "status": MANUAL_CHECK},
        ],
        "members": [],
    }
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def _style_title(ws, cell_range: str, title: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int, start: int, end: int) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True, color="1F2937")
    border = Border(
        left=Side(style="thin", color="B7C9D6"),
        right=Side(style="thin", color="B7C9D6"),
        top=Side(style="thin", color="B7C9D6"),
        bottom=Side(style="thin", color="B7C9D6"),
    )
    for col in range(start, end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _add_status_validation(ws, cell_range: str) -> None:
    dv = DataValidation(type="list", formula1='"%s,%s,%s"' % (CONFIRMED, MANUAL_CHECK, NOT_USED), allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _segments_text(row: dict[str, Any]) -> str:
    return "+".join("%g" % float(value) for value in row.get("segments_mm") or [])


def _segments_sum(row: dict[str, Any]) -> float | str:
    segments = row.get("segments_mm") or []
    return sum(float(value) for value in segments) if segments else ""


def _row_map(input_rows: list[dict[str, Any]], start_row: int = 4) -> dict[str, int]:
    return {row["name"]: index + start_row for index, row in enumerate(input_rows)}


def _quoted(sheet: str, cell: str) -> str:
    return "'%s'!%s" % (sheet, cell)


def _mm(ref: str) -> str:
    return "(%s/1000)" % ref


def _status_formula(row_by_name: dict[str, int], status_col: str, names: list[str], sheet: str | None = None) -> str:
    refs = []
    for name in names:
        cell = "$%s$%d" % (status_col, row_by_name[name])
        refs.append('%s="%s"' % ((_quoted(sheet, cell) if sheet else cell), CONFIRMED))
    return '=IF(AND(%s),"%s","%s")' % (",".join(refs), CONFIRMED, MANUAL_CHECK)


def _input_ref(row_by_name: dict[str, int], name: str, sheet: str | None = None, col: str = "C") -> str:
    cell = "$%s$%d" % (col, row_by_name[name])
    return _quoted(sheet, cell) if sheet else cell


def _write_input_sheet(ws, input_rows: list[dict[str, Any]], title: str, simple: bool = False) -> dict[str, int]:
    headers = ["参数名", "参数含义", "数值", "单位", "校核状态", "备注"] if simple else [
        "参数名", "参数含义", "数值", "单位", "分段尺寸_mm", "分段合计_mm", "分段误差_mm", "来源尺寸链", "校核状态", "备注"
    ]
    last_col = "F" if simple else "J"
    _style_title(ws, "A1:%s1" % last_col, title)
    for index, header in enumerate(headers, start=1):
        ws.cell(row=3, column=index).value = header
    _style_header(ws, 3, 1, len(headers))
    for row_index, row in enumerate(input_rows, start=4):
        values = [
            row.get("name"),
            row.get("label", ""),
            row.get("value", ""),
            row.get("unit", ""),
        ]
        if simple:
            values.extend([row.get("status") or MANUAL_CHECK, row.get("note", "")])
        else:
            values.extend([
                _segments_text(row),
                _segments_sum(row),
                '=IF(F%d="","",F%d-C%d)' % (row_index, row_index, row_index),
                row.get("source_dimension", ""),
                row.get("status") or MANUAL_CHECK,
                row.get("note", ""),
            ])
        for col_index, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col_index).value = value
    max_row = len(input_rows) + 3
    _style_body(ws, 4, max_row, 1, len(headers))
    status_col = "E" if simple else "I"
    _add_status_validation(ws, "%s4:%s%d" % (status_col, status_col, max_row))
    for cell in ws["C"][3:max_row]:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[status_col][3:max_row]:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return _row_map(input_rows)


def _add_optional_image(ws, image_path: str | Path | None, anchor: str = "A5", width: int = 780) -> None:
    if not image_path:
        return
    path = Path(image_path)
    if not path.exists():
        return
    image = ExcelImage(str(path))
    if image.width:
        ratio = width / float(image.width)
        image.width = width
        image.height = int(image.height * ratio)
    ws.add_image(image, anchor)


def _build_full_workbook(layout: dict[str, Any], input_rows: list[dict[str, Any]], project_prefix: str, figure_image: str | Path | None) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    input_ws = wb.create_sheet("关键尺寸输入")
    row_by_name = _write_input_sheet(input_ws, input_rows, "%s Assembly 坐标关键尺寸输入" % project_prefix)
    _set_widths(input_ws, {"A": 20, "B": 30, "C": 12, "D": 10, "E": 28, "F": 16, "G": 16, "H": 38, "I": 16, "J": 52})

    local_ws = wb.create_sheet("斜梁局部截面")
    _style_title(local_ws, "A1:H1", "斜梁局部截面里程")
    local_headers = ["局部截面", "关联全局点", "局部里程_s_m", "相对F里程_m", "来源参数", "计算公式", "校核状态", "备注"]
    for col, header in enumerate(local_headers, start=1):
        local_ws.cell(row=3, column=col).value = header
    _style_header(local_ws, 3, 1, 8)
    local_rows = [
        ["G", "G_global", "=0", "=-C6", "0", "G=0", _status_formula(row_by_name, "I", ["X_F_mm", "Z_F_mm", "GF_mm", "theta_deg"], "关键尺寸输入"), "斜梁 Part 左端局部起点"],
        ["C", "C", "=%s" % _mm(_input_ref(row_by_name, "GC_mm", "关键尺寸输入")), "=C5-C6", "GC_mm", "s_C=GC", _status_formula(row_by_name, "I", ["GC_mm"], "关键尺寸输入"), "斜梁与前斜撑交点截面"],
        ["F", "F", "=%s" % _mm(_input_ref(row_by_name, "GF_mm", "关键尺寸输入")), "=0", "GF_mm", "s_F=GF", _status_formula(row_by_name, "I", ["GF_mm"], "关键尺寸输入"), "斜梁与上立柱参考截面"],
        ["E", "E", "=%s" % _mm(_input_ref(row_by_name, "GE_mm", "关键尺寸输入")), "=C7-C6", "GE_mm", "s_E=GE", _status_formula(row_by_name, "I", ["GE_mm"], "关键尺寸输入"), "斜梁与后斜撑交点截面"],
    ]
    for row_index, row in enumerate(local_rows, start=4):
        for col_index, value in enumerate(row, start=1):
            local_ws.cell(row=row_index, column=col_index).value = value
    _style_body(local_ws, 4, 7, 1, 8)
    _set_widths(local_ws, {"A": 12, "B": 16, "C": 16, "D": 16, "E": 18, "F": 26, "G": 16, "H": 48})

    point_ws = wb.create_sheet("控制点坐标")
    _style_title(point_ws, "A1:I1", "A/B/C/D/E/F 控制点坐标（公式计算）")
    point_headers = ["点名", "点类型", "X_m", "Y_m", "Z_m", "来源参数", "计算公式", "校核状态", "标注说明"]
    for col, header in enumerate(point_headers, start=1):
        point_ws.cell(row=3, column=col).value = header
    _style_header(point_ws, 3, 1, 9)
    theta = _input_ref(row_by_name, "theta_deg", "关键尺寸输入")
    xf = _mm(_input_ref(row_by_name, "X_F_mm", "关键尺寸输入"))
    zf = _mm(_input_ref(row_by_name, "Z_F_mm", "关键尺寸输入"))
    za = _mm(_input_ref(row_by_name, "Z_A_mm", "关键尺寸输入"))
    zbd = _mm(_input_ref(row_by_name, "Z_BD_mm", "关键尺寸输入"))
    r = _mm(_input_ref(row_by_name, "R_hoop_mm", "关键尺寸输入"))
    gc = _mm(_input_ref(row_by_name, "GC_mm", "关键尺寸输入"))
    gf = _mm(_input_ref(row_by_name, "GF_mm", "关键尺寸输入"))
    ge = _mm(_input_ref(row_by_name, "GE_mm", "关键尺寸输入"))
    point_rows = [
        ["O", "N", "=0", "=0", "=0", "原点", "O=(0,0,0)", CONFIRMED, "立柱中心线地面交点"],
        ["A", "N", "=0", "=0", "=%s" % za, "Z_A_mm", "A=(0,0,Z_A)", _status_formula(row_by_name, "I", ["Z_A_mm"], "关键尺寸输入"), "上立柱上顶点"],
        ["B", "N", "=-%s" % r, "=0", "=%s" % zbd, "Z_BD_mm, R_hoop_mm", "B=(-R,0,Z_BD)", _status_formula(row_by_name, "I", ["Z_BD_mm", "R_hoop_mm"], "关键尺寸输入"), "前斜撑与抱箍交点"],
        ["C", "N", "=%s+(%s-%s)*COS(RADIANS(%s))" % (xf, gc, gf, theta), "=0", "=%s+(%s-%s)*SIN(RADIANS(%s))" % (zf, gc, gf, theta), "X_F_mm, Z_F_mm, GC_mm, GF_mm, theta_deg", "C=F+(GC-GF)u", _status_formula(row_by_name, "I", ["X_F_mm", "Z_F_mm", "GC_mm", "GF_mm", "theta_deg"], "关键尺寸输入"), "斜梁与前斜撑交点"],
        ["D", "N", "=%s" % r, "=0", "=%s" % zbd, "Z_BD_mm, R_hoop_mm", "D=(+R,0,Z_BD)", _status_formula(row_by_name, "I", ["Z_BD_mm", "R_hoop_mm"], "关键尺寸输入"), "后斜撑与抱箍交点"],
        ["E", "N", "=%s+(%s-%s)*COS(RADIANS(%s))" % (xf, ge, gf, theta), "=0", "=%s+(%s-%s)*SIN(RADIANS(%s))" % (zf, ge, gf, theta), "X_F_mm, Z_F_mm, GE_mm, GF_mm, theta_deg", "E=F+(GE-GF)u", _status_formula(row_by_name, "I", ["X_F_mm", "Z_F_mm", "GE_mm", "GF_mm", "theta_deg"], "关键尺寸输入"), "斜梁与后斜撑交点"],
        ["F", "N", "=%s" % xf, "=0", "=%s" % zf, "X_F_mm, Z_F_mm", "F=(X_F,0,Z_F)", _status_formula(row_by_name, "I", ["X_F_mm", "Z_F_mm"], "关键尺寸输入"), "斜梁与上立柱/三角连接件参考交点"],
        ["G_global", "REF", "=%s-%s*COS(RADIANS(%s))" % (xf, gf, theta), "=0", "=%s-%s*SIN(RADIANS(%s))" % (zf, gf, theta), "X_F_mm, Z_F_mm, GF_mm, theta_deg", "G_global=F-GF*u", _status_formula(row_by_name, "I", ["X_F_mm", "Z_F_mm", "GF_mm", "theta_deg"], "关键尺寸输入"), "斜梁局部起点派生全局位置"],
    ]
    for row_index, row in enumerate(point_rows, start=4):
        for col_index, value in enumerate(row, start=1):
            point_ws.cell(row=row_index, column=col_index).value = value
    _style_body(point_ws, 4, 11, 1, 9)
    _set_widths(point_ws, {"A": 14, "B": 10, "C": 14, "D": 12, "E": 14, "F": 44, "G": 44, "H": 16, "I": 42})

    check_ws = wb.create_sheet("长度与角度校核")
    _style_title(check_ws, "A1:L1", "局部截面、角度与斜撑长度校核（公式计算）")
    check_headers = ["校核项", "校核类型", "起点", "终点", "计算值", "图纸/输入值", "误差", "允许误差", "是否通过", "校核状态", "计算公式", "备注"]
    for col, header in enumerate(check_headers, start=1):
        check_ws.cell(row=3, column=col).value = header
    _style_header(check_ws, 3, 1, 12)
    check_values = [
        ["GC_GF_GE_ORDER", "斜梁局部截面顺序", "G/C/F", "E", '="GC="&TEXT(%s,"0.000")&", GF="&TEXT(%s,"0.000")&", GE="&TEXT(%s,"0.000")' % (_quoted("斜梁局部截面", "$C$5"), _quoted("斜梁局部截面", "$C$6"), _quoted("斜梁局部截面", "$C$7")), "GC < GF < GE", "", "", '=IF(AND(%s<%s,%s<%s),"%s","%s")' % (_quoted("斜梁局部截面", "$C$5"), _quoted("斜梁局部截面", "$C$6"), _quoted("斜梁局部截面", "$C$6"), _quoted("斜梁局部截面", "$C$7"), PASSED, FAILED), '=IF(AND(I4="%s",%s="%s",%s="%s",%s="%s"),"%s","%s")' % (PASSED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GC_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GF_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GE_mm"]), CONFIRMED, CONFIRMED, MANUAL_CHECK), "GC < GF < GE", "不满足时 C/F/E 截面顺序需人工确认"],
        ["CF_LOCAL", "斜梁局部截面距离", "C", "F", "=%s-%s" % (_quoted("斜梁局部截面", "$C$6"), _quoted("斜梁局部截面", "$C$5")), "GF-GC", "", "", '=IF(E5>0,"%s","%s")' % (PASSED, FAILED), '=IF(AND(I5="%s",%s="%s",%s="%s"),"%s","%s")' % (PASSED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GC_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GF_mm"]), CONFIRMED, CONFIRMED, MANUAL_CHECK), "CF_calc=GF-GC", "用于复核图纸分段尺寸"],
        ["FE_LOCAL", "斜梁局部截面距离", "F", "E", "=%s-%s" % (_quoted("斜梁局部截面", "$C$7"), _quoted("斜梁局部截面", "$C$6")), "GE-GF", "", "", '=IF(E6>0,"%s","%s")' % (PASSED, FAILED), '=IF(AND(I6="%s",%s="%s",%s="%s"),"%s","%s")' % (PASSED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GF_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GE_mm"]), CONFIRMED, CONFIRMED, MANUAL_CHECK), "FE_calc=GE-GF", "用于复核图纸分段尺寸"],
        ["CE_ANGLE", "斜梁角度", "C", "E", "=DEGREES(ATAN((%s-%s)/(%s-%s)))" % (_quoted("控制点坐标", "$E$9"), _quoted("控制点坐标", "$E$7"), _quoted("控制点坐标", "$C$9"), _quoted("控制点坐标", "$C$7")), "=%s" % _input_ref(row_by_name, "theta_deg", "关键尺寸输入"), "=MOD(E7-F7+180,360)-180", "=%s" % _input_ref(row_by_name, "angle_tolerance_deg", "关键尺寸输入"), '=IF(ABS(G7)<=H7,"%s","%s")' % (PASSED, FAILED), '=IF(AND(I7="%s",%s="%s",%s="%s",%s="%s",%s="%s",%s="%s",%s="%s"),"%s","%s")' % (PASSED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["theta_deg"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["X_F_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["Z_F_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GC_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GF_mm"]), CONFIRMED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["GE_mm"]), CONFIRMED, CONFIRMED, MANUAL_CHECK), "DEGREES(ATAN(ΔZ/ΔX))", "C/F/E 应在斜梁轴线上"],
        ["BC", "斜撑长度", "B", "C", "=SQRT((%s-%s)^2+(%s-%s)^2)" % (_quoted("控制点坐标", "$C$7"), _quoted("控制点坐标", "$C$6"), _quoted("控制点坐标", "$E$7"), _quoted("控制点坐标", "$E$6")), "=%s" % _mm(_input_ref(row_by_name, "L_BC_draw_mm", "关键尺寸输入")), "=E8-F8", "=%s" % _input_ref(row_by_name, "control_tolerance_m", "关键尺寸输入"), '=IF(ABS(G8)<=H8,"%s","%s")' % (PASSED, FAILED), '=IF(AND(I8="%s",%s="%s"),"%s","%s")' % (PASSED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["L_BC_draw_mm"]), CONFIRMED, CONFIRMED, MANUAL_CHECK), "sqrt((C.x-B.x)^2+(C.z-B.z)^2)", "误差不超过允许值时通过"],
        ["DE", "斜撑长度", "D", "E", "=SQRT((%s-%s)^2+(%s-%s)^2)" % (_quoted("控制点坐标", "$C$9"), _quoted("控制点坐标", "$C$8"), _quoted("控制点坐标", "$E$9"), _quoted("控制点坐标", "$E$8")), "=%s" % _mm(_input_ref(row_by_name, "L_DE_draw_mm", "关键尺寸输入")), "=E9-F9", "=%s" % _input_ref(row_by_name, "control_tolerance_m", "关键尺寸输入"), '=IF(ABS(G9)<=H9,"%s","%s")' % (PASSED, FAILED), '=IF(AND(I9="%s",%s="%s"),"%s","%s")' % (PASSED, _quoted("关键尺寸输入", "$I$%d" % row_by_name["L_DE_draw_mm"]), CONFIRMED, CONFIRMED, MANUAL_CHECK), "sqrt((E.x-D.x)^2+(E.z-D.z)^2)", "误差不超过允许值时通过"],
    ]
    for row_index, row in enumerate(check_values, start=4):
        for col_index, value in enumerate(row, start=1):
            check_ws.cell(row=row_index, column=col_index).value = value
    _style_body(check_ws, 4, 9, 1, 12)
    _set_widths(check_ws, {"A": 18, "B": 20, "C": 10, "D": 10, "E": 24, "F": 18, "G": 14, "H": 12, "I": 12, "J": 16, "K": 44, "L": 50})

    member_ws = wb.create_sheet("构件轴线校核")
    _style_title(member_ws, "A1:M1", "构件轴线校核（公式计算）")
    member_headers = ["构件名称", "abaqus_part_name", "实例名", "起点", "终点", "轴线长度_m", "校核长度_m", "误差_m", "是否通过", "槽口方向", "旋转规则", "校核状态", "备注"]
    for col, header in enumerate(member_headers, start=1):
        member_ws.cell(row=3, column=col).value = header
    _style_header(member_ws, 3, 1, 13)
    member_rows = layout.get("members") or [
        {"component_name": "前斜撑", "abaqus_part_name": "P_%s_BRACE_FRONT" % project_prefix, "instance_name": "I_%s_BRACE_FRONT_01" % project_prefix, "start": "B", "end": "C", "slot_direction": "+Y", "rotation_rule": "轴线对齐后绕轴线旋转，使槽口朝 +Y"},
        {"component_name": "后斜撑", "abaqus_part_name": "P_%s_BRACE_REAR" % project_prefix, "instance_name": "I_%s_BRACE_REAR_01" % project_prefix, "start": "D", "end": "E", "slot_direction": "+Y", "rotation_rule": "轴线对齐后绕轴线旋转，使槽口朝 +Y"},
        {"component_name": "斜梁", "abaqus_part_name": "P_%s_INCLINED_BEAM" % project_prefix, "instance_name": "I_%s_INCLINED_BEAM_01" % project_prefix, "start": "C", "end": "E", "slot_direction": "+Y", "rotation_rule": "以 G_global 定位，轴线旋转至 theta_deg。"},
    ]
    for row_index, member in enumerate(member_rows[:3], start=4):
        values = [
            member.get("component_name"),
            member.get("abaqus_part_name"),
            member.get("instance_name"),
            member.get("start"),
            member.get("end"),
            "",
            member.get("axis_check_length_m", ""),
            "",
            "",
            member.get("slot_direction", ""),
            member.get("rotation_rule", ""),
            member.get("status", MANUAL_CHECK),
            member.get("note", ""),
        ]
        for col_index, value in enumerate(values, start=1):
            member_ws.cell(row=row_index, column=col_index).value = value
    _style_body(member_ws, 4, 6, 1, 13)
    _set_widths(member_ws, {"A": 14, "B": 34, "C": 34, "D": 10, "E": 10, "F": 14, "G": 14, "H": 12, "I": 14, "J": 10, "K": 56, "L": 14, "M": 52})

    annotation_ws = wb.create_sheet("图纸标注记录")
    _style_title(annotation_ws, "A1:H1", "图纸标注记录（像素位置由输入标定自动计算）")
    annotation_headers = ["标注对象", "对象类型", "图上标签", "X_m", "Z_m", "像素X", "像素Y", "校核状态"]
    for col, header in enumerate(annotation_headers, start=1):
        annotation_ws.cell(row=3, column=col).value = header
    _style_header(annotation_ws, 3, 1, 8)
    for idx, name in enumerate(["O", "A", "B", "C", "D", "E", "F", "G_global"], start=4):
        source_row = idx
        annotation_ws.cell(row=idx, column=1).value = name
        annotation_ws.cell(row=idx, column=2).value = "REF" if name == "G_global" else "N"
        annotation_ws.cell(row=idx, column=3).value = '=A%d&" ("&TEXT(D%d,"0.000")&","&TEXT(E%d,"0.000")&")m"' % (idx, idx, idx)
        annotation_ws.cell(row=idx, column=4).value = "=%s" % _quoted("控制点坐标", "$C$%d" % source_row)
        annotation_ws.cell(row=idx, column=5).value = "=%s" % _quoted("控制点坐标", "$E$%d" % source_row)
        annotation_ws.cell(row=idx, column=6).value = "=ROUND(%s+D%d*%s,0)" % (_input_ref(row_by_name, "origin_px_x", "关键尺寸输入"), idx, _input_ref(row_by_name, "scale_px_per_m", "关键尺寸输入"))
        annotation_ws.cell(row=idx, column=7).value = "=ROUND(%s-E%d*%s,0)" % (_input_ref(row_by_name, "origin_px_y", "关键尺寸输入"), idx, _input_ref(row_by_name, "scale_px_per_m", "关键尺寸输入"))
        annotation_ws.cell(row=idx, column=8).value = "=%s" % _quoted("控制点坐标", "$H$%d" % source_row)
    _style_body(annotation_ws, 4, 11, 1, 8)
    _set_widths(annotation_ws, {"A": 14, "B": 10, "C": 34, "D": 14, "E": 14, "F": 12, "G": 12, "H": 16})

    info_ws = wb.create_sheet("坐标系说明")
    _style_title(info_ws, "A1:B1", "坐标系与使用说明")
    info_rows = [
        ["项目", project_prefix],
        ["单位体系", "m-kg-N-Pa"],
        ["坐标系", layout.get("meta", {}).get("coordinate_system", "")],
        ["原点规则", layout.get("meta", {}).get("origin_rule", "")],
        ["斜梁局部原点", "G 为斜梁 Part 左端局部起点，G_global 仅为派生全局参考点"],
        ["使用方式", "修改“关键尺寸输入” sheet 中黄色输入单元格及状态，其他 sheet 将通过公式自动更新。"],
        ["导出规则", "Step04 会读取输入行并重新计算 Assembly 输入。"],
    ]
    for idx, row in enumerate(info_rows, start=3):
        info_ws.cell(row=idx, column=1).value = row[0]
        info_ws.cell(row=idx, column=2).value = row[1]
    _style_header(info_ws, 3, 1, 1)
    _style_body(info_ws, 3, 9, 2, 2)
    _set_widths(info_ws, {"A": 26, "B": 110})

    figure_ws = wb.create_sheet("控制点示意图")
    _style_title(figure_ws, "A1:H1", "A/B/C/D/E/F 控制点示意图")
    figure_ws["A3"] = "说明：该图用于人工核对控制点命名、坐标轴方向和构件交点位置；计算仍以关键尺寸输入和公式结果为准。"
    figure_ws.merge_cells("A3:H3")
    _add_optional_image(figure_ws, figure_image, "A5")
    return wb


def _build_simple_workbook(layout: dict[str, Any], input_rows: list[dict[str, Any]], project_prefix: str, figure_image: str | Path | None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "坐标计算总表"
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    row_by_name = _write_input_sheet(ws, input_rows, "%s 坐标计算与校核总表" % project_prefix, simple=True)
    _set_widths(ws, {"A": 20, "B": 30, "C": 12, "D": 10, "E": 16, "F": 52, "G": 14, "H": 14, "I": 12, "J": 12})

    output_start = len(input_rows) + 6
    ws.cell(row=output_start, column=1).value = "关键输出参数"
    _style_header(ws, output_start, 1, 6)
    point_header_row = output_start + 1
    for col, header in enumerate(["点名", "X_m", "Y_m", "Z_m", "校核状态", "说明"], start=1):
        ws.cell(row=point_header_row, column=col).value = header
    _style_header(ws, point_header_row, 1, 6)
    theta = _input_ref(row_by_name, "theta_deg")
    xf = _mm(_input_ref(row_by_name, "X_F_mm"))
    zf = _mm(_input_ref(row_by_name, "Z_F_mm"))
    za = _mm(_input_ref(row_by_name, "Z_A_mm"))
    zbd = _mm(_input_ref(row_by_name, "Z_BD_mm"))
    r = _mm(_input_ref(row_by_name, "R_hoop_mm"))
    gc = _mm(_input_ref(row_by_name, "GC_mm"))
    gf = _mm(_input_ref(row_by_name, "GF_mm"))
    ge = _mm(_input_ref(row_by_name, "GE_mm"))
    rows = [
        ["A", "=0", "=0", "=%s" % za, _status_formula(row_by_name, "E", ["Z_A_mm"]), "上立柱上顶点"],
        ["B", "=-%s" % r, "=0", "=%s" % zbd, _status_formula(row_by_name, "E", ["Z_BD_mm", "R_hoop_mm"]), "前斜撑与抱箍交点"],
        ["C", "=%s+(%s-%s)*COS(RADIANS(%s))" % (xf, gc, gf, theta), "=0", "=%s+(%s-%s)*SIN(RADIANS(%s))" % (zf, gc, gf, theta), _status_formula(row_by_name, "E", ["X_F_mm", "Z_F_mm", "GC_mm", "GF_mm", "theta_deg"]), "斜梁与前斜撑交点"],
        ["D", "=%s" % r, "=0", "=%s" % zbd, _status_formula(row_by_name, "E", ["Z_BD_mm", "R_hoop_mm"]), "后斜撑与抱箍交点"],
        ["E", "=%s+(%s-%s)*COS(RADIANS(%s))" % (xf, ge, gf, theta), "=0", "=%s+(%s-%s)*SIN(RADIANS(%s))" % (zf, ge, gf, theta), _status_formula(row_by_name, "E", ["X_F_mm", "Z_F_mm", "GE_mm", "GF_mm", "theta_deg"]), "斜梁与后斜撑交点"],
        ["F", "=%s" % xf, "=0", "=%s" % zf, _status_formula(row_by_name, "E", ["X_F_mm", "Z_F_mm"]), "斜梁与上立柱/三角连接件参考交点"],
        ["G_global", "=%s-%s*COS(RADIANS(%s))" % (xf, gf, theta), "=0", "=%s-%s*SIN(RADIANS(%s))" % (zf, gf, theta), _status_formula(row_by_name, "E", ["X_F_mm", "Z_F_mm", "GF_mm", "theta_deg"]), "斜梁局部起点派生全局位置"],
    ]
    out_rows = {}
    for row_index, row in enumerate(rows, start=output_start + 2):
        out_rows[row[0]] = row_index
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index).value = value
    _style_body(ws, output_start + 2, output_start + 8, 1, 6)

    check_start = output_start + 11
    ws.cell(row=check_start, column=1).value = "校核结果"
    _style_header(ws, check_start, 1, 8)
    for col, header in enumerate(["校核项", "计算值", "图纸/输入值", "误差", "允许误差", "是否通过", "校核状态", "备注"], start=1):
        ws.cell(row=check_start + 1, column=col).value = header
    _style_header(ws, check_start + 1, 1, 8)
    check_rows = [
        ["GC_GF_GE_ORDER", '="GC="&TEXT(%s,"0.000")&", GF="&TEXT(%s,"0.000")&", GE="&TEXT(%s,"0.000")' % (gc, gf, ge), "GC < GF < GE", "", "", '=IF(AND(%s<%s,%s<%s),"%s","%s")' % (gc, gf, gf, ge, PASSED, FAILED), '=IF(AND(F%d="%s",$E$%d="%s",$E$%d="%s",$E$%d="%s"),"%s","%s")' % (check_start + 2, PASSED, row_by_name["GC_mm"], CONFIRMED, row_by_name["GF_mm"], CONFIRMED, row_by_name["GE_mm"], CONFIRMED, CONFIRMED, MANUAL_CHECK), "斜梁局部截面顺序"],
        ["CF_LOCAL", "=%s-%s" % (gf, gc), "GF-GC", "", "", '=IF(B%d>0,"%s","%s")' % (check_start + 3, PASSED, FAILED), '=IF(F%d="%s","%s","%s")' % (check_start + 3, PASSED, CONFIRMED, MANUAL_CHECK), "C-F 局部距离"],
        ["FE_LOCAL", "=%s-%s" % (ge, gf), "GE-GF", "", "", '=IF(B%d>0,"%s","%s")' % (check_start + 4, PASSED, FAILED), '=IF(F%d="%s","%s","%s")' % (check_start + 4, PASSED, CONFIRMED, MANUAL_CHECK), "F-E 局部距离"],
        ["CE_ANGLE", "=DEGREES(ATAN((D%d-D%d)/(B%d-B%d)))" % (out_rows["E"], out_rows["C"], out_rows["E"], out_rows["C"]), "=%s" % theta, "=MOD(B%d-C%d+180,360)-180" % (check_start + 5, check_start + 5), "=%s" % _input_ref(row_by_name, "angle_tolerance_deg"), '=IF(ABS(D%d)<=E%d,"%s","%s")' % (check_start + 5, check_start + 5, PASSED, FAILED), '=IF(F%d="%s","%s","%s")' % (check_start + 5, PASSED, CONFIRMED, MANUAL_CHECK), "CE 与 +X 夹角"],
        ["BC", "=SQRT((B%d-B%d)^2+(D%d-D%d)^2)" % (out_rows["C"], out_rows["B"], out_rows["C"], out_rows["B"]), "=%s" % _mm(_input_ref(row_by_name, "L_BC_draw_mm")), "=B%d-C%d" % (check_start + 6, check_start + 6), "=%s" % _input_ref(row_by_name, "control_tolerance_m"), '=IF(ABS(D%d)<=E%d,"%s","%s")' % (check_start + 6, check_start + 6, PASSED, FAILED), '=IF(F%d="%s","%s","%s")' % (check_start + 6, PASSED, CONFIRMED, MANUAL_CHECK), "前斜撑长度校核"],
        ["DE", "=SQRT((B%d-B%d)^2+(D%d-D%d)^2)" % (out_rows["E"], out_rows["D"], out_rows["E"], out_rows["D"]), "=%s" % _mm(_input_ref(row_by_name, "L_DE_draw_mm")), "=B%d-C%d" % (check_start + 7, check_start + 7), "=%s" % _input_ref(row_by_name, "control_tolerance_m"), '=IF(ABS(D%d)<=E%d,"%s","%s")' % (check_start + 7, check_start + 7, PASSED, FAILED), '=IF(F%d="%s","%s","%s")' % (check_start + 7, PASSED, CONFIRMED, MANUAL_CHECK), "后斜撑长度校核"],
    ]
    for row_index, row in enumerate(check_rows, start=check_start + 2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index).value = value
    _style_body(ws, check_start + 2, check_start + 7, 1, 8)

    figure_start = check_start + 10
    ws.cell(row=figure_start, column=1).value = "控制点示意图"
    _style_header(ws, figure_start, 1, 8)
    _add_optional_image(ws, figure_image, "A%d" % (figure_start + 2), width=720)
    ws.sheet_view.showGridLines = False
    return wb


def create_coordinate_formula_workbooks(
    layout_path: str | Path,
    full_output_path: str | Path,
    simple_output_path: str | Path,
    project_prefix: str,
    figure_image: str | Path | None = None,
    control_tolerance_m: float = 0.001,
    angle_tolerance_deg: float = 0.05,
) -> tuple[Path, Path]:
    layout = _copy_layout_with_prefix(load_coordinate_layout(layout_path), project_prefix)
    input_rows = _dimension_rows(layout, control_tolerance_m, angle_tolerance_deg)
    full_wb = _build_full_workbook(layout, input_rows, project_prefix, figure_image)
    simple_wb = _build_simple_workbook(layout, input_rows, project_prefix, figure_image)

    full_output = Path(full_output_path)
    simple_output = Path(simple_output_path)
    full_output.parent.mkdir(parents=True, exist_ok=True)
    simple_output.parent.mkdir(parents=True, exist_ok=True)
    full_wb.save(full_output)
    simple_wb.save(simple_output)
    return full_output, simple_output


def create_coordinate_template_workbook(
    output_path: str | Path,
    project_prefix: str = "PROJECT_PREFIX",
    figure_image: str | Path | None = None,
    control_tolerance_m: float = 0.001,
    angle_tolerance_deg: float = 0.05,
) -> Path:
    """Create the blank Step03 Excel template consumed by Step04."""

    with tempfile.TemporaryDirectory() as tmp:
        layout_json = Path(tmp) / "coordinate_layout_template.json"
        write_coordinate_layout_template(layout_json, project_prefix)
        layout = _copy_layout_with_prefix(load_coordinate_layout(layout_json), project_prefix)

    input_rows = _dimension_rows(layout, control_tolerance_m, angle_tolerance_deg)
    workbook = _build_simple_workbook(layout, input_rows, project_prefix, figure_image)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
