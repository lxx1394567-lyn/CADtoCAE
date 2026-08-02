from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .main_frame_assembly import export_main_frame_assembly
from .part_script import PROJECT_PREFIX_RE, project_prefix_from_path


DEBUG_REPORT_SUBDIR = Path("过程文件") / "调试文件"


@dataclass
class AssemblyScriptOutput:
    coordinate_workbook_path: str
    status: str
    project_prefix: str
    project_dir: str
    copied_coordinate_workbook_path: str | None
    copied_components_json_path: str | None
    assembly_json_path: str | None
    script_paths: list[str]
    report_path: str
    warning_count: int
    error_count: int
    messages: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _write_report(path: str | Path, payload: dict[str, Any]) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def step04_debug_report_path(output_root: str | Path, project_prefix_value: str) -> Path:
    return Path(output_root) / DEBUG_REPORT_SUBDIR / ("%s_step04_assembly_script_report.json" % project_prefix_value)


def _prefix_from_components_json(components_json: str | Path | None) -> str | None:
    if not components_json:
        return None
    path = Path(components_json)
    if not path.exists():
        return None
    text = path.read_text(encoding="utf-8", errors="ignore")
    match = PROJECT_PREFIX_RE.search(text)
    return match.group(0).upper() if match else None


def _prefix_from_workbook_text(xlsx: str | Path) -> str | None:
    workbook = load_workbook(xlsx, read_only=True, data_only=False)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), max_col=min(sheet.max_column, 6), values_only=True):
                for value in row:
                    if value is None:
                        continue
                    match = PROJECT_PREFIX_RE.search(str(value))
                    if match:
                        return match.group(0).upper()
    finally:
        workbook.close()
    return None


def infer_project_prefix_from_coordinate_workbook(xlsx: str | Path, components_json: str | Path | None = None) -> str:
    prefix = project_prefix_from_path(xlsx)
    if prefix:
        return prefix
    prefix = _prefix_from_workbook_text(xlsx)
    if prefix:
        return prefix
    prefix = _prefix_from_components_json(components_json)
    if prefix:
        return prefix
    raise ValueError("无法从坐标表或 Step02 组件数据推断项目名称前缀，请将坐标表命名为 <project_prefix>_coordinate_formula_simple_fixed.xlsx。")


def locate_components_source(coordinate_workbook: str | Path, project_prefix: str | None = None) -> Path | None:
    workbook = Path(coordinate_workbook)
    prefix = project_prefix or project_prefix_from_path(workbook) or _prefix_from_workbook_text(workbook)
    candidates: list[Path] = []
    for parent in [workbook.parent, *list(workbook.parents)[:4]]:
        if prefix:
            candidates.append(parent / ("%s_create_parts_in_cae.py" % prefix))
            candidates.append(parent / ("%s_components.json" % prefix))
        candidates.extend(
            [
                parent / "components.json",
                parent / "json" / "components.json",
            ]
        )
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists():
            return candidate
    return None


def locate_components_json(coordinate_workbook: str | Path, project_prefix: str | None = None) -> Path | None:
    return locate_components_source(coordinate_workbook, project_prefix)


def generate_assembly_scripts_from_workbook(
    coordinate_workbook: str | Path,
    output_root: str | Path,
    components_json: str | Path | None = None,
    project_prefix_value: str | None = None,
    model_name: str | None = None,
    overwrite: bool = False,
) -> AssemblyScriptOutput:
    workbook = Path(coordinate_workbook)
    messages: list[str] = []
    output_dir = Path(output_root)
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        prefix = project_prefix_value or infer_project_prefix_from_coordinate_workbook(workbook, components_json)
        components = Path(components_json) if components_json else locate_components_source(workbook, prefix)
        if not components or not components.exists():
            raise FileNotFoundError("Step02 component data not found. Put <project_prefix>_create_parts_in_cae.py or <project_prefix>_components.json in the coordinate workbook folder.")

        components_prefix = _prefix_from_components_json(components)
        if components_prefix and components_prefix != prefix:
            raise ValueError("Step02 component data prefix %s does not match coordinate workbook prefix %s." % (components_prefix, prefix))

        if model_name and model_name != prefix:
            raise ValueError("Step04 model name must match project prefix %s." % prefix)

        actual_model_name = prefix
        scripts_dir = output_dir
        reports_dir = output_dir / DEBUG_REPORT_SUBDIR

        json_file, scripts, payload = export_main_frame_assembly(
            workbook,
            components,
            None,
            scripts_dir,
            reports_dir=reports_dir,
            project_code=prefix,
            model_name=actual_model_name,
        )

        warnings = list(payload.get("warnings", []))
        errors = list(payload.get("errors", []))
        status = "ok" if not errors else "needs_review"
        messages.append("Generated %s Abaqus assembly script(s) with embedded data. Abaqus model name: %s." % (len(scripts), actual_model_name))
        if warnings:
            messages.append("%s warnings found; please review the coordinate workbook." % len(warnings))
        if errors:
            messages.append("%s errors found; fix the coordinate workbook and generate again." % len(errors))

        report = _write_report(
            step04_debug_report_path(output_dir, prefix),
            {
                "status": status,
                "project_prefix": prefix,
                "source_coordinate_workbook": str(workbook.resolve()),
                "source_components": str(components.resolve()),
                "assembly_json": None,
                "data_mode": "embedded_in_py",
                "abaqus_scripts": [str(path.resolve()) for path in scripts],
                "model_name": actual_model_name,
                "warnings": warnings,
                "errors": errors,
                "messages": messages,
            },
        )
        return AssemblyScriptOutput(
            coordinate_workbook_path=str(workbook.resolve()),
            status=status,
            project_prefix=prefix,
            project_dir=str(output_dir.resolve()),
            copied_coordinate_workbook_path=str(workbook.resolve()),
            copied_components_json_path=str(components.resolve()),
            assembly_json_path=None,
            script_paths=[str(path.resolve()) for path in scripts],
            report_path=str(report.resolve()),
            warning_count=len(warnings),
            error_count=len(errors),
            messages=messages,
        )
    except Exception as exc:
        prefix = project_prefix_value or project_prefix_from_path(workbook) or "UNKNOWN_PROJECT"
        report = _write_report(
            step04_debug_report_path(output_dir, prefix),
            {
                "status": "failed",
                "source_coordinate_workbook": str(workbook.resolve()),
                "components_json": str(Path(components_json).resolve()) if components_json else None,
                "messages": [str(exc)],
            },
        )
        return AssemblyScriptOutput(
            coordinate_workbook_path=str(workbook.resolve()),
            status="failed",
            project_prefix=prefix,
            project_dir=str(output_dir.resolve()),
            copied_coordinate_workbook_path=None,
            copied_components_json_path=None,
            assembly_json_path=None,
            script_paths=[],
            report_path=str(report.resolve()),
            warning_count=0,
            error_count=1,
            messages=[str(exc)],
        )


def batch_generate_assembly_scripts(
    coordinate_workbooks: Iterable[str | Path],
    output_root: str | Path,
    components_json: str | Path | None = None,
    model_name: str | None = None,
    overwrite: bool = False,
) -> list[AssemblyScriptOutput]:
    root = Path(output_root)
    root.mkdir(parents=True, exist_ok=True)
    outputs: list[AssemblyScriptOutput] = []
    for workbook in coordinate_workbooks:
        outputs.append(
            generate_assembly_scripts_from_workbook(
                workbook,
                root,
                components_json=components_json,
                model_name=model_name,
                overwrite=overwrite,
            )
        )
    return outputs
