from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .standards import (
    ParsedSpec,
    component_code_from_part_name,
    derive_component_row,
    effective_model_policy,
    has_complete_model_dimensions,
    load_standards,
    material_properties,
    mm_to_m,
    parse_spec,
    section_kind_and_model_params,
)


RAW_HEADERS = ["类别", "序号", "名称", "规格", "长度_mm", "数量", "备注", "来源页码", "识别置信度"]
COMPONENT_HEADERS = [
    "支架类型",
    "角度",
    "阵列布置",
    "构件名称",
    "abaqus_part_name",
    "规格",
    "长度_mm",
    "长度_m",
    "数量",
    "材料牌号",
    "建模方式",
    "单元类型",
    "截面类型",
    "截面参数",
    "厚度_mm",
    "厚度_m",
]
STEP02_GUIDE_HEADERS = {
    "支架类型": "用于推断项目名前缀。文件名已有 SP_SC_ANG20 这类前缀时，以文件名前缀优先。",
    "角度": "用于推断项目名前缀，例如 20 或 26.5。",
    "构件名称": "用于人工核对和构件角色识别。",
    "abaqus_part_name": "Abaqus Part 名称，是 Step02 最关键的识别字段，格式建议为 P_<项目名前缀>_<构件英文名>。",
    "规格": "Step02 会从本列重新解析截面。支持 C80x40x10x2.0、L90x56x5.0、Φ159x3.0、Φ180x70x5.0、Φ10、M8 等格式。",
    "长度_mm": "C 型钢、圆管、角钢、撑杆、圆杆等线性构件必须填写。",
    "数量": "Step02 会写入 JSON，后续 assembly 会使用。",
    "材料牌号": "必须填写，建议使用 Q235 B、Q355 B、Q420 B、Q550 B、6063-T5。",
    "建模方式": "要自动生成 Part，应填写 壳单元 或 实体单元。人工模板不会自动建 Part。",
    "单元类型": "壳单元通常为 S4R，实体单元通常为 C3D8R。",
}
STEP02_REQUIRED_HEADERS = list(STEP02_GUIDE_HEADERS)
FORMULA_LINKED_COMPONENT_HEADERS = {
    "构件名称",
    "abaqus_part_name",
    "规格",
    "长度_mm",
    "长度_m",
    "数量",
    "材料牌号",
    "建模方式",
    "单元类型",
}


def _excel_string(value: Any) -> str:
    return '"%s"' % str(value).replace('"', '""')


def _sheet_cell_ref(sheet_name: str, column: str, row_index: int) -> str:
    safe_sheet = sheet_name.replace("'", "''")
    return "'%s'!%s%s" % (safe_sheet, column, row_index)


def _same_row_link_formula(sheet_name: str, column: str, row_index: int) -> str:
    source = _sheet_cell_ref(sheet_name, column, row_index)
    return '=IF(TRIM(%s&"")="","",%s)' % (source, source)


def _nested_if_equals(text_expr: str, pairs: Iterable[tuple[str, str]], default_expr: str) -> str:
    result = default_expr
    seen: set[str] = set()
    normalized_pairs: list[tuple[str, str]] = []
    for key, value_expr in pairs:
        if key in seen:
            continue
        seen.add(key)
        normalized_pairs.append((key, value_expr))
    for key, value_expr in reversed(normalized_pairs):
        result = "IF(%s=%s,%s,%s)" % (text_expr, _excel_string(key), value_expr, result)
    return result


def _support_type_code_expr(cell_ref: str, standards: dict[str, Any]) -> str:
    text_expr = 'UPPER(TRIM(%s&""))' % cell_ref
    pairs: list[tuple[str, str]] = []
    for canonical, item in standards["support_types"].items():
        code = str(item["code"])
        candidates = [canonical, code, *(item.get("aliases") or [])]
        pairs.extend((str(candidate).strip().upper(), _excel_string(code)) for candidate in candidates)
    default_expr = 'UPPER(SUBSTITUTE(TRIM(%s&"")," ","_"))' % cell_ref
    return _nested_if_equals(text_expr, pairs, default_expr)


def _angle_code_expr(cell_ref: str) -> str:
    text_expr = (
        'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE('
        'UPPER(TRIM(%s&""))," ",""),"°",""),"度",""),"DEG","")'
    ) % cell_ref
    with_prefix = 'IF(LEFT(%s,3)="ANG",%s,"ANG"&%s)' % (text_expr, text_expr, text_expr)
    return 'SUBSTITUTE(%s,".","P")' % with_prefix


def _component_role_expr(
    component_name_ref: str,
    standards: dict[str, Any],
    role_key: str,
    default_value: str,
) -> str:
    text_expr = 'TRIM(%s&"")' % component_name_ref
    pairs = [
        (component_name, _excel_string(role.get(role_key, default_value)))
        for component_name, role in standards["component_roles"].items()
    ]
    return _nested_if_equals(text_expr, pairs, _excel_string(default_value))


def _model_policy_expr(component_name_ref: str, standards: dict[str, Any]) -> str:
    labels = standards["model_policy_labels"]
    text_expr = 'TRIM(%s&"")' % component_name_ref
    pairs = [
        (component_name, _excel_string(labels.get(str(role.get("model_policy", "")), role.get("model_policy", ""))))
        for component_name, role in standards["component_roles"].items()
    ]
    default_label = labels.get("MANUAL_TEMPLATE", "人工模板")
    return _nested_if_equals(text_expr, pairs, _excel_string(default_label))


def _material_grade_formula(sheet_name: str, column: str, row_index: int) -> str:
    source = _sheet_cell_ref(sheet_name, column, row_index)
    raw_expr = 'TRIM(%s&"")' % source
    normalized_expr = 'UPPER(SUBSTITUTE(%s," ",""))' % raw_expr
    pairs = [
        ("Q235B", _excel_string("Q235 B")),
        ("Q355B", _excel_string("Q355 B")),
        ("Q420B", _excel_string("Q420 B")),
        ("Q550B", _excel_string("Q550 B")),
        ("6063T5", _excel_string("6063-T5")),
        ("6063-T5", _excel_string("6063-T5")),
    ]
    mapped_expr = _nested_if_equals(normalized_expr, pairs, raw_expr)
    return '=IF(%s="","",%s)' % (raw_expr, mapped_expr)


def _part_name_formula(row_index: int, standards: dict[str, Any]) -> str:
    support_ref = "A%s" % row_index
    angle_ref = "B%s" % row_index
    component_ref = "D%s" % row_index
    support_code = _support_type_code_expr(support_ref, standards)
    angle_code = _angle_code_expr(angle_ref)
    component_code = _component_role_expr(component_ref, standards, "code", "UNKNOWN_COMPONENT")
    return (
        '=IF(OR(TRIM(%s&"")="",TRIM(%s&"")="",TRIM(%s&"")=""),"",'
        '"P_"&%s&"_"&%s&"_"&%s)'
    ) % (support_ref, angle_ref, component_ref, support_code, angle_code, component_code)


def _component_workbook_row(
    component_row: dict[str, Any],
    raw_sheet_name: str,
    row_index: int,
    standards: dict[str, Any],
) -> list[Any]:
    formulas = {
        "构件名称": _same_row_link_formula(raw_sheet_name, "C", row_index),
        "abaqus_part_name": _part_name_formula(row_index, standards),
        "规格": _same_row_link_formula(raw_sheet_name, "D", row_index),
        "长度_mm": _same_row_link_formula(raw_sheet_name, "E", row_index),
        "长度_m": '=IFERROR(IF(TRIM(G%s&"")="","",VALUE(G%s)/1000),"")' % (row_index, row_index),
        "数量": _same_row_link_formula(raw_sheet_name, "F", row_index),
        "材料牌号": _material_grade_formula(raw_sheet_name, "G", row_index),
        "建模方式": "=%s" % _model_policy_expr("D%s" % row_index, standards),
        "单元类型": "=%s" % _component_role_expr("D%s" % row_index, standards, "element_type", "C3D8R"),
    }
    return [formulas.get(header, component_row.get(header, "")) for header in COMPONENT_HEADERS]


def read_raw_material_csv(path: str | Path) -> list[dict[str, Any]]:
    csv_path = Path(path)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    return normalize_raw_rows(rows)


def normalize_raw_rows(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_rows: list[dict[str, Any]] = []
    last_remark = ""
    for row in rows:
        normalized = {header: row.get(header, "") for header in RAW_HEADERS}
        remark = str(normalized.get("备注", "")).strip()
        if remark:
            last_remark = remark
        elif last_remark:
            normalized["备注"] = last_remark
        normalized_rows.append(normalized)
    return normalized_rows


def build_component_rows(
    raw_rows: list[dict[str, Any]],
    support_type: str,
    angle: str,
    array_layout: str,
    standards_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    standards = load_standards(standards_path)
    return [
        derive_component_row(row, support_type, angle, array_layout, standards)
        for row in raw_rows
    ]


def _append_table(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_row = max(ws.max_row, 1)
    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}{last_row}"
    if last_row >= 2:
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ref


def _style_sheet(ws, widths: dict[str, float] | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    body_font = Font(name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.sheet_view.showGridLines = False

    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    else:
        for index in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(index)].width = 14


def _add_review_validations(ws) -> None:
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    policy_col = header_to_col.get("建模方式")
    element_col = header_to_col.get("单元类型")
    if policy_col:
        col_letter = get_column_letter(policy_col)
        dv = DataValidation(type="list", formula1='"壳单元,实体单元,连接器简化,人工模板"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}500")
    if element_col:
        col_letter = get_column_letter(element_col)
        dv = DataValidation(type="list", formula1='"S4R,C3D8R"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}500")


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _add_issue(issue_map: dict[str, list[str]], header: str, message: str) -> None:
    messages = issue_map.setdefault(header, [])
    if message not in messages:
        messages.append(message)


def _step02_issue_cells(component_row: dict[str, Any]) -> dict[str, list[str]]:
    issue_map: dict[str, list[str]] = {}
    _complete, issues = has_complete_model_dimensions(component_row)
    for issue in issues:
        if "abaqus_part_name" in issue:
            _add_issue(issue_map, "abaqus_part_name", issue)
        elif any(token in issue for token in ["规格", "高度", "翼缘", "卷边", "厚度", "外径", "直径", "边长", "公称"]):
            _add_issue(issue_map, "规格", issue)
        elif "长度" in issue or "弯折" in issue:
            _add_issue(issue_map, "长度_mm", issue)
        elif "数量" in issue:
            _add_issue(issue_map, "数量", issue)
        elif "材料牌号" in issue:
            _add_issue(issue_map, "材料牌号", issue)
        elif "建模方式" in issue or "人工模板" in issue or "连接器" in issue:
            _add_issue(issue_map, "建模方式", issue)
        elif "构件名称" in issue:
            _add_issue(issue_map, "构件名称", issue)
        elif "构件代码" in issue or "Part名称" in issue:
            _add_issue(issue_map, "abaqus_part_name", issue)
        else:
            _add_issue(issue_map, "abaqus_part_name", issue)

    policy = str(component_row.get("建模方式", "")).strip()
    if policy in {"壳单元", "实体单元", "SHELL", "SOLID"} and _is_blank(component_row.get("单元类型")):
        _add_issue(issue_map, "单元类型", "单元类型缺失")

    return issue_map


def _apply_step02_guidance(ws, component_rows: list[dict[str, Any]]) -> None:
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    header_fill = PatternFill("solid", fgColor="FFC00000")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    issue_fill = PatternFill("solid", fgColor="FFFFC7CE")
    issue_font = Font(name="Microsoft YaHei", size=10, color="FF9C0006")

    for header, note in STEP02_GUIDE_HEADERS.items():
        column = header_to_col.get(header)
        if not column:
            continue
        cell = ws.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.comment = Comment(note, "CADtoCAE")

    for row_index, component_row in enumerate(component_rows, start=2):
        issue_map = _step02_issue_cells(component_row)
        for header, header_messages in issue_map.items():
            column = header_to_col.get(header)
            if not column:
                continue
            cell = ws.cell(row=row_index, column=column)
            cell.fill = issue_fill
            cell.font = issue_font
            cell.comment = Comment("；".join(header_messages), "CADtoCAE")


def create_material_workbook(
    raw_rows: list[dict[str, Any]],
    support_type: str,
    angle: str,
    array_layout: str,
    output_path: str | Path,
    standards_path: str | Path | None = None,
) -> Path:
    raw_rows = normalize_raw_rows(raw_rows)
    standards = load_standards(standards_path)
    component_rows = [
        derive_component_row(row, support_type, angle, array_layout, standards)
        for row in raw_rows
    ]

    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    default_ws = wb.active
    wb.remove(default_ws)

    raw_ws = wb.create_sheet("原始材料表")
    _append_table(
        raw_ws,
        RAW_HEADERS,
        [[row.get(header, "") for header in RAW_HEADERS] for row in raw_rows],
        "RawMaterialTable",
    )
    _style_sheet(
        raw_ws,
        {
            "A": 10,
            "B": 8,
            "C": 16,
            "D": 22,
            "E": 12,
            "F": 10,
            "G": 14,
            "H": 12,
            "I": 14,
        },
    )

    component_ws = wb.create_sheet("建模构件表")
    _append_table(
        component_ws,
        COMPONENT_HEADERS,
        [
            _component_workbook_row(row, raw_ws.title, row_index, standards)
            for row_index, row in enumerate(component_rows, start=2)
        ],
        "ComponentModelTable",
    )
    _style_sheet(
        component_ws,
        {
            "A": 16,
            "B": 10,
            "C": 14,
            "D": 16,
            "E": 34,
            "F": 22,
            "G": 12,
            "H": 12,
            "I": 10,
            "J": 12,
            "K": 12,
            "L": 10,
            "M": 14,
            "N": 38,
            "O": 10,
            "P": 10,
        },
    )
    _add_review_validations(component_ws)
    _apply_step02_guidance(component_ws, component_rows)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def _is_formula_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _coerce_number(value: Any) -> float | int | None:
    if _is_blank(value):
        return None
    text = str(value).strip().replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _parse_section_params_text(value: Any) -> dict[str, float | int | str]:
    if _is_blank(value):
        return {}
    params: dict[str, float | int | str] = {}
    for item in re.split(r"[;；]", str(value)):
        if "=" not in item:
            continue
        key, raw_value = item.split("=", 1)
        key = key.strip()
        raw_value = raw_value.strip()
        if not key:
            continue
        numeric_value = _coerce_number(raw_value)
        params[key] = numeric_value if numeric_value is not None else raw_value
    return params


def _section_code_from_table(section_type: str, params: dict[str, Any], fallback: str) -> str:
    if fallback and fallback not in {"UNKNOWN", "UNSPEC"}:
        return fallback
    safe_type = re.sub(r"[^A-Za-z0-9]+", "_", section_type.upper()).strip("_") or "TABLE"
    safe_params = "_".join(
        "%s%s" % (re.sub(r"[^A-Za-z0-9]+", "", str(key)), str(value).replace(".", "P"))
        for key, value in sorted(params.items())
    )
    return ("TABLE_%s_%s" % (safe_type, safe_params)).rstrip("_")


def _parsed_spec_for_export(row: dict[str, Any]) -> ParsedSpec:
    parsed = parse_spec(row.get("规格", ""))
    if parsed.status == "已解析":
        return parsed

    section_type = str(row.get("截面类型", "") or "").strip()
    params = _parse_section_params_text(row.get("截面参数"))
    thickness_mm = _coerce_number(row.get("厚度_mm"))
    if thickness_mm is not None and "厚度_mm" not in params:
        params["厚度_mm"] = thickness_mm
    if not section_type or not params:
        return parsed

    return ParsedSpec(
        section_type=section_type,
        section_params=params,
        thickness_mm=float(thickness_mm) if thickness_mm is not None else parsed.thickness_mm,
        section_code=_section_code_from_table(section_type, params, parsed.section_code),
        status="已解析",
    )


def _worksheet_rows_by_header(wb: Any, sheet_name: str) -> tuple[list[dict[str, Any]], list[str]]:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append({header: value for header, value in zip(headers, values)})
    return rows, headers


def _sheet_rows_by_header(path: str | Path, sheet_name: str, data_only: bool = True) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=data_only, read_only=True)
    try:
        rows, _headers = _worksheet_rows_by_header(wb, sheet_name)
        return rows
    finally:
        wb.close()


def _derive_component_row_for_formula_fallback(
    raw_row: dict[str, Any],
    support_type: Any,
    angle: Any,
    array_layout: Any,
    standards: dict[str, Any],
) -> dict[str, Any]:
    return derive_component_row(
        raw_row,
        str(support_type or ""),
        angle if angle is not None else "",
        str(array_layout or ""),
        standards,
    )


def read_component_rows_for_processing(
    path: str | Path,
    standards_path: str | Path | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    value_wb = load_workbook(path, data_only=True, read_only=True)
    formula_wb = load_workbook(path, data_only=False, read_only=True)
    try:
        value_rows, headers = _worksheet_rows_by_header(value_wb, "建模构件表")
        formula_rows, _formula_headers = _worksheet_rows_by_header(formula_wb, "建模构件表")
        if "原始材料表" not in value_wb.sheetnames:
            return value_rows, headers

        has_link_formulas = any(
            _is_formula_value(row.get(header))
            for row in formula_rows
            for header in FORMULA_LINKED_COMPONENT_HEADERS
        )
        if not has_link_formulas:
            return value_rows, headers

        raw_rows, _raw_headers = _worksheet_rows_by_header(value_wb, "原始材料表")
        raw_rows = normalize_raw_rows(raw_rows)
        standards = load_standards(standards_path)
        merged_rows: list[dict[str, Any]] = []
        max_rows = max(len(value_rows), len(formula_rows))
        for index in range(max_rows):
            value_row = value_rows[index] if index < len(value_rows) else {}
            formula_row = formula_rows[index] if index < len(formula_rows) else {}
            raw_row = raw_rows[index] if index < len(raw_rows) else {}
            support_type = value_row.get("支架类型") or formula_row.get("支架类型") or ""
            angle = value_row.get("角度") or formula_row.get("角度") or ""
            array_layout = value_row.get("阵列布置") or formula_row.get("阵列布置") or ""
            derived_row = {}
            if raw_row:
                try:
                    derived_row = _derive_component_row_for_formula_fallback(
                        raw_row,
                        support_type,
                        angle,
                        array_layout,
                        standards,
                    )
                except Exception:
                    derived_row = {
                        "构件名称": raw_row.get("名称", ""),
                        "规格": raw_row.get("规格", ""),
                        "长度_mm": raw_row.get("长度_mm", ""),
                        "数量": raw_row.get("数量", ""),
                        "材料牌号": raw_row.get("备注", ""),
                    }

            merged_row: dict[str, Any] = {}
            for header in headers:
                if header in FORMULA_LINKED_COMPONENT_HEADERS and _is_formula_value(formula_row.get(header)):
                    merged_row[header] = derived_row.get(header, value_row.get(header))
                else:
                    merged_row[header] = value_row.get(header)
            merged_rows.append(merged_row)
        return merged_rows, headers
    finally:
        value_wb.close()
        formula_wb.close()


def load_approved_components(path: str | Path, standards_path: str | Path | None = None) -> list[dict[str, Any]]:
    rows, _headers = read_component_rows_for_processing(path, standards_path)
    return [
        row
        for row in rows
        if str(row.get("校核状态", "")).strip().lower() in {"已确认", "approved"}
    ]


def load_dimension_complete_components(path: str | Path, standards_path: str | Path | None = None) -> list[dict[str, Any]]:
    rows, _headers = read_component_rows_for_processing(path, standards_path)
    complete_rows = []
    for row in rows:
        complete, _issues = has_complete_model_dimensions(row)
        if complete:
            complete_rows.append(row)
    return complete_rows


def export_abaqus_json(
    workbook_path: str | Path,
    output_path: str | Path,
    standards_path: str | Path | None = None,
    selection: str = "approved",
) -> Path:
    standards = load_standards(standards_path)
    components = []
    if selection == "approved":
        selected_rows = load_approved_components(workbook_path, standards_path)
    elif selection == "complete":
        selected_rows = load_dimension_complete_components(workbook_path, standards_path)
    else:
        raise ValueError(f"未知导出选择模式: {selection}")

    for row in selected_rows:
        parsed = _parsed_spec_for_export(row)
        section_kind, section_params_m = section_kind_and_model_params(parsed)
        material = material_properties(str(row.get("材料牌号", "") or ""), standards)
        model_policy, element_type = effective_model_policy(row)
        length_m = mm_to_m(row.get("长度_mm"))
        thickness_m = mm_to_m(parsed.thickness_mm)
        component_code = row.get("构件代码") or component_code_from_part_name(row.get("abaqus_part_name"))
        components.append(
            {
                "part_name": row.get("abaqus_part_name"),
                "support_type": row.get("支架类型"),
                "angle": row.get("角度"),
                "component_name": row.get("构件名称"),
                "component_code": component_code,
                "spec": row.get("规格"),
                "length_m": length_m,
                "quantity": row.get("数量"),
                "material": material,
                "model_policy": model_policy,
                "element_type": element_type,
                "section_kind": section_kind,
                "section_params_m": section_params_m,
                "thickness_m": thickness_m,
                "section_code": parsed.section_code,
                "model_units": {"length": "m", "mass": "kg", "force": "N", "stress": "Pa"},
            }
        )
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as handle:
        json.dump({"components": components}, handle, ensure_ascii=False, indent=2)
    return output
