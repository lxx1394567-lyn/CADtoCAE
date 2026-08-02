from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

from cadtocae.runs import create_run_paths, update_manifest
from cadtocae.standards import project_prefix
from cadtocae.workbook import export_abaqus_json
from make_cae_runner import generate_cae_runner


def _prefix_from_latest(outputs_root: str | Path) -> str | None:
    latest = Path(outputs_root) / "latest_run_manifest.json"
    if not latest.exists():
        return None
    payload = json.loads(latest.read_text(encoding="utf-8"))
    return payload.get("project_prefix") or payload.get("project_code")


def _prefix_from_workbook(xlsx: str | Path, standards: str | Path) -> str | None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    if "建模构件表" not in wb.sheetnames:
        return None
    ws = wb["建模构件表"]
    headers = [cell.value for cell in ws[1]]
    try:
        support_col = headers.index("支架类型") + 1
        angle_col = headers.index("角度") + 1
    except ValueError:
        return None
    for row in range(2, ws.max_row + 1):
        support_type = ws.cell(row=row, column=support_col).value
        angle = ws.cell(row=row, column=angle_col).value
        if support_type and angle:
            return project_prefix(str(support_type), angle, standards)
    return None


def _default_excel(run_dir: Path, project_prefix_value: str) -> Path:
    for name in ("%s_components.xlsx" % project_prefix_value, "%s_components_v2.xlsx" % project_prefix_value):
        run_excel = run_dir / "workbooks" / name
        if run_excel.exists():
            return run_excel
    return run_dir / "workbooks" / ("%s_components.xlsx" % project_prefix_value)


def _copy_input_workbook(source: Path, run_workbooks_dir: Path) -> Path:
    target = run_workbooks_dir / source.name
    if source.resolve() != target.resolve():
        shutil.copy2(source, target)
    return target


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Step02: generate Abaqus Part creation script from the Part modeling Excel.")
    parser.add_argument("--xlsx", default=None, help="Part modeling Excel. Defaults to run_dir/workbooks/<project_prefix>_components.xlsx.")
    parser.add_argument("--selection", choices=["approved", "complete"], default="complete")
    parser.add_argument("--project-prefix", "--project-code", dest="project_prefix", default=None)
    parser.add_argument("--standards", default=str(PROJECT_ROOT / "config" / "standards.json"))
    parser.add_argument("--outputs-root", default="outputs")
    parser.add_argument("--run-dir", default="latest", help="Run directory. Defaults to latest.")
    parser.add_argument("--components-json", default=None, help="Optional legacy JSON export path. By default component data is embedded in the generated Part script.")
    parser.add_argument("--part-script", default=None, help="Optional Abaqus script path. Defaults to run_dir/abaqus_scripts/<project>_create_parts_in_cae.py.")
    parser.add_argument("--model-name", default=None, help="Abaqus model name. Defaults to <project_prefix>.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    prefix = args.project_prefix or _prefix_from_latest(args.outputs_root)
    if not prefix and args.xlsx:
        prefix = _prefix_from_workbook(args.xlsx, args.standards)
    if not prefix:
        raise ValueError("Cannot determine project prefix. Pass --project-prefix or run Step01 first.")

    run_dir = args.run_dir
    if args.xlsx and str(args.run_dir).strip().lower() == "latest":
        latest_manifest = Path(args.outputs_root) / "latest_run_manifest.json"
        if not latest_manifest.exists():
            run_dir = None

    run_paths = create_run_paths(prefix, args.outputs_root, run_dir)
    xlsx = Path(args.xlsx) if args.xlsx else _default_excel(run_paths.root, prefix)
    if not xlsx.exists():
        raise FileNotFoundError("Part modeling Excel not found: %s" % xlsx)

    copied_xlsx = _copy_input_workbook(xlsx, run_paths.workbooks)
    part_script = Path(args.part_script) if args.part_script else run_paths.abaqus_scripts / ("%s_create_parts_in_cae.py" % prefix)
    model_name = args.model_name or prefix

    with tempfile.TemporaryDirectory() as tmp:
        components_json = Path(args.components_json) if args.components_json else Path(tmp) / ("%s_components.json" % prefix)
        exported_json = export_abaqus_json(copied_xlsx, components_json, args.standards, selection=args.selection)
        script = generate_cae_runner(exported_json, part_script, model_name=model_name, save_as_path=None)
        exported_json_path = Path(exported_json).resolve() if args.components_json else None

    outputs = {
        "workbooks": {"components": str(copied_xlsx.resolve())},
        "abaqus_scripts": {"create_parts": str(Path(script).resolve())},
        "metadata": {"components_data": "embedded_in_part_script"},
    }
    if exported_json_path:
        outputs["json"] = {"components": str(exported_json_path)}

    update_manifest(
        run_paths,
        prefix,
        "step02_part_script",
        inputs={
            "components_excel": str(copied_xlsx.resolve()),
            "selection": args.selection,
            "model_name": model_name,
            "standards": str(Path(args.standards).resolve()),
        },
        outputs=outputs,
    )

    print("Run directory: %s" % run_paths.root.resolve())
    if exported_json_path:
        print("Wrote components JSON: %s" % exported_json_path)
    else:
        print("Embedded components data in Part Abaqus script.")
    print("Wrote Part Abaqus script: %s" % Path(script).resolve())
    print("Manifest: %s" % run_paths.manifest.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
